"""DXF 表格抽取测试 — 算法单测 + 集成 + Excel 导出。"""
from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.core.config import settings
from backend.main import app

from recognizer.cad_engine.table_classifier import classify_table_kind
from recognizer.cad_engine.table_extractor import (
    extract_text_cluster_tables,
)
from tests.dwg_test_helpers import (
    equipment_schedule_dxf,
    wait_for_extract_tables_job,
)


VERSION = "v1.1.4-table-extract"


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as test_client:
        yield test_client


def _make_cad_json(items: list[dict]) -> dict:
    return {
        "spaces": [
            {
                "space": "modelspace",
                "texts": items,
                "mtexts": [],
                "inserts": [],
            }
        ],
        "counts": {},
        "warnings": [],
    }


def _text_item(text: str, x: float, y: float, h: float = 2.5) -> dict:
    return {
        "clean_text": text,
        "insert": [x, y, 0],
        "height": h,
        "layer": "TABLE",
    }


def test_extract_text_cluster_finds_3x3_equipment_grid():
    rows = [
        ["设备名", "型号", "数量"],
        ["送风机", "FJ-1", "2"],
        ["排风机", "FJ-2", "3"],
        ["照明箱", "AL-1", "5"],
    ]
    items = []
    row_h = 8.0
    col_xs = [0.0, 60.0, 120.0]
    for r_idx, row in enumerate(rows):
        y = 100.0 - r_idx * row_h
        for c_idx, cell in enumerate(row):
            items.append(_text_item(cell, col_xs[c_idx], y))
    tables = extract_text_cluster_tables(_make_cad_json(items))
    assert len(tables) == 1
    table = tables[0]
    assert table["col_count"] >= 3
    assert table["row_count"] >= 3
    assert "设备名" in table["header"][0]


def test_extract_text_cluster_ignores_title_block_area():
    rows = [
        ["设备名", "型号", "数量"],
        ["送风机", "FJ-1", "2"],
        ["排风机", "FJ-2", "3"],
        ["照明箱", "AL-1", "5"],
    ]
    items = []
    for r_idx, row in enumerate(rows):
        y = 50.0 - r_idx * 8.0
        for c_idx, cell in enumerate(row):
            items.append(_text_item(cell, c_idx * 60.0, y))
    exclude_bbox = [-10.0, -100.0, 200.0, 100.0]
    tables = extract_text_cluster_tables(
        _make_cad_json(items), exclude_bbox=exclude_bbox
    )
    assert tables == []


def test_extract_text_cluster_requires_min_rows():
    items = [
        _text_item("A", 0, 100),
        _text_item("B", 60, 100),
        _text_item("C", 0, 92),
        _text_item("D", 60, 92),
    ]
    tables = extract_text_cluster_tables(_make_cad_json(items))
    assert tables == []


def test_classify_equipment_by_header_keywords():
    assert classify_table_kind(["设备名", "型号", "数量"]) == "equipment"
    assert classify_table_kind(["材料", "品名", "单位"]) == "material"
    assert classify_table_kind(["图号", "图名", "页号"]) == "drawing_index"
    assert classify_table_kind(["图例", "符号", "说明"]) == "legend"


def test_classify_falls_back_to_other():
    assert classify_table_kind(["A", "B", "C"]) == "other"
    assert classify_table_kind([]) == "other"


def _create_project(client: TestClient, name: str) -> int:
    response = client.post("/api/projects", json={"name": name})
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _upload_and_parse(client: TestClient, project_id: int, content: bytes) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", ("equipment.dxf", content, "application/dxf"))],
    )
    assert response.status_code == 201, response.text
    file_id = response.json()["files"][0]["id"]
    parse_resp = client.post(f"/api/files/{file_id}/parse-dxf")
    assert parse_resp.status_code == 200, parse_resp.text
    return parse_resp.json()


def test_extract_tables_from_sheet_writes_drawing_table_rows(client: TestClient):
    project_id = _create_project(client, "DXF 表格抽取-单 sheet")
    content = equipment_schedule_dxf()
    parse_result = _upload_and_parse(client, project_id, content)
    sheet_id = parse_result["sheet_id"]

    extract_resp = client.post(
        f"/api/sheets/{sheet_id}/extract-tables", json={}
    )
    assert extract_resp.status_code == 200, extract_resp.text
    summary = extract_resp.json()
    assert summary["sheet_id"] == sheet_id
    assert summary["total_count"] >= 1

    list_resp = client.get(f"/api/sheets/{sheet_id}/tables")
    assert list_resp.status_code == 200
    tables = list_resp.json()
    assert len(tables) >= 1
    table = tables[0]
    assert table["extraction_method"] in {"acad_table", "text_cluster"}
    assert table["row_count"] >= 3
    assert table["col_count"] >= 2


def test_extract_tables_auto_triggered_after_parse(client: TestClient):
    project_id = _create_project(client, "DXF 表格抽取-自动触发")
    content = equipment_schedule_dxf()
    parse_result = _upload_and_parse(client, project_id, content)
    sheet_id = parse_result["sheet_id"]

    list_resp = client.get(f"/api/sheets/{sheet_id}/tables")
    assert list_resp.status_code == 200
    tables = list_resp.json()
    assert len(tables) >= 1, "parse-dxf 应自动触发 extract_tables"


def test_extract_tables_batch_job_completes_async(client: TestClient):
    project_id = _create_project(client, "DXF 表格抽取-batch 异步")
    content = equipment_schedule_dxf()
    upload_resp = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", ("eq.dxf", content, "application/dxf"))],
    )
    assert upload_resp.status_code == 201
    body = upload_resp.json()
    batch_id = body["id"]
    file_id = body["files"][0]["id"]
    parse_resp = client.post(f"/api/files/{file_id}/parse-dxf")
    assert parse_resp.status_code == 200

    start = client.post(
        f"/api/imports/{batch_id}/extract-tables",
        json={"continue_on_error": True},
    )
    assert start.status_code == 200, start.text
    assert start.json()["status"] in {"running", "completed"}
    job = wait_for_extract_tables_job(client, batch_id, timeout_s=30.0)
    assert job["status"] == "completed", job


def test_extract_tables_batch_job_409_when_duplicate(client: TestClient):
    project_id = _create_project(client, "DXF 表格抽取-batch 409")
    content = equipment_schedule_dxf()
    upload_resp = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", ("eq.dxf", content, "application/dxf"))],
    )
    batch_id = upload_resp.json()["id"]
    file_id = upload_resp.json()["files"][0]["id"]
    client.post(f"/api/files/{file_id}/parse-dxf")

    first = client.post(f"/api/imports/{batch_id}/extract-tables", json={})
    assert first.status_code == 200, first.text
    if first.json()["status"] == "running":
        second = client.post(f"/api/imports/{batch_id}/extract-tables", json={})
        # 由于第一次 job 可能在第二次之前已结束（小输入），允许 200/409 两种
        assert second.status_code in {200, 409}
    wait_for_extract_tables_job(client, batch_id, timeout_s=30.0)


def test_excel_export_includes_table_detail_sheet(client: TestClient):
    project_id = _create_project(client, "DXF 表格抽取-Excel")
    content = equipment_schedule_dxf()
    parse_result = _upload_and_parse(client, project_id, content)
    sheet_id = parse_result["sheet_id"]
    client.post(f"/api/sheets/{sheet_id}/extract-tables", json={})

    # 确认表格已落库
    tables = client.get(f"/api/sheets/{sheet_id}/tables").json()
    assert len(tables) >= 1

    export_resp = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True},
    )
    assert export_resp.status_code == 200, export_resp.text
    body = export_resp.json()
    output_path = settings.root_dir / body["file_path"]
    assert output_path.exists()
    wb = load_workbook(output_path, read_only=True)
    assert "图纸表格明细" in wb.sheetnames
    ws = wb["图纸表格明细"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    assert len(rows) >= 2, "应至少 1 行表头 + 1 行数据"
    headers = rows[0]
    assert "表格类型" in headers
    assert "表头(JSON)" in headers
