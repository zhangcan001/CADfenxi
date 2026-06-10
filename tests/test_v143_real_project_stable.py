"""v1.4.3 真实项目稳定包整理测试。"""
from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.core.config import settings
from backend.main import app
from scripts import build_portable_package
from test_v142_excel_delivery_polish import (
    EXPECTED_SHEETS,
    test_v142_excel_delivery_order_headers_styles_and_safe_details,
    test_v142_import_and_core_flows_do_not_regress,
)
from test_v14_real_project_trial_guard import _run_realistic_trial_project


VERSION = "v1.5.1-fast-delivery-package-fix"


def test_v143_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v143_real_project_import_deep_extract_excel_backup_and_health(tmp_path: Path):
    with TestClient(app) as client:
        result = _run_realistic_trial_project(client, tmp_path)

    assert result["upload"]["file_type_counts"] == {"pdf": 1, "dxf": 2, "dwg": 1, "unsupported": 0}
    assert result["pdf_split"].status_code == 200
    assert result["pipeline"]["status"] == "success"
    assert result["pipeline"]["summary"]["converted_success"] == 1
    assert result["pipeline"]["summary"]["parse_success"] >= 3
    assert result["consistency"].status_code == 200
    assert result["consistency"].json()["by_code"].get("CROSS_DRAWING_NAME_CONFLICT", 0) >= 1
    assert result["before_state"] == result["after_state"]
    assert result["backup"]["file_size"] > 0
    assert result["restore"].status_code == 200
    assert result["restored_export"]["ledger_row_count"] == result["export"]["ledger_row_count"]
    assert result["project_health"].status_code == 200
    assert result["restored_health"].status_code == 200
    assert result["system_health"].status_code == 200

    workbook = load_workbook(settings.root_dir / result["export"]["file_path"], read_only=True)
    try:
        assert workbook.sheetnames == EXPECTED_SHEETS
        assert workbook["图纸总台账"].max_row - 1 == len(result["sheets"])
        table_text = "\n".join(str(cell) for row in workbook["图纸表格明细"].iter_rows(values_only=True) for cell in row if cell)
        block_text = "\n".join(str(cell) for row in workbook["图纸块统计"].iter_rows(values_only=True) for cell in row if cell)
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        assert "送风机" in table_text
        assert "LAMP" in block_text
        assert "同图号图名不一致" in issue_text
    finally:
        workbook.close()


def test_v143_excel_delivery_and_core_flows_do_not_regress(tmp_path: Path):
    test_v142_excel_delivery_order_headers_styles_and_safe_details(tmp_path)
    test_v142_import_and_core_flows_do_not_regress(tmp_path)


def test_v143_release_report_and_portable_metadata_exist():
    report = settings.root_dir / "docs" / "FAST_RELEASE_REPORT_v1.4.3.md"
    content = report.read_text(encoding="utf-8")

    assert report.is_file()
    assert "v1.4.3-fast-real-project-stable" in content
    assert "导入" in content
    assert "Excel" in content
    assert "备份恢复" in content
    assert "健康检查" in content
    assert build_portable_package.DEFAULT_VERSION == VERSION
    assert build_portable_package.package_name() == f"工程图纸智能台账识别系统-{VERSION}"
