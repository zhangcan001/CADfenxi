from datetime import date
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue


def create_project_file(client: TestClient):
    project = client.post("/api/projects", json={"name": f"Excel模板项目-{uuid4().hex[:8]}"})
    assert project.status_code == 201
    project_id = project.json()["id"]
    upload = client.post(
        f"/api/projects/{project_id}/imports",
        data={"batch_name": "Excel 模板批次"},
        files=[("files", ("template.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"))],
    )
    assert upload.status_code == 201
    return project_id, upload.json()["id"], upload.json()["files"][0]["id"], project.json()["name"]


def add_file(project_id: int, batch_id: int, name: str, source_format: str) -> int:
    with SessionLocal() as db:
        drawing_file = DrawingFile(
            project_id=project_id,
            batch_id=batch_id,
            original_name=name,
            file_ext=f".{source_format}",
            source_format=source_format,
            file_size=10,
            file_hash=uuid4().hex,
            storage_path=f"app_data/temp/{uuid4().hex}_{name}",
            status="preprocessed",
            convert_status="success" if source_format == "dwg" else None,
            converted_format="dxf" if source_format == "dwg" else None,
            converted_file_path=f"app_data/temp/{uuid4().hex}.dxf" if source_format == "dwg" else None,
        )
        db.add(drawing_file)
        db.flush()
        file_id = drawing_file.id
        db.commit()
    return file_id


def add_sheet(project_id: int, batch_id: int, file_id: int, drawing_no: str, discipline: str, source: str = "pdf_text") -> int:
    with SessionLocal() as db:
        page_no = db.query(DrawingSheet).filter(DrawingSheet.file_id == file_id).count() + 1
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=page_no,
            sheet_type="unknown",
            drawing_no=drawing_no,
            drawing_name=f"{discipline}平面图",
            discipline=discipline,
            version="A",
            issue_date=date(2026, 5, 21),
            status="confirmed",
            review_status="confirmed",
            trust_level="A",
            confidence_score=96,
        )
        db.add(sheet)
        db.flush()
        sheet_id = sheet.id
        for field_name, value in {
            "drawing_no": drawing_no,
            "drawing_name": f"{discipline}平面图",
            "discipline": discipline,
            "version": "A",
            "issue_date": "2026-05-21",
        }.items():
            db.add(
                FieldValue(
                    project_id=project_id,
                    batch_id=batch_id,
                    file_id=file_id,
                    sheet_id=sheet_id,
                    field_name=field_name,
                    raw_value=value,
                    normalized_value=value,
                    display_value=value,
                    final_source=source,
                    confidence=96,
                    is_reviewed=True,
                )
            )
        db.commit()
    return sheet_id


def add_issue(project_id: int, batch_id: int, file_id: int, sheet_id: int) -> None:
    with SessionLocal() as db:
        db.add(
            DrawingIssue(
                project_id=project_id,
                batch_id=batch_id,
                file_id=file_id,
                sheet_id=sheet_id,
                issue_code="DRAWING_NO_EMPTY",
                severity="warning",
                message="测试问题",
                suggestion="请复核图号",
                status="open",
            )
        )
        db.commit()


def export_workbook(client: TestClient, project_id: int):
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200
    path = settings.root_dir / response.json()["file_path"]
    assert path.exists()
    return response.json(), load_workbook(path)


def test_excel_export_template_contains_required_sheets_headers_and_styles():
    with TestClient(app) as client:
        project_id, batch_id, pdf_file, project_name = create_project_file(client)
        dxf_file = add_file(project_id, batch_id, "template.dxf", "dxf")
        dwg_file = add_file(project_id, batch_id, "template.dwg", "dwg")
        add_sheet(project_id, batch_id, pdf_file, "建施-01", "建筑", "pdf_text")
        add_sheet(project_id, batch_id, dxf_file, "电施-01", "电气", "cad_block_attr")
        dwg_sheet = add_sheet(project_id, batch_id, dwg_file, "水施-01", "给排水", "cad_text")
        add_issue(project_id, batch_id, dwg_file, dwg_sheet)
        result, workbook = export_workbook(client, project_id)

    assert result["ledger_row_count"] == 3
    assert result["issue_row_count"] == 1
    assert ["图纸总台账", "问题清单", "专业汇总", "图纸表格明细", "图纸块统计", "校核状态汇总", "导出说明"] == workbook.sheetnames
    assert [cell.value for cell in workbook["图纸总台账"][1]] == [
        "序号", "项目名称", "专业", "图纸编号", "图纸名称", "版本", "出图日期",
        "文件格式", "原始文件名", "来源文件", "识别来源", "可信等级", "识别评分",
        "校核状态", "确认时间", "问题数量", "错误数量", "警告数量", "备注",
    ]
    assert workbook["图纸总台账"].max_row == 4
    assert workbook["问题清单"].max_row == 2
    assert workbook["图纸总台账"].freeze_panes == "A2"
    assert workbook["图纸总台账"].auto_filter.ref is not None
    assert workbook["图纸总台账"]["A1"].font.bold is True
    assert workbook["导出说明"]["B3"].value == settings.version
    assert workbook["导出说明"]["B4"].value == project_name
    formats = [workbook["图纸总台账"].cell(row=row, column=8).value for row in range(2, 5)]
    assert {"PDF", "DXF", "DWG转换"} == set(formats)


def test_excel_summary_sheets_count_disciplines_and_review_status():
    with TestClient(app) as client:
        project_id, batch_id, pdf_file, _project_name = create_project_file(client)
        add_sheet(project_id, batch_id, pdf_file, "建施-01", "建筑")
        add_sheet(project_id, batch_id, pdf_file, "建施-02", "建筑")
        structure = add_sheet(project_id, batch_id, pdf_file, "结施-01", "结构")
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, structure)
            sheet.review_status = "unreviewed"
            sheet.trust_level = "D"
            db.commit()
        _result, workbook = export_workbook(client, project_id)

    discipline_rows = {
        workbook["专业汇总"].cell(row=row, column=1).value: [
            workbook["专业汇总"].cell(row=row, column=col).value for col in range(2, 10)
        ]
        for row in range(2, workbook["专业汇总"].max_row + 1)
    }
    assert discipline_rows["建筑"][0] == 2
    assert discipline_rows["建筑"][1] == 2
    assert discipline_rows["结构"][0] == 1
    assert discipline_rows["结构"][2] == 1

    review_summary = {
        workbook["校核状态汇总"].cell(row=row, column=1).value: workbook["校核状态汇总"].cell(row=row, column=2).value
        for row in range(2, workbook["校核状态汇总"].max_row + 1)
    }
    assert review_summary["图纸总数"] == 3
    assert review_summary["已确认图纸"] == 2
    assert review_summary["未校核图纸"] == 1
    assert review_summary["D 级图纸"] == 1
