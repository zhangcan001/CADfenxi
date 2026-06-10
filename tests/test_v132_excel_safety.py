"""v1.3.2 Excel 深度抽取导出安全收口测试。"""
from __future__ import annotations

import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.drawing_table import DrawingTable
from backend.services import issue_service
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    dxf_with_insert_blocks,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from recognizer.cad_engine.cad_json_writer import cad_parse_output_path, write_cad_json
from test_full_flow_stability_v055 import make_pdf_bytes
from test_project_backup_restore import backup_project, create_project, upload_files
from test_v11_fast_ux import export_excel
from test_v13_deep_extract_review import add_confirmed_sheet, make_cad_json, post_import, text_item, upload_and_parse_dxf


VERSION = "v1.5.1-fast-delivery-package-fix"


def _issue_statuses(project_id: int) -> dict[int, str]:
    with SessionLocal() as db:
        rows = db.scalars(select(DrawingIssue).where(DrawingIssue.project_id == project_id)).all()
        return {row.id: row.status for row in rows}


def _review_statuses(project_id: int) -> dict[int, str]:
    with SessionLocal() as db:
        rows = db.scalars(select(DrawingSheet).where(DrawingSheet.project_id == project_id)).all()
        return {row.id: row.review_status for row in rows}


def test_v132_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v132_excel_deep_extract_sheets_are_isolated_and_safe():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.2 Excel 安全")
        upload, parse = upload_and_parse_dxf(
            client,
            project_id,
            dxf_with_insert_blocks(
                [
                    {"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (10, 0)]},
                    {"name": "A1_FRAME", "layer": "0", "positions": [(0, 50)]},
                ]
            ),
            filename="excel-safety.dxf",
        )
        sheet_id = parse["sheet_id"]
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            assert sheet is not None
            cad_json = make_cad_json(
                texts=[
                    text_item("设备名", 0, 120),
                    text_item("型号", 60, 120),
                    text_item("数量", 120, 120),
                    text_item("风机", 0, 112),
                    text_item("F-1", 60, 112),
                    text_item("2", 120, 112),
                    text_item("水泵", 0, 104),
                    text_item("P-1", 60, 104),
                    text_item("1", 120, 104),
                ],
                inserts=[
                    {"block_name": "LAMP", "layer": "EE-LIGHT", "insert": [0, 0, 0], "attribs": []},
                    {"block_name": "LAMP", "layer": "EE-LIGHT", "insert": [10, 0, 0], "attribs": []},
                    {"block_name": "A1_FRAME", "layer": "0", "insert": [0, 50, 0], "attribs": []},
                ],
            )
            write_cad_json(cad_json, cad_parse_output_path(settings.root_dir, sheet.project_id, sheet.id))
            db.commit()
        table_extract = client.post(f"/api/sheets/{sheet_id}/extract-tables", json={"force": True})
        block_extract = client.post(f"/api/sheets/{sheet_id}/extract-blocks", json={"force": True})
        with SessionLocal() as db:
            table = db.scalar(select(DrawingTable).where(DrawingTable.sheet_id == sheet_id))
            assert table is not None
            warnings = json.loads(table.warnings_json or "[]")
            if "LOW_CONFIDENCE_TABLE" not in warnings:
                warnings.append("LOW_CONFIDENCE_TABLE")
            table.warnings_json = json.dumps(warnings, ensure_ascii=False)
            db.commit()
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-132", "drawing_name": "人工确认总台账", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3.2"})
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            assert sheet is not None
            issue_service.add_issue_for_sheet(
                db,
                sheet,
                "CROSS_DRAWING_NAME_CONFLICT",
                severity="warning",
                message="图号「人工-132」对应多个图名：人工确认总台账 / 其他图名。",
                suggestion="请核对同图号图纸是否为不同版本或图名录入错误。",
            )
            db.commit()
        check = client.post(f"/api/projects/{project_id}/exports/check")
        before_reviews = _review_statuses(project_id)
        before_issues = _issue_statuses(project_id)
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        after_reviews = _review_statuses(project_id)
        after_issues = _issue_statuses(project_id)

    assert table_extract.status_code == 200
    assert block_extract.status_code == 200
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert check.status_code == 200
    check_data = check.json()
    assert check_data["can_export"] is True
    assert check_data["drawing_table_count"] >= 1
    assert check_data["low_confidence_table_count"] >= 1
    assert check_data["block_stats_sheet_count"] == 1
    assert check_data["consistency_issue_count"] >= 1
    assert "低可信表格" in "\n".join(check_data["warnings"])
    assert export.status_code == 200
    assert export.json()["precheck"]["drawing_table_count"] >= 1
    assert export.json()["precheck"]["low_confidence_table_count"] >= 1
    assert export.json()["precheck"]["block_stats_sheet_count"] == 1
    assert export.json()["precheck"]["consistency_issue_count"] >= 1
    assert before_reviews == after_reviews
    assert before_issues == after_issues

    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=False)
    try:
        expected = {
            "图纸总台账",
            "问题清单",
            "专业汇总",
            "校核状态汇总",
            "导出说明",
            "图纸表格明细",
            "图纸块统计",
        }
        assert expected.issubset(set(workbook.sheetnames))
        for worksheet in workbook.worksheets:
            assert worksheet.freeze_panes == "A2"
            assert worksheet.auto_filter.ref is not None

        ledger_headers = [cell.value for cell in workbook["图纸总台账"][1]]
        assert "表头(JSON)" not in ledger_headers
        assert "数据(JSON)" not in ledger_headers
        assert "块名" not in ledger_headers
        assert "数量" not in ledger_headers
        assert workbook["图纸总台账"]["D2"].value == "人工-132"
        assert workbook["图纸总台账"]["E2"].value == "人工确认总台账"

        table_headers = [cell.value for cell in workbook["图纸表格明细"][1]]
        block_headers = [cell.value for cell in workbook["图纸块统计"][1]]
        assert {"Sheet ID", "图号", "图名", "可信度", "表头(JSON)", "数据(JSON)"}.issubset(set(table_headers))
        assert {"Sheet ID", "图号", "图名", "块名", "数量"}.issubset(set(block_headers))
        table_text = "\n".join(str(cell) for row in workbook["图纸表格明细"].iter_rows(values_only=True) for cell in row if cell)
        block_text = "\n".join(str(cell) for row in workbook["图纸块统计"].iter_rows(values_only=True) for cell in row if cell)
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        info_text = "\n".join(str(cell) for row in workbook["导出说明"].iter_rows(values_only=True) for cell in row if cell)
        assert "低可信" in table_text
        assert "LAMP" in block_text
        assert "A1_FRAME" not in block_text
        assert "同图号图名不一致" in issue_text
        assert "图纸表格明细" in info_text and "不会被表格明细或块统计覆盖" in info_text
    finally:
        workbook.close()


def test_v132_pdf_dxf_dwg_cad_preview_backup_health_do_not_regress(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        pdf_project_id = create_project(client, "v1.3.2 PDF 回归")
        pdf_upload = upload_files(client, pdf_project_id, [("v132.pdf", make_pdf_bytes("图号 建施-132"), "application/pdf")])
        pdf_split = client.post(f"/api/files/{pdf_upload['files'][0]['id']}/split")

        project_id = create_project(client, "v1.3.2 核心闭环")
        cad_upload = upload_files(
            client,
            project_id,
            [
                ("v132-pipeline.dxf", dxf_with_insert_blocks([{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0)]}]), "application/dxf"),
                ("v132-pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        pipeline = run_cad_pipeline_blocking(
            client,
            cad_upload["id"],
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        first_sheet_id = sheets[0]["id"]
        update = client.patch(
            f"/api/sheets/{first_sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-132R", "drawing_name": "Excel 安全回归", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{first_sheet_id}/confirm", json={"force": True, "note": "v1.3.2"})
        preview = client.post(f"/api/sheets/{first_sheet_id}/cad-preview")
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        system_health = client.get("/api/system/health-check")

    assert pdf_split.status_code == 200
    assert len(pdf_split.json()["sheets"]) >= 1
    assert pipeline["summary"]["dwg_files"] == 1
    assert pipeline["summary"]["converted_success"] == 1
    assert pipeline["summary"]["parse_success"] >= 2
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert preview.status_code == 200
    assert export["ledger_row_count"] >= 1
    assert backup["backup_id"] > 0
    assert project_health.status_code == 200
    assert system_health.status_code == 200
