import io
from pathlib import Path
from uuid import uuid4

import ezdxf
import pymupdf as fitz
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.cad_conversion_run import CadConversionRun
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.export_record import ExportRecord
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from backend.models.review_audit_log import ReviewAuditLog
from dwg_test_helpers import (
    DWG_BYTES,
    DXF_TEXT,
    clear_converter_tables,
    create_converter_setting,
    run_cad_pipeline_blocking,
    write_mock_converter,
)


def make_pdf_bytes(text: str = "建施-55 稳定性PDF") -> bytes:
    document = fitz.open()
    page = document.new_page(width=360, height=240)
    page.insert_text((36, 72), text, fontsize=12)
    payload = document.tobytes()
    document.close()
    return payload


def title_block_dxf(drawing_no: str = "建施-55", drawing_name: str = "稳定性平面图") -> bytes:
    doc = ezdxf.new("R2010")
    doc.layers.add("ARCH")
    msp = doc.modelspace()
    msp.add_text(drawing_no, dxfattribs={"layer": "ARCH", "insert": (10, 10, 0)})
    msp.add_mtext(drawing_name, dxfattribs={"layer": "ARCH", "insert": (10, 20, 0)})
    block = doc.blocks.new(name="TITLE_BLOCK")
    for tag in ["DRAWING_NO", "DRAWING_NAME", "DATE", "REV"]:
        block.add_attdef(tag, insert=(0, 0, 0), dxfattribs={"height": 3.5})
    insert = msp.add_blockref("TITLE_BLOCK", (0, 0, 0))
    insert.add_auto_attribs(
        {"DRAWING_NO": drawing_no, "DRAWING_NAME": drawing_name, "DATE": "2026-05-21", "REV": "A"}
    )
    stream = io.StringIO()
    doc.write(stream)
    return stream.getvalue().encode("utf-8")


def create_project(client: TestClient, name: str) -> int:
    response = client.post("/api/projects", json={"name": f"{name}-{uuid4().hex[:8]}"})
    assert response.status_code == 201
    return response.json()["id"]


def upload(client: TestClient, project_id: int, files: list[tuple[str, bytes, str]]) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", item) for item in files],
    )
    assert response.status_code == 201
    return response.json()


def export_excel(client: TestClient, project_id: int) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200
    return response.json()


def run_pipeline(client: TestClient, batch_id: int, steps: list[str] | None = None, skip_completed: bool = True) -> dict:
    payload = {
        "steps": steps or ["convert_dwg", "prepare_dxf_sheet", "parse_dxf", "generate_candidates", "fuse_fields"],
        "skip_completed": skip_completed,
        "continue_on_error": True,
    }
    return run_cad_pipeline_blocking(client, batch_id, payload)


def counts(sheet_id: int) -> tuple[int, int]:
    with SessionLocal() as db:
        candidate_count = db.scalar(
            select(func.count()).select_from(RecognitionCandidate).where(RecognitionCandidate.sheet_id == sheet_id)
        ) or 0
        open_issue_count = db.scalar(
            select(func.count()).select_from(DrawingIssue).where(
                DrawingIssue.sheet_id == sheet_id,
                DrawingIssue.status == "open",
            )
        ) or 0
    return candidate_count, open_issue_count


def sheet_snapshot(sheet_id: int) -> dict:
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        values = [
            (value.field_name, value.display_value, value.is_reviewed)
            for value in db.scalars(select(FieldValue).where(FieldValue.sheet_id == sheet_id)).all()
        ]
        issue_statuses = [
            issue.status for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
        ]
        return {
            "drawing_no": sheet.drawing_no,
            "status": sheet.status,
            "review_status": sheet.review_status,
            "values": sorted(values),
            "issues": sorted(issue_statuses),
        }


def test_v055_portable_health_and_storage_regression():
    with TestClient(app) as client:
        health = client.get("/api/health")

    assert health.status_code == 200
    assert health.json()["status"] == "ok"
    assert health.json()["version"] == settings.version
    assert health.json()["storage"] == "ok"
    assert Path("scripts/start_local.bat").exists()
    assert Path("scripts/local_launcher.py").exists()
    assert Path("scripts/check_env.bat").exists()
    assert settings.check_storage() is True


def test_v055_pdf_minimal_full_flow_and_excel_consistency():
    with TestClient(app) as client:
        project_id = create_project(client, "PDF全流程稳定性")
        upload_result = upload(client, project_id, [("建施-55_PDF稳定性.pdf", make_pdf_bytes(), "application/pdf")])
        batch_id = upload_result["id"]
        file_id = upload_result["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        assert split.status_code == 200
        crop = client.post(f"/api/imports/{batch_id}/title-crops")
        text = client.post(f"/api/imports/{batch_id}/extract-text")
        ocr = client.post(f"/api/imports/{batch_id}/ocr-titles")
        sheet_id = split.json()["sheets"][0]["id"]
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        review = client.post(
            f"/api/review/sheets/{sheet_id}/update-fields",
            json={"fields": {"drawing_no": "PDF-人工-55", "drawing_name": "PDF人工图名", "discipline": "建筑"}},
        )
        before = sheet_snapshot(sheet_id)
        export = export_excel(client, project_id)
        after = sheet_snapshot(sheet_id)

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert crop.status_code == 200
    assert text.status_code == 200
    assert ocr.status_code == 200
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert review.status_code == 200
    assert export["ledger_row_count"] == 1
    assert workbook["图纸总台账"].max_row - 1 == 1
    assert workbook["图纸总台账"].cell(2, 4).value == "PDF-人工-55"
    assert before == after


def test_v055_dxf_flow_idempotency_and_error_paths():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF全流程稳定性")
        upload_result = upload(client, project_id, [("建施-55_稳定性.dxf", title_block_dxf(), "application/dxf")])
        file_id = upload_result["files"][0]["id"]
        first_prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        second_prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        first_parse = client.post(f"/api/files/{file_id}/parse-dxf")
        second_parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = first_parse.json()["sheet_id"]
        summary = client.get(f"/api/sheets/{sheet_id}/cad-parse")
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        first_counts = counts(sheet_id)
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        second_counts = counts(sheet_id)
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        third_counts = counts(sheet_id)
        export = export_excel(client, project_id)
        broken = upload(client, project_id, [("broken.dxf", b"not a dxf", "application/dxf")])
        broken_parse = client.post(f"/api/files/{broken['files'][0]['id']}/parse-dxf")

    with SessionLocal() as db:
        sheet_total = db.scalar(select(func.count()).select_from(DrawingSheet).where(DrawingSheet.file_id == file_id))
        cad_candidate = db.scalar(
            select(RecognitionCandidate.id).where(
                RecognitionCandidate.sheet_id == sheet_id,
                RecognitionCandidate.source_type.like("cad_%"),
            )
        )
    assert first_prepare.status_code == 200
    assert second_prepare.status_code == 200
    assert second_prepare.json()["created"] is False
    assert first_parse.status_code == 200
    assert second_parse.status_code == 200
    assert summary.status_code == 200
    assert sheet_total == 1
    assert first_counts[0] == second_counts[0]
    assert second_counts[1] == third_counts[1]
    assert cad_candidate is not None
    assert export["ledger_row_count"] == 1
    assert broken_parse.status_code == 200
    assert broken_parse.json()["status"] == "failed"
    assert broken_parse.json()["error_code"] in {"DXF_PARSE_FAILED", "DXF_OPEN_FAILED"}


def test_v055_dwg_mock_conversion_full_flow_and_missing_converter(tmp_path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        no_config_project = create_project(client, "DWG未配置稳定性")
        no_config_upload = upload(client, no_config_project, [("need-config.dwg", DWG_BYTES, "application/acad")])
        no_config = client.post(f"/api/files/{no_config_upload['files'][0]['id']}/convert-dwg-to-dxf")
        create_converter_setting(client, converter)
        project_id = create_project(client, "DWG转换稳定性")
        upload_result = upload(client, project_id, [("stable.dwg", DWG_BYTES, "application/acad")])
        file_id = upload_result["files"][0]["id"]
        convert = client.post(f"/api/files/{file_id}/convert-dwg-to-dxf")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        generate = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)
        runs = client.get(f"/api/files/{file_id}/cad-conversion-runs")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert no_config.status_code == 400
    assert no_config.json()["detail"]["error_code"] == "CONVERTER_NOT_CONFIGURED"
    assert convert.status_code == 200
    assert convert.json()["status"] == "success"
    assert (settings.root_dir / convert.json()["converted_file_path"]).exists()
    assert parse.status_code == 200
    assert generate.status_code == 200
    assert fuse.status_code == 200
    assert runs.json()[0]["status"] == "success"
    assert workbook["图纸总台账"].cell(2, 8).value == "DWG转换"
    with SessionLocal() as db:
        drawing_file = db.get(DrawingSheet, sheet_id).file
        assert drawing_file.convert_status == "success"
        assert db.scalar(select(CadConversionRun.id).where(CadConversionRun.source_file_id == file_id)) is not None


def test_v055_cad_pipeline_mixed_batch_repeat_and_pdf_untouched(tmp_path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "CADPipeline稳定性")
        upload_result = upload(
            client,
            project_id,
            [
                ("pipe.pdf", make_pdf_bytes("PDF 不应被 CAD pipeline 处理"), "application/pdf"),
                ("pipe.dxf", DXF_TEXT.encode("utf-8"), "application/dxf"),
                ("pipe.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        batch_id = upload_result["id"]
        pdf_file_id = next(item["id"] for item in upload_result["files"] if item["source_format"] == "pdf")
        pdf_split = client.post(f"/api/files/{pdf_file_id}/split")
        first = run_pipeline(client, batch_id)
        second = run_pipeline(client, batch_id)
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        export = export_excel(client, project_id)

    assert pdf_split.status_code == 200
    assert first["summary"]["pdf_files"] == 1
    assert first["summary"]["dxf_files"] == 1
    assert first["summary"]["dwg_files"] == 1
    assert first["summary"]["converted_success"] == 1
    assert first["summary"]["parse_success"] == 2
    assert second["status"] in {"skipped", "success"}
    assert {sheet["source_format"] for sheet in sheets} >= {"pdf", "dxf", "dwg"}
    assert export["ledger_row_count"] == len(sheets)
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(DrawingSheet).where(DrawingSheet.file_id == pdf_file_id)) == 1


def test_v055_manual_confirmation_survives_regeneration_pipeline_and_export(tmp_path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "人工确认保护稳定性")
        upload_result = upload(client, project_id, [("manual.dxf", title_block_dxf("建施-03"), "application/dxf")])
        batch_id = upload_result["id"]
        first = run_pipeline(client, batch_id, steps=["prepare_dxf_sheet", "parse_dxf", "generate_candidates", "fuse_fields"])
        sheet_id = first["steps"][1]["items"][0]["sheet_id"]
        update = client.post(
            f"/api/review/sheets/{sheet_id}/update-fields",
            json={"fields": {"drawing_no": "建施-03A", "drawing_name": "人工确认图名", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/review/sheets/{sheet_id}/confirm", json={"force": False, "note": "确认"})
        assert update.status_code == 200
        assert confirm.status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        run_pipeline(client, batch_id, steps=["generate_candidates", "fuse_fields"], skip_completed=False)
        sheet = client.get(f"/api/sheets/{sheet_id}").json()
        export = export_excel(client, project_id)

    workbook = load_workbook(settings.root_dir / export["file_path"])
    values = client_values(sheet_id)
    assert sheet["drawing_no"] == "建施-03A"
    assert sheet["review_status"] == "confirmed"
    assert values["drawing_no"]["is_reviewed"] is True
    assert workbook["图纸总台账"].cell(2, 4).value == "建施-03A"
    with SessionLocal() as db:
        assert db.scalar(select(ReviewAuditLog.id).where(ReviewAuditLog.sheet_id == sheet_id)) is not None


def test_v055_empty_export_and_unsupported_upload_errors_are_clear():
    with TestClient(app) as client:
        project_id = create_project(client, "异常提示稳定性")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        unsupported = client.post(
            f"/api/projects/{project_id}/imports",
            files=[("files", ("bad.txt", b"hello", "text/plain"))],
        )

    assert export.status_code == 409
    assert "当前项目没有图纸" in export.json()["detail"]["summary_message"]
    assert unsupported.status_code == 201
    unsupported_body = unsupported.json()
    assert unsupported_body["total_selected"] == 1
    assert unsupported_body["imported_count"] == 0
    assert unsupported_body["unsupported_count"] == 1
    assert unsupported_body["items"][0]["status"] == "unsupported"
    assert unsupported_body["items"][0]["error_code"] == "UNSUPPORTED_FILE_TYPE"
    assert "不支持" in unsupported_body["items"][0]["message"]


def client_values(sheet_id: int) -> dict[str, dict]:
    with SessionLocal() as db:
        return {
            value.field_name: {
                "display_value": value.display_value,
                "is_reviewed": value.is_reviewed,
                "final_source": value.final_source,
            }
            for value in db.scalars(select(FieldValue).where(FieldValue.sheet_id == sheet_id)).all()
        }
