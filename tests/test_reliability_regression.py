from pathlib import Path

import pymupdf as fitz
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.export_record import ExportRecord
from backend.models.field_value import FieldValue
from backend.models.review_audit_log import ReviewAuditLog


def make_pdf_bytes(text: str = "Drawing A-001") -> bytes:
    document = fitz.open()
    page = document.new_page(width=300, height=220)
    page.insert_text((36, 72), text, fontsize=12)
    data = document.tobytes()
    document.close()
    return data


def create_project(client: TestClient, name: str = "可靠性测试项目") -> int:
    response = client.post("/api/projects", json={"name": name})
    assert response.status_code == 201
    return response.json()["id"]


def upload_pdf(client: TestClient, project_id: int, filename: str = "test.pdf", content: bytes | None = None):
    return client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", (filename, content or make_pdf_bytes(), "application/pdf"))],
    )


def add_sheet(project_id: int, batch_id: int, file_id: int, page_no: int = 1):
    with SessionLocal() as db:
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=page_no,
            sheet_type="unknown",
            drawing_no=None,
            drawing_name=None,
            discipline=None,
            version=None,
            status="recognized",
            review_status="unreviewed",
            trust_level="A",
            confidence_score=95,
        )
        db.add(sheet)
        db.commit()
        db.refresh(sheet)
        return sheet.id


def add_field_value(sheet_id: int, field_name: str, display_value: str, is_reviewed: bool = False):
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        value = FieldValue(
            project_id=sheet.project_id,
            batch_id=sheet.batch_id,
            file_id=sheet.file_id,
            sheet_id=sheet.id,
            field_name=field_name,
            raw_value=display_value,
            normalized_value=display_value,
            display_value=display_value,
            final_source="machine",
            confidence=90,
            is_reviewed=is_reviewed,
        )
        db.add(value)
        if field_name in {"drawing_no", "drawing_name", "discipline", "version"}:
            setattr(sheet, field_name, display_value)
        db.commit()
        db.refresh(value)
        return value.id


def add_issue(sheet_id: int, issue_code: str = "DRAWING_NO_EMPTY", severity: str = "error", status: str = "open"):
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        issue = DrawingIssue(
            project_id=sheet.project_id,
            batch_id=sheet.batch_id,
            file_id=sheet.file_id,
            sheet_id=sheet.id,
            issue_code=issue_code,
            severity=severity,
            message="测试问题",
            suggestion="请处理",
            status=status,
        )
        db.add(issue)
        db.commit()
        db.refresh(issue)
        return issue.id


def test_full_flow_smoke():
    with TestClient(app) as client:
        project_id = create_project(client)
        upload = upload_pdf(client, project_id, "full-flow.pdf")
        assert upload.status_code == 201
        batch_id = upload.json()["id"]
        file_id = upload.json()["files"][0]["id"]

        split = client.post(f"/api/files/{file_id}/split")
        assert split.status_code == 200
        sheet_id = split.json()["sheets"][0]["id"]

        title_crop = client.post(f"/api/sheets/{sheet_id}/title-crop")
        assert title_crop.status_code == 200
        client.post(f"/api/sheets/{sheet_id}/extract-text")
        client.post(f"/api/sheets/{sheet_id}/ocr-title")
        client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        review = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "A-001", "drawing_name": "测试图", "discipline": "建筑"}},
        )
        assert review.status_code == 200
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True})
        assert confirm.status_code == 200
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert export.status_code == 200
    data = export.json()
    assert Path(settings.root_dir / data["file_path"]).exists()
    with SessionLocal() as db:
        assert db.scalar(select(DrawingSheet.id).where(DrawingSheet.id == sheet_id)) is not None
        assert db.scalar(select(FieldValue.id).where(FieldValue.sheet_id == sheet_id)) is not None
        assert db.scalar(select(ReviewAuditLog.id).where(ReviewAuditLog.sheet_id == sheet_id)) is not None
        assert db.scalar(select(ExportRecord.id).where(ExportRecord.id == data["export_id"])) is not None


def test_manual_review_protected_from_refusion():
    with TestClient(app) as client:
        project_id = create_project(client, "人工保护项目")
        upload = upload_pdf(client, project_id, "protect.pdf")
        file_id = upload.json()["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        add_field_value(sheet_id, "drawing_no", "A", is_reviewed=False)
        client.patch(f"/api/sheets/{sheet_id}/fields", json={"fields": {"drawing_no": "B"}})
        client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True})
        client.post(f"/api/sheets/{sheet_id}/fuse-fields")

    with SessionLocal() as db:
        value = db.scalar(
            select(FieldValue).where(FieldValue.sheet_id == sheet_id, FieldValue.field_name == "drawing_no")
        )
    assert value is not None
    assert value.display_value == "B"


def test_export_consistency_and_non_mutation():
    with TestClient(app) as client:
        project_id = create_project(client, "导出一致性项目")
        upload = upload_pdf(client, project_id, "export.pdf")
        file_id = upload.json()["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        add_field_value(sheet_id, "drawing_no", "A-100")
        add_field_value(sheet_id, "drawing_name", "总图")
        add_field_value(sheet_id, "discipline", "建筑")
        add_issue(sheet_id, "DRAWING_NO_EMPTY", "warning", "open")
        before_sheet = client.get(f"/api/sheets/{sheet_id}").json()
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        after_sheet = client.get(f"/api/sheets/{sheet_id}").json()

    assert export.status_code == 200
    workbook = load_workbook(settings.root_dir / export.json()["file_path"])
    ledger = workbook["图纸总台账"]
    issues = workbook["问题清单"]
    assert ledger.max_row - 1 == 1
    assert issues.max_row - 1 == 1
    assert ledger.cell(2, 4).value == "A-100"
    assert ledger.cell(2, 5).value == "总图"
    assert ledger.cell(2, 3).value == "建筑"
    assert before_sheet["status"] == after_sheet["status"]
    assert before_sheet["review_status"] == after_sheet["review_status"]


def test_repeat_operations_idempotent():
    with TestClient(app) as client:
        project_id = create_project(client, "幂等项目")
        upload = upload_pdf(client, project_id, "repeat.pdf")
        batch_id = upload.json()["id"]
        file_id = upload.json()["files"][0]["id"]

        first_split = client.post(f"/api/files/{file_id}/split")
        second_split = client.post(f"/api/files/{file_id}/split")
        assert first_split.status_code == 200
        assert second_split.status_code == 200
        assert len(second_split.json()["sheets"]) == len(first_split.json()["sheets"])

        sheet_id = first_split.json()["sheets"][0]["id"]
        client.post(f"/api/sheets/{sheet_id}/title-crop")
        client.post(f"/api/sheets/{sheet_id}/title-crop")
        client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export_1 = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        export_2 = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert export_1.status_code == 200
    assert export_2.status_code == 200
    assert export_1.json()["file_name"] != export_2.json()["file_name"]
    with SessionLocal() as db:
        assert db.scalar(
            select(func.count()).select_from(DrawingSheet).where(DrawingSheet.file_id == file_id)
        ) == len(first_split.json()["sheets"])
        assert db.scalar(
            select(func.count()).select_from(ExportRecord).where(ExportRecord.project_id == project_id)
        ) >= 2


def test_error_paths_return_clear_codes():
    with TestClient(app) as client:
        project_id = create_project(client, "错误项目")
        non_pdf = client.post(
            f"/api/projects/{project_id}/imports",
            files=[("files", ("note.txt", b"hello", "text/plain"))],
        )
        damaged_upload = upload_pdf(client, project_id, "damaged.pdf", b"not a real pdf")
        damaged_split = client.post(f"/api/files/{damaged_upload.json()['files'][0]['id']}/split")
        upload = upload_pdf(client, project_id, "missing-assets.pdf")
        sheet_id = client.post(f"/api/files/{upload.json()['files'][0]['id']}/split").json()["sheets"][0]["id"]
        sheet = client.get(f"/api/sheets/{sheet_id}").json()
        Path(settings.root_dir / sheet["preview_path"]).unlink()
        missing_preview = client.post(f"/api/sheets/{sheet_id}/title-crop")

        crop_upload = upload_pdf(client, project_id, "missing-crop.pdf")
        crop_sheet_id = client.post(
            f"/api/files/{crop_upload.json()['files'][0]['id']}/split"
        ).json()["sheets"][0]["id"]
        crop = client.post(f"/api/sheets/{crop_sheet_id}/title-crop")
        Path(settings.root_dir / crop.json()["title_crop_path"]).unlink()
        missing_crop = client.post(f"/api/sheets/{crop_sheet_id}/ocr-title")
        missing_sheet = client.post("/api/sheets/999999/title-crop")
        missing_project = client.get("/api/projects/999999")
        missing_export = client.get("/api/exports/999999/download")

    assert non_pdf.status_code == 400
    assert non_pdf.json()["detail"]["error_code"] == "UNSUPPORTED_FORMAT"
    assert damaged_split.status_code == 400
    assert damaged_split.json()["detail"]["error_code"] == "PDF_OPEN_FAILED"
    assert missing_preview.status_code == 400
    assert missing_preview.json()["detail"]["error_code"] == "PREVIEW_FILE_MISSING"
    assert missing_crop.status_code == 400
    assert missing_crop.json()["detail"]["error_code"] == "TITLE_CROP_FILE_MISSING"
    assert missing_sheet.status_code == 404
    assert missing_project.status_code == 404
    assert missing_export.status_code == 404


def test_deleted_project_is_protected_when_related_data_exists():
    with TestClient(app) as client:
        empty_project = create_project(client, "可删项目")
        response = client.delete(f"/api/projects/{empty_project}")
        assert response.status_code == 204

        project_id = create_project(client, "保护项目")
        upload = upload_pdf(client, project_id, "keep.pdf")
        file_id = upload.json()["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        assert split.status_code == 200
        response = client.delete(f"/api/projects/{project_id}")

    assert response.status_code == 400


def test_issue_history_survives_refusion():
    with TestClient(app) as client:
        project_id = create_project(client, "问题历史项目")
        upload = upload_pdf(client, project_id, "issue-history.pdf")
        file_id = upload.json()["files"][0]["id"]
        sheet_id = client.post(f"/api/files/{file_id}/split").json()["sheets"][0]["id"]
        issue_id = add_issue(sheet_id, "DRAWING_NAME_EMPTY", "warning", "open")
        resolved = client.patch(f"/api/issues/{issue_id}", json={"status": "resolved", "note": "已处理"})
        refusion = client.post(f"/api/sheets/{sheet_id}/fuse-fields")

    assert resolved.status_code == 200
    assert refusion.status_code == 200
    with SessionLocal() as db:
        issue = db.get(DrawingIssue, issue_id)
        assert issue is not None
        assert issue.status == "resolved"


def test_batch_confirm_only_confirms_a_without_open_error():
    with TestClient(app) as client:
        project_id = create_project(client, "批量确认项目")
        upload = upload_pdf(client, project_id, "batch-confirm.pdf")
        batch_id = upload.json()["id"]
        file_id = upload.json()["files"][0]["id"]
        sheet_ok = add_sheet(project_id, batch_id, file_id, page_no=1)
        sheet_error = add_sheet(project_id, batch_id, file_id, page_no=2)
        sheet_low = add_sheet(project_id, batch_id, file_id, page_no=3)
        sheet_failed = add_sheet(project_id, batch_id, file_id, page_no=4)
        for sheet_id in [sheet_ok, sheet_error, sheet_low, sheet_failed]:
            add_field_value(sheet_id, "drawing_no", f"A-{sheet_id}")
            add_field_value(sheet_id, "drawing_name", "测试图")
            add_field_value(sheet_id, "discipline", "建筑")
        with SessionLocal() as db:
            db.get(DrawingSheet, sheet_low).trust_level = "C"
            failed = db.get(DrawingSheet, sheet_failed)
            failed.status = "failed"
            failed.trust_level = "D"
            db.commit()
        add_issue(sheet_error, "FIELD_CONFLICT_HIGH", "error", "open")
        response = client.post(
            f"/api/projects/{project_id}/batch-confirm",
            json={"sheet_ids": [sheet_ok, sheet_error, sheet_low, sheet_failed], "note": "可靠性批量确认"},
        )

    assert response.status_code == 200
    data = response.json()
    assert data["confirmed_count"] == 1
    assert {item["sheet_id"] for item in data["skipped"]} == {sheet_error, sheet_low, sheet_failed}
    with SessionLocal() as db:
        assert db.get(DrawingSheet, sheet_ok).status == "confirmed"
        assert db.scalar(select(ReviewAuditLog.id).where(ReviewAuditLog.sheet_id == sheet_ok)) is not None
