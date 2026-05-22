import io
from pathlib import Path

import ezdxf
import pymupdf as fitz
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    create_project,
    write_mock_converter,
)
from scripts.build_portable_package import DEFAULT_VERSION


def make_trial_pdf(text: str = "图号：建施-01\n图名：一层平面图\n专业：建筑") -> bytes:
    document = fitz.open()
    page = document.new_page(width=420, height=300)
    page.insert_text((48, 60), text, fontsize=12)
    data = document.tobytes()
    document.close()
    return data


def make_trial_dxf(text: str = "图号：建施-02 图名：二层平面图 专业：建筑") -> bytes:
    document = ezdxf.new("R2010")
    document.modelspace().add_text(text, dxfattribs={"insert": (0, 0, 0)})
    stream = io.StringIO()
    document.write(stream)
    return stream.getvalue().encode("utf-8")


def upload_files(client: TestClient, project_id: int, files: list[tuple[str, bytes, str]]) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        data={"batch_name": "v0.5 真实项目试用回归批次"},
        files=[("files", item) for item in files],
    )
    assert response.status_code == 201, response.text
    return response.json()


def run_cad_pipeline(client: TestClient, batch_id: int, **overrides) -> dict:
    payload = {
        "steps": ["convert_dwg", "prepare_dxf_sheet", "parse_dxf", "generate_candidates", "fuse_fields"],
        "skip_completed": True,
        "continue_on_error": True,
    }
    payload.update(overrides)
    response = client.post(f"/api/imports/{batch_id}/cad-pipeline", json=payload)
    assert response.status_code == 200, response.text
    return response.json()


def field_values_for_sheet(sheet_id: int) -> dict[str, FieldValue]:
    with SessionLocal() as db:
        values = db.query(FieldValue).filter(FieldValue.sheet_id == sheet_id).all()
        return {value.field_name: value for value in values}


def test_pdf_minimal_real_trial_flow_runs_through_candidates_fusion_and_review():
    clear_converter_tables()
    with TestClient(app) as client:
        project_id = create_project(client, "v0.5真实项目试用-PDF")
        upload = upload_files(client, project_id, [("建施-01_一层平面图.pdf", make_trial_pdf(), "application/pdf")])
        file_id = upload["files"][0]["id"]

        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        crop = client.post(f"/api/sheets/{sheet_id}/title-crop")
        text = client.post(f"/api/sheets/{sheet_id}/extract-text")
        ocr = client.post(f"/api/sheets/{sheet_id}/ocr-title")
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fusion = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        manual = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-PDF-001"}, "note": "v0.5 试用校核"},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "试用确认"})
        sheet = client.get(f"/api/sheets/{sheet_id}").json()

    assert split.status_code == 200
    assert crop.status_code == 200
    assert text.status_code == 200
    assert ocr.status_code == 200
    assert candidates.status_code == 200
    assert fusion.status_code == 200
    assert manual.status_code == 200
    assert confirm.status_code == 200
    assert sheet["drawing_no"] == "人工-PDF-001"
    assert sheet["status"] == "confirmed"


def test_dxf_minimal_real_trial_flow_runs_through_parse_candidates_and_fusion():
    clear_converter_tables()
    with TestClient(app) as client:
        project_id = create_project(client, "v0.5真实项目试用-DXF")
        upload = upload_files(client, project_id, [("建施-02_二层平面图.dxf", make_trial_dxf(), "application/dxf")])
        file_id = upload["files"][0]["id"]

        prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fusion = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        sheet = client.get(f"/api/sheets/{sheet_id}").json()
        cad_summary = client.get(f"/api/sheets/{sheet_id}/cad-parse").json()

    assert prepare.status_code == 200
    assert parse.status_code == 200
    assert candidates.status_code == 200
    assert fusion.status_code == 200
    assert sheet["source_format"] == "dxf"
    assert cad_summary["counts"]["text_count"] >= 1
    assert field_values_for_sheet(sheet_id)


def test_dwg_mock_conversion_then_dxf_flow_runs_through():
    clear_converter_tables()
    with TestClient(app) as client:
        project_id = create_project(client, "v0.5真实项目试用-DWG")
        upload = upload_files(client, project_id, [("结构-01_基础平面图.dwg", DWG_BYTES, "application/acad")])
        file_id = upload["files"][0]["id"]
        converter = write_mock_converter(Path(settings.temp_dir))
        create_converter_setting(client, converter)

        convert = client.post(f"/api/files/{file_id}/convert-dwg-to-dxf")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fusion = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        sheet = client.get(f"/api/sheets/{sheet_id}").json()

    assert convert.status_code == 200
    assert convert.json()["status"] == "success"
    assert parse.status_code == 200
    assert candidates.status_code == 200
    assert fusion.status_code == 200
    assert sheet["source_format"] == "dwg"
    assert field_values_for_sheet(sheet_id)


def test_mixed_cad_pipeline_excel_consistency_and_manual_review_protection():
    clear_converter_tables()
    with TestClient(app) as client:
        converter = write_mock_converter(Path(settings.temp_dir))
        create_converter_setting(client, converter)
        project_id = create_project(client, "v0.5真实项目试用-混合")
        upload = upload_files(
            client,
            project_id,
            [
                ("建施-03_三层平面图.dxf", make_trial_dxf("图号：建施-03 图名：三层平面图 专业：建筑"), "application/dxf"),
                ("结施-04_基础平面图.dwg", DWG_BYTES, "application/acad"),
                ("建施-05_说明.pdf", make_trial_pdf("图号：建施-05\n图名：建筑说明\n专业：建筑"), "application/pdf"),
            ],
        )
        pdf_file_id = next(item["id"] for item in upload["files"] if item["source_format"] == "pdf")
        pdf_split = client.post(f"/api/files/{pdf_file_id}/split")
        pipeline = run_cad_pipeline(client, upload["id"])
        sheets_before = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        cad_sheet = next(item for item in sheets_before if item["source_format"] in {"dxf", "dwg"})
        manual = client.patch(
            f"/api/sheets/{cad_sheet['id']}/fields",
            json={"fields": {"drawing_no": "人工-CAD-保护"}, "note": "保护人工字段"},
        )
        rerun = run_cad_pipeline(client, upload["id"], steps=["generate_candidates", "fuse_fields"], skip_completed=False)
        export_check_before = client.post(f"/api/projects/{project_id}/exports/check").json()
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        sheets_after = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        export_check_after = client.post(f"/api/projects/{project_id}/exports/check").json()
        protected = client.get(f"/api/sheets/{cad_sheet['id']}").json()

    workbook = load_workbook(settings.root_dir / export.json()["file_path"])
    ledger_data_rows = workbook["图纸总台账"].max_row - 1

    assert pdf_split.status_code == 200
    assert pipeline["status"] == "success"
    assert pipeline["summary"]["converted_success"] == 1
    assert pipeline["summary"]["parse_success"] == 2
    assert manual.status_code == 200
    assert rerun["status"] == "success"
    assert export.status_code == 200
    assert protected["drawing_no"] == "人工-CAD-保护"
    assert len(sheets_before) == len(sheets_after) == ledger_data_rows
    assert [(item["id"], item["status"], item["review_status"]) for item in sheets_before] == [
        (item["id"], item["status"], item["review_status"]) for item in sheets_after
    ]
    assert export_check_before["unconfirmed_count"] == export_check_after["unconfirmed_count"]


def test_real_trial_regression_keeps_portable_startup_and_p0_p1_tests_in_suite():
    root = Path(__file__).resolve().parents[1]

    assert DEFAULT_VERSION == settings.version
    required_tests = [
        "tests/test_portable_package.py",
        "tests/test_portable_startup_diagnostics.py",
        "tests/test_portable_stable_package.py",
        "tests/test_health.py",
        "tests/test_exports.py",
        "tests/test_review_workbench.py",
        "tests/test_cad_pipeline.py",
    ]
    for relative in required_tests:
        assert (root / relative).is_file()

    with SessionLocal() as db:
        assert db.query(DrawingSheet).count() >= 0
