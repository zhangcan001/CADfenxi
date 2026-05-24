from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.cad_conversion_run import CadConversionRun
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_evidence import FieldEvidence
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from backend.models.review_audit_log import ReviewAuditLog
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf


def create_project(client: TestClient, name: str) -> int:
    response = client.post("/api/projects", json={"name": name})
    assert response.status_code == 201, response.text
    return response.json()["id"]


def upload_files(client: TestClient, project_id: int, files: list[tuple[str, bytes, str]]) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", item) for item in files],
    )
    assert response.status_code == 201, response.text
    return response.json()


def run_pipeline(client: TestClient, batch_id: int) -> dict:
    payload = {
        "steps": ["convert_dwg", "prepare_dxf_sheet", "parse_dxf", "generate_candidates", "fuse_fields"],
        "skip_completed": True,
        "continue_on_error": True,
    }
    return run_cad_pipeline_blocking(client, batch_id, payload)


def export_excel(client: TestClient, project_id: int) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200, response.text
    return response.json()


def project_counts(project_id: int) -> dict[str, int]:
    with SessionLocal() as db:
        return {
            "drawing_files": db.scalar(select(func.count()).select_from(DrawingFile).where(DrawingFile.project_id == project_id)) or 0,
            "drawing_sheets": db.scalar(select(func.count()).select_from(DrawingSheet).where(DrawingSheet.project_id == project_id)) or 0,
            "recognition_candidates": db.scalar(select(func.count()).select_from(RecognitionCandidate).where(RecognitionCandidate.project_id == project_id)) or 0,
            "field_values": db.scalar(select(func.count()).select_from(FieldValue).where(FieldValue.project_id == project_id)) or 0,
            "drawing_issues": db.scalar(select(func.count()).select_from(DrawingIssue).where(DrawingIssue.project_id == project_id)) or 0,
            "review_audit_logs": db.scalar(select(func.count()).select_from(ReviewAuditLog).where(ReviewAuditLog.project_id == project_id)) or 0,
            "cad_conversion_runs": db.scalar(select(func.count()).select_from(CadConversionRun).where(CadConversionRun.project_id == project_id)) or 0,
        }


def assert_counts_and_paths(client: TestClient, original_id: int, restored_id: int) -> None:
    assert project_counts(restored_id) == project_counts(original_id)
    with SessionLocal() as db:
        marker = f"project_{restored_id}"
        old_marker = f"project_{original_id}"
        files = db.scalars(select(DrawingFile).where(DrawingFile.project_id == restored_id)).all()
        sheets = db.scalars(select(DrawingSheet).where(DrawingSheet.project_id == restored_id)).all()
        values = db.scalars(select(FieldValue).where(FieldValue.project_id == restored_id)).all()
        evidence = db.scalars(
            select(FieldEvidence)
            .join(FieldValue, FieldEvidence.field_value_id == FieldValue.id)
            .where(FieldValue.project_id == restored_id)
        ).all()
        for drawing_file in files:
            assert marker in drawing_file.storage_path
            assert old_marker not in drawing_file.storage_path
            assert Path(settings.root_dir / drawing_file.storage_path).exists()
            if drawing_file.converted_file_path:
                assert marker in drawing_file.converted_file_path
                assert old_marker not in drawing_file.converted_file_path
                assert Path(settings.root_dir / drawing_file.converted_file_path).exists()
        for sheet in sheets:
            for path in [sheet.preview_path, sheet.thumbnail_path, sheet.title_crop_path]:
                if path:
                    assert marker in path
                    assert old_marker not in path
                    assert Path(settings.root_dir / path).exists()
        assert all(value.project_id == restored_id for value in values)
        assert all(db.get(RecognitionCandidate, item.candidate_id) is not None for item in evidence)
    integrity = client.get(f"/api/projects/{restored_id}/integrity-check")
    assert integrity.status_code == 200
    assert integrity.json()["path_check"]["invalid_paths"] == 0


def backup_restore(client: TestClient, project_id: int) -> int:
    backup = client.post(f"/api/projects/{project_id}/backup")
    assert backup.status_code == 200, backup.text
    verify = client.get(f"/api/backups/{backup.json()['backup_id']}/verify")
    assert verify.status_code == 200
    assert verify.json()["valid"] is True
    restore = client.post(
        f"/api/backups/{backup.json()['backup_id']}/restore",
        json={"restore_mode": "new_project"},
    )
    assert restore.status_code == 200, restore.text
    return restore.json()["new_project_id"]


def test_pdf_project_backup_restore_realistic_counts_paths_review_and_export():
    with TestClient(app) as client:
        project_id = create_project(client, "v071-PDF")
        upload = upload_files(client, project_id, [("建施-91_一层平面图.pdf", make_pdf_bytes("图号 建施-91\n图名 一层平面图"), "application/pdf")])
        file_id = upload["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        assert client.post(f"/api/sheets/{sheet_id}/title-crop").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/extract-text").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        assert client.patch(f"/api/sheets/{sheet_id}/fields", json={"fields": {"drawing_no": "人工-PDF-91"}, "note": "v071"}).status_code == 200
        export_before = export_excel(client, project_id)
        restored_id = backup_restore(client, project_id)
        export_after = export_excel(client, restored_id)
        restored_sheets = client.get(f"/api/projects/{restored_id}/sheets?page_size=100")
        assert_counts_and_paths(client, project_id, restored_id)

    assert restored_sheets.status_code == 200
    assert load_workbook(settings.root_dir / export_before["file_path"])["图纸总台账"].max_row == load_workbook(settings.root_dir / export_after["file_path"])["图纸总台账"].max_row
    with SessionLocal() as db:
        assert db.scalar(select(FieldValue).where(FieldValue.project_id == restored_id, FieldValue.raw_value == "人工-PDF-91")) is not None
        assert db.scalar(select(ReviewAuditLog).where(ReviewAuditLog.project_id == restored_id)) is not None


def test_dxf_project_backup_restore_realistic_counts_and_cad_json():
    with TestClient(app) as client:
        project_id = create_project(client, "v071-DXF")
        upload = upload_files(
            client,
            project_id,
            [
                ("建施-92_一层.dxf", title_block_dxf("建施-92", "一层平面图"), "application/dxf"),
                ("建施-93_二层.dxf", title_block_dxf("建施-93", "二层平面图"), "application/dxf"),
            ],
        )
        for item in upload["files"]:
            assert client.post(f"/api/files/{item['id']}/prepare-dxf-sheet").status_code == 200
            parse = client.post(f"/api/files/{item['id']}/parse-dxf")
            sheet_id = parse.json()["sheet_id"]
            assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
            assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        export_excel(client, project_id)
        restored_id = backup_restore(client, project_id)
        export_excel(client, restored_id)
        cad_summary = client.get(f"/api/projects/{restored_id}/sheets?page_size=100").json()["items"][0]
        assert_counts_and_paths(client, project_id, restored_id)

    assert cad_summary["source_format"] == "dxf"


def test_dwg_conversion_project_backup_restore_realistic_counts_and_converted_paths(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "v071-DWG")
        upload = upload_files(client, project_id, [("结构-94.dwg", DWG_BYTES, "application/acad")])
        file_id = upload["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/convert-dwg-to-dxf").status_code == 200
        assert client.post(f"/api/files/{file_id}/prepare-dxf-sheet").status_code == 200
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        export_excel(client, project_id)
        restored_id = backup_restore(client, project_id)
        export_excel(client, restored_id)
        assert_counts_and_paths(client, project_id, restored_id)

    assert project_counts(restored_id)["cad_conversion_runs"] == project_counts(project_id)["cad_conversion_runs"]


def test_mixed_project_backup_restore_realistic_pipeline_manual_review_and_delete_backup(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "v071-MIX")
        upload = upload_files(
            client,
            project_id,
            [
                ("建施-95.pdf", make_pdf_bytes("图号 建施-95\n图名 PDF图纸"), "application/pdf"),
                ("建施-96.dxf", title_block_dxf("建施-96", "DXF图纸"), "application/dxf"),
                ("建施-97.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        pdf_file_id = next(item["id"] for item in upload["files"] if item["source_format"] == "pdf")
        assert client.post(f"/api/files/{pdf_file_id}/split").status_code == 200
        pipeline = run_pipeline(client, upload["id"])
        assert pipeline["status"] == "success"
        sheet = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"][0]
        assert client.patch(f"/api/sheets/{sheet['id']}/fields", json={"fields": {"drawing_no": "人工-MIX-95"}, "note": "mix"}).status_code == 200
        export_excel(client, project_id)
        backup = client.post(f"/api/projects/{project_id}/backup").json()
        restored = client.post(f"/api/backups/{backup['backup_id']}/restore", json={"restore_mode": "new_project"}).json()
        restored_id = restored["new_project_id"]
        delete_backup = client.delete(f"/api/backups/{backup['backup_id']}")
        restored_project = client.get(f"/api/projects/{restored_id}")
        export_excel(client, restored_id)
        assert_counts_and_paths(client, project_id, restored_id)

    assert delete_backup.status_code == 204
    assert restored_project.status_code == 200
    with SessionLocal() as db:
        assert db.scalar(select(FieldValue).where(FieldValue.project_id == restored_id, FieldValue.raw_value == "人工-MIX-95")) is not None
