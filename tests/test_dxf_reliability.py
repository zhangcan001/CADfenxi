import io
from pathlib import Path

import ezdxf
import pymupdf as fitz
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.export_record import ExportRecord
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from backend.models.recognition_run import RecognitionRun
from backend.models.review_audit_log import ReviewAuditLog


def create_project(client: TestClient, name: str = "DXF 可靠性测试") -> int:
    response = client.post("/api/projects", json={"name": name})
    assert response.status_code == 201
    return response.json()["id"]


def make_pdf_bytes(text: str = "建施-88 PDF回归图") -> bytes:
    document = fitz.open()
    page = document.new_page(width=360, height=240)
    page.insert_text((36, 72), text, fontsize=12)
    payload = document.tobytes()
    document.close()
    return payload


def dxf_bytes(builder=None) -> bytes:
    doc = ezdxf.new("R2010")
    if builder is not None:
        builder(doc)
    stream = io.StringIO()
    doc.write(stream)
    return stream.getvalue().encode("utf-8")


def title_block_dxf(
    drawing_no: str = "建施-03",
    drawing_name: str = "一层平面图",
    discipline_layer: str = "ARCH",
) -> bytes:
    def build(doc):
        doc.layers.add(discipline_layer)
        doc.layers.add("TITLE")
        msp = doc.modelspace()
        msp.add_text("建施-04", dxfattribs={"layer": "TITLE", "insert": (10, 10, 0)})
        msp.add_mtext("二层平面图", dxfattribs={"layer": "TITLE", "insert": (10, 20, 0)})
        block = doc.blocks.new(name="TITLE_BLOCK")
        for tag in ["DRAWING_NO", "DRAWING_NAME", "DATE", "REV"]:
            block.add_attdef(tag, insert=(0, 0, 0), dxfattribs={"layer": "TITLE", "height": 3.5})
        insert = msp.add_blockref("TITLE_BLOCK", (0, 0, 0), dxfattribs={"layer": "TITLE"})
        insert.add_auto_attribs(
            {
                "DRAWING_NO": drawing_no,
                "DRAWING_NAME": drawing_name,
                "DATE": "2026-05-20",
                "REV": "A版",
            }
        )

    return dxf_bytes(build)


def layer_only_dxf() -> bytes:
    def build(doc):
        for layer in ["ARCH", "STRUCT", "ELEC"]:
            doc.layers.add(layer)
        doc.modelspace().add_line((0, 0), (100, 0), dxfattribs={"layer": "ARCH"})

    return dxf_bytes(build)


def upload_file(client: TestClient, project_id: int, name: str, content: bytes, mime: str) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", (name, content, mime))],
    )
    assert response.status_code == 201
    return response.json()


def upload_dxf(
    client: TestClient,
    project_id: int,
    content: bytes | None = None,
    name: str = "建施-03_一层平面图.dxf",
) -> dict:
    return upload_file(client, project_id, name, content or title_block_dxf(), "application/dxf")


def upload_pdf(client: TestClient, project_id: int, name: str = "建施-88_PDF回归图.pdf") -> dict:
    return upload_file(client, project_id, name, make_pdf_bytes(), "application/pdf")


def run_dxf_flow(client: TestClient, project_id: int, content: bytes | None = None) -> tuple[dict, dict, dict, dict]:
    upload = upload_dxf(client, project_id, content)
    file_id = upload["files"][0]["id"]
    prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
    assert prepare.status_code == 200
    parse = client.post(f"/api/files/{file_id}/parse-dxf")
    assert parse.status_code == 200
    assert parse.json()["status"] == "success"
    generate = client.post(f"/api/sheets/{parse.json()['sheet_id']}/generate-candidates")
    assert generate.status_code == 200
    fuse = client.post(f"/api/sheets/{parse.json()['sheet_id']}/fuse-fields")
    assert fuse.status_code == 200
    return upload, prepare.json(), parse.json(), fuse.json()


def field_map(client: TestClient, sheet_id: int) -> dict[str, dict]:
    values = client.get(f"/api/sheets/{sheet_id}/field-values")
    assert values.status_code == 200
    return {item["field_name"]: item for item in values.json()}


def open_issue_count(project_id: int, sheet_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(
            select(func.count())
            .select_from(DrawingIssue)
            .where(
                DrawingIssue.project_id == project_id,
                DrawingIssue.sheet_id == sheet_id,
                DrawingIssue.status == "open",
            )
        ) or 0


def export_project(client: TestClient, project_id: int) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200
    return response.json()


def test_dxf_complete_main_flow_with_review_and_excel_export():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF完整主流程")
        upload, _, parse, _ = run_dxf_flow(client, project_id)
        sheet_id = parse["sheet_id"]
        review = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-03A"}, "note": "DXF人工校核"},
        )
        export = export_project(client, project_id)

    assert review.status_code == 200
    assert (settings.root_dir / parse["output_path"]).exists()
    assert (settings.root_dir / export["file_path"]).exists()
    with SessionLocal() as db:
        assert db.scalar(select(DrawingSheet.id).where(DrawingSheet.id == sheet_id)) is not None
        assert db.scalar(select(FieldValue.id).where(FieldValue.sheet_id == sheet_id)) is not None
        assert db.scalar(select(DrawingIssue.id).where(DrawingIssue.sheet_id == sheet_id)) is not None
        assert db.scalar(select(ReviewAuditLog.id).where(ReviewAuditLog.sheet_id == sheet_id)) is not None
        assert db.scalar(select(ExportRecord.id).where(ExportRecord.id == export["export_id"])) is not None
        assert db.scalar(
            select(RecognitionCandidate.id).where(
                RecognitionCandidate.sheet_id == sheet_id,
                RecognitionCandidate.source_type.like("cad_%"),
            )
        ) is not None
        assert upload["files"][0]["source_format"] == "dxf"


def test_pdf_and_dxf_mixed_project_regression_exports_both_rows():
    with TestClient(app) as client:
        project_id = create_project(client, "PDF和DXF混合回归")
        pdf_upload = upload_pdf(client, project_id)
        pdf_file_id = pdf_upload["files"][0]["id"]
        pdf_split = client.post(f"/api/files/{pdf_file_id}/split")
        assert pdf_split.status_code == 200
        pdf_sheet_id = pdf_split.json()["sheets"][0]["id"]
        assert client.post(f"/api/sheets/{pdf_sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{pdf_sheet_id}/fuse-fields").status_code == 200

        _, _, dxf_parse, _ = run_dxf_flow(client, project_id)
        dxf_sheet_id = dxf_parse["sheet_id"]
        sheets = client.get(f"/api/projects/{project_id}/sheets").json()["items"]
        export = export_project(client, project_id)

    workbook = load_workbook(settings.root_dir / export["file_path"])
    ledger = workbook["图纸总台账"]
    exported_names = [ledger.cell(row, 9).value for row in range(2, ledger.max_row + 1)]
    with SessionLocal() as db:
        pdf_sheet = db.get(DrawingSheet, pdf_sheet_id)
        dxf_sheet = db.get(DrawingSheet, dxf_sheet_id)
        assert pdf_sheet is not None and pdf_sheet.preview_path
        assert dxf_sheet is not None and dxf_sheet.preview_path is None
        assert db.scalar(
            select(RecognitionRun.id).where(
                RecognitionRun.sheet_id == dxf_sheet_id,
                RecognitionRun.run_type.in_(["pdf_text", "title_ocr"]),
            )
        ) is None
    assert {item["source_format"] for item in sheets} >= {"pdf", "dxf"}
    assert len(exported_names) == 2
    assert any(str(name).endswith(".pdf") for name in exported_names)
    assert any(str(name).endswith(".dxf") for name in exported_names)


def test_dxf_manual_reviewed_field_is_protected_after_regeneration_and_refusion():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF人工字段保护")
        _, _, parse, _ = run_dxf_flow(client, project_id, title_block_dxf("建施-03"))
        sheet_id = parse["sheet_id"]
        manual = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-03A"}, "note": "人工确认图号"},
        )
        assert manual.status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        values = field_map(client, sheet_id)

    assert values["drawing_no"]["display_value"] == "建施-03A"
    assert values["drawing_no"]["is_reviewed"] is True
    assert values["drawing_no"]["final_source"] == "manual"


def test_dxf_repeat_operations_are_idempotent_except_runs_and_exports():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF幂等性")
        upload = upload_dxf(client, project_id)
        file_id = upload["files"][0]["id"]
        first_prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        second_prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        first_parse = client.post(f"/api/files/{file_id}/parse-dxf")
        second_parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = first_parse.json()["sheet_id"]
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        first_candidate_count = len(client.get(f"/api/sheets/{sheet_id}/candidates").json())
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        second_candidate_count = len(client.get(f"/api/sheets/{sheet_id}/candidates").json())
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        first_issue_count = open_issue_count(project_id, sheet_id)
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        second_issue_count = open_issue_count(project_id, sheet_id)
        export_1 = export_project(client, project_id)
        export_2 = export_project(client, project_id)

    assert first_prepare.status_code == 200
    assert second_prepare.status_code == 200
    assert first_prepare.json()["sheet_id"] == second_prepare.json()["sheet_id"]
    assert second_prepare.json()["created"] is False
    assert first_parse.status_code == 200
    assert second_parse.status_code == 200
    assert first_candidate_count == second_candidate_count
    assert first_issue_count == second_issue_count
    assert export_1["export_id"] != export_2["export_id"]
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(DrawingSheet).where(DrawingSheet.file_id == file_id)) == 1
        assert db.scalar(
            select(func.count())
            .select_from(RecognitionRun)
            .where(RecognitionRun.sheet_id == sheet_id, RecognitionRun.run_type == "cad_parse")
        ) >= 2


def test_dxf_error_paths_return_clear_structured_codes():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF异常路径")
        dwg = client.post(
            f"/api/projects/{project_id}/imports",
            files=[("files", ("bad.dwg", b"dwg", "application/acad"))],
        )
        dwg_parse = client.post(f"/api/files/{dwg.json()['files'][0]['id']}/parse-dxf")
        pdf_upload = upload_pdf(client, project_id, "parse-target.pdf")
        pdf_parse = client.post(f"/api/files/{pdf_upload['files'][0]['id']}/parse-dxf")
        broken_upload = upload_dxf(client, project_id, b"broken dxf", "broken.dxf")
        broken_parse = client.post(f"/api/files/{broken_upload['files'][0]['id']}/parse-dxf")
        unparsed_upload = upload_dxf(client, project_id, title_block_dxf(), "unparsed.dxf")
        unparsed_prepare = client.post(f"/api/files/{unparsed_upload['files'][0]['id']}/prepare-dxf-sheet")
        unparsed_candidates = client.post(
            f"/api/sheets/{unparsed_prepare.json()['sheet_id']}/generate-candidates"
        )
        empty_upload = upload_dxf(client, project_id, dxf_bytes(lambda doc: doc.modelspace().add_line((0, 0), (1, 1))), "empty.dxf")
        empty_parse = client.post(f"/api/files/{empty_upload['files'][0]['id']}/parse-dxf")
        missing_file = client.post("/api/files/999999999/parse-dxf")
        missing_sheet = client.post("/api/sheets/999999999/generate-candidates")

    assert dwg.status_code == 201
    assert dwg.json()["files"][0]["source_format"] == "dwg"
    assert dwg_parse.status_code == 400
    assert dwg_parse.json()["detail"]["error_code"] == "DWG_NOT_CONVERTED"
    assert pdf_parse.status_code == 400
    assert pdf_parse.json()["detail"]["error_code"] == "UNSUPPORTED_CAD_FORMAT"
    assert broken_parse.status_code == 200
    assert broken_parse.json()["status"] == "failed"
    assert broken_parse.json()["error_code"] in {"DXF_PARSE_FAILED", "DXF_OPEN_FAILED"}
    assert unparsed_candidates.status_code == 404
    assert unparsed_candidates.json()["detail"]["error_code"] == "CAD_PARSE_NOT_FOUND"
    assert empty_parse.status_code == 200
    assert "DXF_EMPTY_CONTENT" in empty_parse.json()["warnings"]
    assert missing_file.status_code == 404
    assert missing_sheet.status_code == 404


def test_cad_layer_only_affects_discipline():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF图层限制")
        _, _, parse, _ = run_dxf_flow(client, project_id, layer_only_dxf())
        sheet_id = parse["sheet_id"]
        candidates = client.get(f"/api/sheets/{sheet_id}/candidates").json()
        values = field_map(client, sheet_id)

    assert all(item["field_name"] == "discipline" for item in candidates if item["source_type"] == "cad_layer")
    assert values["discipline"]["final_source"] in {"cad_layer", "mixed"}
    for field_name in ["drawing_no", "drawing_name", "version", "issue_date"]:
        assert field_name not in values or values[field_name]["final_source"] != "cad_layer"


def test_dxf_excel_export_matches_database_and_does_not_mutate_state():
    with TestClient(app) as client:
        project_id = create_project(client, "DXF导出一致性")
        _, _, parse, _ = run_dxf_flow(client, project_id)
        sheet_id = parse["sheet_id"]
        before_sheet = client.get(f"/api/sheets/{sheet_id}").json()
        with SessionLocal() as db:
            before_issues = {
                issue.id: issue.status
                for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
            }
        export = export_project(client, project_id)
        after_sheet = client.get(f"/api/sheets/{sheet_id}").json()

    workbook = load_workbook(settings.root_dir / export["file_path"])
    ledger = workbook["图纸总台账"]
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        after_issues = {
            issue.id: issue.status
            for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
        }
    assert sheet is not None
    assert ledger.max_row - 1 == 1
    assert ledger.cell(2, 4).value == sheet.drawing_no
    assert ledger.cell(2, 5).value == sheet.drawing_name
    assert ledger.cell(2, 3).value == sheet.discipline
    assert before_sheet["status"] == after_sheet["status"]
    assert before_sheet["review_status"] == after_sheet["review_status"]
    assert before_issues == after_issues
