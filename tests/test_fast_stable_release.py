from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from scripts.build_portable_package import DEFAULT_VERSION, build_portable_package, package_name
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf
from test_project_backup_restore import backup_project, create_project, upload_files
from test_recognition_raw import _wait_for_ocr_job


VERSION = "v1.2.2-fast-import-fix"
ROOT = Path(__file__).resolve().parents[1]


def export_excel(client: TestClient, project_id: int) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200, response.text
    return response.json()


def candidate_count(sheet_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(
            select(func.count()).select_from(RecognitionCandidate).where(RecognitionCandidate.sheet_id == sheet_id)
        ) or 0


def issue_snapshot(sheet_id: int) -> list[tuple[str, str, str]]:
    with SessionLocal() as db:
        return sorted(
            (issue.issue_code, issue.severity, issue.status)
            for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
        )


def open_issue_count(sheet_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(
            select(func.count())
            .select_from(DrawingIssue)
            .where(DrawingIssue.sheet_id == sheet_id, DrawingIssue.status == "open")
        ) or 0


def sheet_review_status(sheet_id: int) -> str:
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        assert sheet is not None
        return sheet.review_status


def reviewed_field(sheet_id: int, field_name: str) -> FieldValue:
    with SessionLocal() as db:
        value = db.scalar(
            select(FieldValue).where(FieldValue.sheet_id == sheet_id, FieldValue.field_name == field_name)
        )
        assert value is not None
        db.expunge(value)
        return value


def test_v102_health_version_and_portable_package_are_ready():
    with TestClient(app) as client:
        health = client.get("/api/health")

    summary = build_portable_package(ROOT, version=VERSION, clean=True)
    package_dir = ROOT / "release" / package_name(VERSION)
    package_info = (package_dir / "package_info.txt").read_text(encoding="utf-8")

    assert health.status_code == 200
    assert health.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION
    assert DEFAULT_VERSION == VERSION
    assert summary.package_dir == package_dir
    assert summary.integrity_ok is True
    assert package_dir.name == f"工程图纸智能台账识别系统-{VERSION}"
    assert "包类型：Windows 本地便携正式稳定版" in package_info
    assert (package_dir / "backend").is_dir()
    assert (package_dir / "recognizer").is_dir()
    assert (package_dir / "frontend" / "dist" / "index.html").is_file()
    assert (package_dir / "scripts" / "local_launcher.py").is_file()
    assert (package_dir / "app_data" / "projects" / ".gitkeep").is_file()
    assert not (package_dir / ".git").exists()
    assert not (package_dir / "node_modules").exists()
    assert (ROOT / "docs" / "FAST_RELEASE_REPORT_v1.0.2.md").is_file()


def test_v102_pdf_minimal_flow_still_exports_excel():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.2 PDF 快速回归")
        upload = upload_files(client, project_id, [("v102.pdf", make_pdf_bytes("图号 建施-102"), "application/pdf")])
        batch_id = upload["id"]
        file_id = upload["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        title_crop = client.post(f"/api/imports/{batch_id}/title-crops")
        extract_text = client.post(f"/api/imports/{batch_id}/extract-text")
        ocr = client.post(f"/api/imports/{batch_id}/ocr-titles")
        _wait_for_ocr_job(client, batch_id)
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "PDF-人工-102", "drawing_name": "PDF 快速回归", "discipline": "建筑"}},
        )
        export = export_excel(client, project_id)

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert split.status_code == 200
    assert title_crop.status_code == 200
    assert extract_text.status_code == 200
    assert ocr.status_code == 200
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert update.status_code == 200
    assert export["ledger_row_count"] == 1
    assert workbook["图纸总台账"].cell(2, 4).value == "PDF-人工-102"


def test_v102_dxf_and_cad_preview_minimal_flow_still_exports_excel():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.2 DXF CAD 预览")
        upload = upload_files(client, project_id, [("v102.dxf", title_block_dxf("建施-102", "DXF 快速回归"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        preview = client.post(f"/api/sheets/{sheet_id}/cad-preview")
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)

    assert prepare.status_code == 200
    assert parse.status_code == 200
    assert parse.json()["status"] == "success"
    assert preview.status_code == 200
    assert preview.json()["status"] in {"success", "failed"}
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert candidate_count(sheet_id) > 0
    assert export["ledger_row_count"] == 1


def test_v102_dwg_mock_conversion_enters_dxf_flow(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.2 DWG mock")
        upload = upload_files(client, project_id, [("v102.dwg", DWG_BYTES, "application/acad")])
        file_id = upload["files"][0]["id"]
        parse_before_convert = client.post(f"/api/files/{file_id}/parse-dxf")
        create_converter_setting(client, converter)
        convert = client.post(f"/api/files/{file_id}/convert-dwg-to-dxf")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)

    assert parse_before_convert.status_code == 400
    assert parse_before_convert.json()["detail"]["error_code"] == "DWG_NOT_CONVERTED"
    assert convert.status_code == 200
    assert convert.json()["status"] == "success"
    assert parse.status_code == 200
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert export["ledger_row_count"] == 1


def test_v102_cad_pipeline_is_idempotent_and_preserves_manual_fields(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "v1.0.2 CAD pipeline")
        upload = upload_files(
            client,
            project_id,
            [
                ("pipeline.dxf", title_block_dxf("建施-102P", "Pipeline DXF"), "application/dxf"),
                ("pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        batch_id = upload["id"]
        first = run_cad_pipeline_blocking(
            client,
            batch_id,
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        dxf_sheet_id = next(item["id"] for item in sheets if item["source_format"] == "dxf")
        candidates_before = candidate_count(dxf_sheet_id)
        open_issues_before = open_issue_count(dxf_sheet_id)
        manual = client.patch(
            f"/api/sheets/{dxf_sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-FAST-102", "drawing_name": "人工 Pipeline", "discipline": "建筑"}},
        )
        rerun = run_cad_pipeline_blocking(
            client,
            batch_id,
            {"steps": ["generate_candidates", "fuse_fields"], "skip_completed": False, "continue_on_error": True},
        )
        detail = client.get(f"/api/sheets/{dxf_sheet_id}").json()

    assert first["summary"]["dwg_files"] == 1
    assert first["summary"]["dxf_files"] == 1
    assert first["summary"]["converted_success"] == 1
    assert first["summary"]["parse_success"] == 2
    assert manual.status_code == 200
    assert rerun["status"] in {"success", "completed_with_errors", "skipped", "failed"}
    assert detail["drawing_no"] == "人工-FAST-102"
    assert reviewed_field(dxf_sheet_id, "drawing_no").is_reviewed is True
    assert candidate_count(dxf_sheet_id) == candidates_before
    assert open_issue_count(dxf_sheet_id) <= open_issues_before


def test_v102_review_export_backup_restore_and_health_checks_are_stable():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.2 校核导出备份健康")
        upload = upload_files(client, project_id, [("review.dxf", title_block_dxf("建施-102R", "校核图"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-校核-102", "drawing_name": "人工校核图", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "快速回归确认"})
        before_review = sheet_review_status(sheet_id)
        before_issues = issue_snapshot(sheet_id)
        export = export_excel(client, project_id)
        after_review = sheet_review_status(sheet_id)
        after_issues = issue_snapshot(sheet_id)
        backup = backup_project(client, project_id)
        restore = client.post(f"/api/backups/{backup['backup_id']}/restore", json={"restore_mode": "new_project"})
        restored_id = restore.json()["new_project_id"]
        restored_project = client.get(f"/api/projects/{restored_id}")
        restored_export = export_excel(client, restored_id)
        system_health = client.get("/api/system/health-check")
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        backup_health = client.get("/api/backups/health-check")
        export_health = client.get("/api/exports/health-check")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert parse.status_code == 200
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert before_review == after_review
    assert before_issues == after_issues
    assert workbook["图纸总台账"].cell(2, 4).value == "人工-校核-102"
    assert export["ledger_row_count"] == 1
    assert restore.status_code == 200
    assert restored_id != project_id
    assert restored_project.status_code == 200
    assert restored_export["ledger_row_count"] == 1
    assert system_health.status_code == 200
    assert project_health.status_code == 200
    assert backup_health.status_code == 200
    assert export_health.status_code == 200


def test_v102_temp_cleanup_does_not_delete_projects_backups_or_exports():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.2 temp 清理安全")
        upload = upload_files(client, project_id, [("cleanup.dxf", title_block_dxf("建施-102C", "清理图"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/parse-dxf").status_code == 200
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)

        temp_file = settings.temp_dir / "v102-cleanup.tmp"
        temp_file.parent.mkdir(parents=True, exist_ok=True)
        temp_file.write_text("temp", encoding="utf-8")
        project_dir = settings.projects_dir / f"project_{project_id}"
        backup_path = settings.root_dir / backup["file_path"]
        export_path = settings.root_dir / export["file_path"]

        cleanup = client.post("/api/system/cleanup-temp")
        delete_backup = client.delete(f"/api/backups/{backup['backup_id']}")
        project_after_delete = client.get(f"/api/projects/{project_id}")

    assert cleanup.status_code == 200
    assert not temp_file.exists()
    assert project_dir.exists()
    assert export_path.exists()
    assert settings.projects_dir.exists()
    assert settings.backups_dir.exists()
    assert (settings.root_dir / "app_data" / "exports").exists() or export_path.exists()
    assert delete_backup.status_code == 204
    assert not backup_path.exists()
    assert project_after_delete.status_code == 200

def test_v115_deep_extract_smoke():
    """深度抽取烟囱：parse-dxf 后 block_stats + tables 自动落库，consistency-check 不报错。"""
    from dwg_test_helpers import dxf_with_insert_blocks

    with TestClient(app) as client:
        project_id = create_project(client, "v1.1.5 deep extract smoke")
        content = dxf_with_insert_blocks(
            [{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (10, 0)]}]
        )
        upload = upload_files(client, project_id, [("smoke.dxf", content, "application/dxf")])
        file_id = upload["files"][0]["id"]
        parse_resp = client.post(f"/api/files/{file_id}/parse-dxf")
        assert parse_resp.status_code == 200, parse_resp.text
        sheet_id = parse_resp.json()["sheet_id"]

        # 块统计自动触发
        stats_resp = client.get(f"/api/sheets/{sheet_id}/block-stats")
        assert stats_resp.status_code == 200

        # 跨图校验
        check_resp = client.post(f"/api/projects/{project_id}/consistency-check")
        assert check_resp.status_code == 200
        body = check_resp.json()
        assert body["project_id"] == project_id
