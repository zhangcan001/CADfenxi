from openpyxl import load_workbook
from sqlalchemy import select
from fastapi.testclient import TestClient

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.export_record import ExportRecord


def create_project_and_file(client: TestClient):
    project = client.post("/api/projects", json={"name": "导出测试项目"})
    assert project.status_code == 201
    project_id = project.json()["id"]
    upload = client.post(
        f"/api/projects/{project_id}/imports",
        data={"batch_name": "导出批次"},
        files=[("files", ("export.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"))],
    )
    assert upload.status_code == 201
    return project_id, upload.json()["id"], upload.json()["files"][0]["id"]


def add_sheet(project_id: int, batch_id: int, file_id: int, drawing_no: str | None, status: str = "recognized", review_status: str = "confirmed"):
    with SessionLocal() as db:
        current_count = db.query(DrawingSheet).filter(DrawingSheet.file_id == file_id).count()
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=current_count + 1,
            drawing_no=drawing_no,
            drawing_name="二层平面图",
            discipline="建筑",
            version="A",
            status=status,
            review_status=review_status,
            trust_level="A",
            confidence_score=95,
        )
        db.add(sheet)
        db.flush()
        sheet_id = sheet.id
        db.commit()
    return sheet_id


def add_issue(project_id: int, batch_id: int, file_id: int, sheet_id: int, severity: str = "error", status: str = "open"):
    with SessionLocal() as db:
        issue = DrawingIssue(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            sheet_id=sheet_id,
            issue_code="DRAWING_NO_EMPTY",
            severity=severity,
            message="测试问题",
            suggestion="请处理",
            status=status,
        )
        db.add(issue)
        db.flush()
        issue_id = issue.id
        db.commit()
    return issue_id


def test_export_check_empty_project():
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "空导出项目"}).json()
        response = client.post(f"/api/projects/{project['id']}/exports/check")

    assert response.status_code == 200
    assert response.json()["unconfirmed_count"] == 0


def test_export_check_counts_all_warnings():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project_and_file(client)
        sheet_1 = add_sheet(project_id, batch_id, file_id, "建施-03", review_status="unreviewed")
        sheet_2 = add_sheet(project_id, batch_id, file_id, None, status="failed", review_status="unreviewed")
        add_sheet(project_id, batch_id, file_id, "建施-03")
        add_issue(project_id, batch_id, file_id, sheet_1, "error", "open")
        response = client.post(f"/api/projects/{project_id}/exports/check")

    data = response.json()
    assert data["unconfirmed_count"] == 2
    assert data["open_error_count"] == 1
    assert data["failed_count"] == 1
    assert data["empty_drawing_no_count"] == 1
    assert data["duplicate_drawing_no_count"] == 2


def test_export_allows_incomplete_with_warnings():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project_and_file(client)
        add_sheet(project_id, batch_id, file_id, None, review_status="unreviewed")
        response = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": False, "include_issues": True, "filter": None},
        )

    assert response.status_code == 200
    assert response.json()["warning_summary"]["empty_drawing_no_count"] == 1
    assert response.json()["warning_count"] > 0


def test_export_excel_success_and_record_download():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project_and_file(client)
        sheet_id = add_sheet(project_id, batch_id, file_id, "建施-03", review_status="unreviewed")
        add_issue(project_id, batch_id, file_id, sheet_id, "warning", "open")
        before = client.get(f"/api/sheets/{sheet_id}").json()
        response = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        history = client.get(f"/api/projects/{project_id}/exports")
        download = client.get(response.json()["download_url"])
        after = client.get(f"/api/sheets/{sheet_id}").json()

    assert response.status_code == 200
    data = response.json()
    path = settings.root_dir / data["file_path"]
    assert path.exists()
    assert history.status_code == 200
    assert history.json()[0]["export_id"] == data["export_id"]
    assert download.status_code == 200
    assert before["status"] == after["status"]
    assert before["review_status"] == after["review_status"]
    with SessionLocal() as db:
        assert db.scalar(select(ExportRecord.id).where(ExportRecord.id == data["export_id"])) is not None


def test_excel_contains_expected_sheets_and_rows():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project_and_file(client)
        sheet_id = add_sheet(project_id, batch_id, file_id, "建施-03")
        add_issue(project_id, batch_id, file_id, sheet_id, "info", "ignored")
        response = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    workbook = load_workbook(settings.root_dir / response.json()["file_path"])
    assert "图纸总台账" in workbook.sheetnames
    assert "问题清单" in workbook.sheetnames
    assert "导出说明" in workbook.sheetnames
    assert workbook["图纸总台账"].max_row == 2
    assert workbook["问题清单"].max_row == 2
