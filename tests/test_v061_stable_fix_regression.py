from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue
from scripts.build_portable_package import build_portable_package, package_name
from scripts.local_launcher import check_startup_requirements
from test_full_flow_stability_v055 import (
    make_pdf_bytes,
    title_block_dxf,
    create_project,
    upload,
    export_excel,
    run_pipeline,
    counts,
    sheet_snapshot,
    client_values,
)
from dwg_test_helpers import DWG_BYTES, clear_converter_tables, create_converter_setting, write_mock_converter


VERSION = settings.version


def test_v061_health_version_portable_package_and_startup_diagnostics(tmp_path: Path):
    root = Path(__file__).resolve().parents[1]

    with TestClient(app) as client:
        health = client.get("/api/health")

    summary = build_portable_package(root, version=VERSION, clean=True)
    checks = check_startup_requirements(tmp_path)

    assert health.status_code == 200
    assert health.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION
    assert summary.package_dir == root / "release" / package_name(VERSION)
    assert summary.integrity_ok is True
    assert VERSION in summary.package_info_path.read_text(encoding="utf-8")
    assert any(check.error_code == "FRONTEND_DIST_NOT_FOUND" for check in checks)


def test_v061_pdf_minimal_flow_export_uses_manual_value_and_does_not_mutate_state():
    with TestClient(app) as client:
        project_id = create_project(client, "v061-PDF")
        upload_result = upload(client, project_id, [("建施-61_PDF修复.pdf", make_pdf_bytes(), "application/pdf")])
        batch_id = upload_result["id"]
        file_id = upload_result["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]

        assert client.post(f"/api/imports/{batch_id}/title-crops").status_code == 200
        assert client.post(f"/api/imports/{batch_id}/extract-text").status_code == 200
        assert client.post(f"/api/imports/{batch_id}/ocr-titles").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        manual = client.post(
            f"/api/review/sheets/{sheet_id}/update-fields",
            json={"fields": {"drawing_no": "PDF-人工-61", "drawing_name": "人工图名", "discipline": "建筑"}},
        )
        before = sheet_snapshot(sheet_id)
        export = export_excel(client, project_id)
        after = sheet_snapshot(sheet_id)

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert split.status_code == 200
    assert manual.status_code == 200
    assert workbook["图纸总台账"].cell(2, 4).value == "PDF-人工-61"
    assert before == after


def test_v061_dxf_flow_repeated_candidates_fusion_and_broken_dxf_error_are_stable():
    with TestClient(app) as client:
        project_id = create_project(client, "v061-DXF")
        upload_result = upload(client, project_id, [("建施-61_稳定修复.dxf", title_block_dxf(), "application/dxf")])
        file_id = upload_result["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/prepare-dxf-sheet").status_code == 200
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        first_counts = counts(sheet_id)
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        second_counts = counts(sheet_id)
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        third_counts = counts(sheet_id)
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        fourth_counts = counts(sheet_id)
        export = export_excel(client, project_id)
        broken = upload(client, project_id, [("broken-v061.dxf", b"not a dxf", "application/dxf")])
        broken_parse = client.post(f"/api/files/{broken['files'][0]['id']}/parse-dxf")

    assert parse.status_code == 200
    assert first_counts[0] == second_counts[0]
    assert third_counts[1] == fourth_counts[1]
    assert export["ledger_row_count"] == 1
    assert broken_parse.status_code == 200
    assert broken_parse.json()["status"] == "failed"
    assert broken_parse.json()["error_code"] in {"DXF_PARSE_FAILED", "DXF_OPEN_FAILED"}


def test_v061_dwg_mock_conversion_pipeline_and_unconfigured_converter_errors(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        no_config_project = create_project(client, "v061-DWG未配置")
        no_config_upload = upload(client, no_config_project, [("need-config-v061.dwg", DWG_BYTES, "application/acad")])
        no_config = client.post(f"/api/files/{no_config_upload['files'][0]['id']}/convert-dwg-to-dxf")

        create_converter_setting(client, converter)
        project_id = create_project(client, "v061-DWG")
        upload_result = upload(client, project_id, [("pipe-v061.dwg", DWG_BYTES, "application/acad")])
        batch_id = upload_result["id"]
        first = run_pipeline(client, batch_id)
        second = run_pipeline(client, batch_id)
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        export = export_excel(client, project_id)

    assert no_config.status_code == 400
    assert no_config.json()["detail"]["error_code"] == "CONVERTER_NOT_CONFIGURED"
    assert first["status"] == "success"
    assert first["summary"]["converted_success"] == 1
    assert first["summary"]["parse_success"] == 1
    assert second["status"] in {"skipped", "success"}
    assert len(sheets) == 1
    assert export["ledger_row_count"] == 1


def test_v061_review_batch_confirm_safety_and_manual_field_survives_cad_pipeline(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "v061-校核保护")
        upload_result = upload(client, project_id, [("manual-v061.dxf", title_block_dxf("建施-61"), "application/dxf")])
        batch_id = upload_result["id"]
        first = run_pipeline(client, batch_id, steps=["prepare_dxf_sheet", "parse_dxf", "generate_candidates", "fuse_fields"])
        sheet_id = first["steps"][1]["items"][0]["sheet_id"]
        update = client.post(
            f"/api/review/sheets/{sheet_id}/update-fields",
            json={"fields": {"drawing_no": "建施-61A", "drawing_name": "人工确认图名", "discipline": "建筑"}},
        )
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            db.add(
                DrawingIssue(
                    project_id=sheet.project_id,
                    batch_id=sheet.batch_id,
                    file_id=sheet.file_id,
                    sheet_id=sheet.id,
                    issue_code="OPEN_ERROR",
                    severity="error",
                    message="测试 open error",
                    suggestion="请先处理错误",
                    status="open",
                )
            )
            db.commit()
        batch_confirm = client.post(
            f"/api/projects/{project_id}/batch-confirm",
            json={"sheet_ids": [sheet_id], "confirm_mode": "selected", "only_without_errors": True},
        )
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        run_pipeline(client, batch_id, steps=["generate_candidates", "fuse_fields"], skip_completed=False)
        sheet_after = client.get(f"/api/sheets/{sheet_id}").json()

    assert update.status_code == 200
    assert batch_confirm.status_code == 200
    assert batch_confirm.json()["confirmed_count"] == 0
    assert batch_confirm.json()["skipped_count"] == 1
    assert sheet_after["drawing_no"] == "建施-61A"
    assert client_values(sheet_id)["drawing_no"]["is_reviewed"] is True


def test_v061_excel_export_does_not_change_review_or_issue_status_and_template_version():
    with TestClient(app) as client:
        project_id = create_project(client, "v061-导出一致性")
        upload_result = upload(client, project_id, [("export-v061.dxf", title_block_dxf("建施-62"), "application/dxf")])
        file_id = upload_result["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/prepare-dxf-sheet").status_code == 200
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        before = sheet_snapshot(sheet_id)
        export = export_excel(client, project_id)
        after = sheet_snapshot(sheet_id)

    workbook = load_workbook(settings.root_dir / export["file_path"])
    with SessionLocal() as db:
        field_count = db.scalar(select(func.count()).select_from(FieldValue).where(FieldValue.sheet_id == sheet_id)) or 0
    assert workbook["导出说明"]["B3"].value == VERSION
    assert workbook["图纸总台账"].max_row - 1 == export["ledger_row_count"]
    assert before["review_status"] == after["review_status"]
    assert before["issues"] == after["issues"]
    assert field_count >= 1
