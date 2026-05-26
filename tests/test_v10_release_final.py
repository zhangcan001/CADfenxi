import zipfile
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.export_record import ExportRecord
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from dwg_test_helpers import (
    DWG_BYTES,
    DXF_TEXT,
    clear_converter_tables,
    create_converter_setting,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from scripts.build_portable_package import DEFAULT_VERSION, build_portable_package, package_name
from test_cad_preview import prepare_dxf_sheet
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf
from test_project_backup_restore import backup_project, create_project, upload_files
from test_recognition_raw import _wait_for_ocr_job


VERSION = "v1.1.5-deep-extract"
ROOT = Path(__file__).resolve().parents[1]


def export_excel(client: TestClient, project_id: int) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200, response.text
    return response.json()


def sheet_snapshot(sheet_id: int) -> dict:
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        assert sheet is not None
        values = [
            (value.field_name, value.display_value, value.is_reviewed, value.final_source)
            for value in db.scalars(select(FieldValue).where(FieldValue.sheet_id == sheet_id)).all()
        ]
        issues = [
            (issue.issue_code, issue.status)
            for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
        ]
        return {
            "status": sheet.status,
            "review_status": sheet.review_status,
            "drawing_no": sheet.drawing_no,
            "values": sorted(values),
            "issues": sorted(issues),
        }


def field_value(sheet_id: int, field_name: str) -> FieldValue:
    with SessionLocal() as db:
        value = db.scalar(
            select(FieldValue).where(FieldValue.sheet_id == sheet_id, FieldValue.field_name == field_name)
        )
        assert value is not None
        db.expunge(value)
        return value


def candidate_count(sheet_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(
            select(func.count()).select_from(RecognitionCandidate).where(RecognitionCandidate.sheet_id == sheet_id)
        ) or 0


def open_issue_statuses(sheet_id: int) -> list[str]:
    with SessionLocal() as db:
        return [
            issue.status
            for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
        ]


def test_v10_version_release_docs_and_package_defaults():
    with TestClient(app) as client:
        health = client.get("/api/health")

    assert health.status_code == 200
    assert health.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION
    assert DEFAULT_VERSION == VERSION
    assert package_name() == f"工程图纸智能台账识别系统-{VERSION}"
    assert (ROOT / "docs" / "RELEASE_CHECKLIST_v1.0-local-stable.md").is_file()
    assert (ROOT / "docs" / "FINAL_ACCEPTANCE_v1.0-local-stable.md").is_file()
    assert (ROOT / "docs" / "RELEASE_CHECKLIST_v1.0-local-stable.md").is_file()
    assert "v1.0-local-stable 发布说明" in (ROOT / "RELEASE_NOTES.md").read_text(encoding="utf-8")


def test_v10_pdf_minimal_flow_and_export_state_are_stable():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 PDF 最终回归")
        upload = upload_files(client, project_id, [("v092.pdf", make_pdf_bytes("图号 建施-092"), "application/pdf")])
        batch_id = upload["id"]
        file_id = upload["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        title_crop = client.post(f"/api/imports/{batch_id}/title-crops")
        extract_text = client.post(f"/api/imports/{batch_id}/extract-text")
        ocr = client.post(f"/api/imports/{batch_id}/ocr-titles")
        _wait_for_ocr_job(client, batch_id)
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        update = client.post(
            f"/api/review/sheets/{sheet_id}/update-fields",
            json={"fields": {"drawing_no": "PDF-人工-092", "drawing_name": "PDF 人工图名", "discipline": "建筑"}},
        )
        before = sheet_snapshot(sheet_id)
        export = export_excel(client, project_id)
        after = sheet_snapshot(sheet_id)
        preview = client.get(f"/api/sheets/{sheet_id}/preview")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert split.status_code == 200, split.text
    assert title_crop.status_code == 200, title_crop.text
    assert extract_text.status_code == 200, extract_text.text
    assert ocr.status_code == 200, ocr.text
    assert candidates.status_code == 200, candidates.text
    assert fuse.status_code == 200, fuse.text
    assert update.status_code == 200, update.text
    assert export["ledger_row_count"] == 1
    assert workbook["图纸总台账"].cell(2, 4).value == "PDF-人工-092"
    assert before == after
    assert preview.status_code == 200
    assert preview.headers["content-type"].startswith("image/png")


def test_v10_dxf_minimal_flow_cad_json_candidates_and_excel_are_stable():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 DXF 最终回归")
        upload = upload_files(client, project_id, [("v092.dxf", title_block_dxf("建施-092", "DXF 回归"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        summary = client.get(f"/api/sheets/{sheet_id}/cad-parse")
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)

    assert prepare.status_code == 200, prepare.text
    assert parse.status_code == 200, parse.text
    assert parse.json()["status"] == "success"
    assert (settings.root_dir / parse.json()["output_path"]).is_file()
    assert summary.status_code == 200, summary.text
    assert summary.json()["counts"]["text_count"] + summary.json()["counts"]["mtext_count"] + summary.json()["counts"]["attrib_count"] > 0
    assert candidates.status_code == 200, candidates.text
    assert fuse.status_code == 200, fuse.text
    assert candidate_count(sheet_id) > 0
    assert export["ledger_row_count"] == 1


def test_v10_dwg_mock_conversion_flow_does_not_directly_parse_dwg(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 DWG 最终回归")
        upload = upload_files(client, project_id, [("v092.dwg", DWG_BYTES, "application/acad")])
        file_id = upload["files"][0]["id"]
        parse_before_convert = client.post(f"/api/files/{file_id}/parse-dxf")
        create_converter_setting(client, converter)
        convert = client.post(f"/api/files/{file_id}/convert-dwg-to-dxf")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)
        runs = client.get(f"/api/files/{file_id}/cad-conversion-runs")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert parse_before_convert.status_code == 400
    assert parse_before_convert.json()["detail"]["error_code"] == "DWG_NOT_CONVERTED"
    assert convert.status_code == 200, convert.text
    assert convert.json()["status"] == "success"
    assert (settings.root_dir / convert.json()["converted_file_path"]).is_file()
    assert parse.status_code == 200, parse.text
    assert candidates.status_code == 200, candidates.text
    assert fuse.status_code == 200, fuse.text
    assert runs.status_code == 200
    assert runs.json()[0]["status"] == "success"
    assert workbook["图纸总台账"].cell(2, 8).value == "DWG转换"


def test_v10_cad_pipeline_review_workbench_and_manual_field_protection(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "v1.0 pipeline 最终回归")
        upload = upload_files(
            client,
            project_id,
            [
                ("pipeline.pdf", make_pdf_bytes("PDF 不应被 CAD pipeline 误处理"), "application/pdf"),
                ("pipeline.dxf", DXF_TEXT.encode("utf-8"), "application/dxf"),
                ("pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        batch_id = upload["id"]
        pdf_file_id = next(item["id"] for item in upload["files"] if item["source_format"] == "pdf")
        assert client.post(f"/api/files/{pdf_file_id}/split").status_code == 200
        pipeline = run_cad_pipeline_blocking(
            client,
            batch_id,
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        dxf_sheet_id = next(item["id"] for item in sheets if item["source_format"] == "dxf")
        update = client.patch(
            f"/api/sheets/{dxf_sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-092", "drawing_name": "人工保护图名", "discipline": "建筑"}},
        )
        blocked_confirm = client.post(
            f"/api/sheets/{dxf_sheet_id}/confirm",
            json={"force": False, "note": "v1.0 open error 拦截"},
        )
        confirm = client.post(f"/api/sheets/{dxf_sheet_id}/confirm", json={"force": True, "note": "v1.0 确认"})
        rerun = run_cad_pipeline_blocking(
            client,
            batch_id,
            {"steps": ["generate_candidates", "fuse_fields"], "skip_completed": False, "continue_on_error": True},
        )
        after = client.get(f"/api/sheets/{dxf_sheet_id}").json()
        batch_confirm = client.post(
            f"/api/projects/{project_id}/batch-confirm",
            json={"sheet_ids": [dxf_sheet_id], "note": "最终回归批量确认"},
        )
        export = export_excel(client, project_id)

    data = pipeline
    assert data["summary"]["pdf_files"] == 1
    assert data["summary"]["dxf_files"] == 1
    assert data["summary"]["dwg_files"] == 1
    assert data["summary"]["converted_success"] == 1
    assert data["summary"]["parse_success"] == 2
    assert data["summary"]["cad_preview_success"] >= 1
    assert update.status_code == 200, update.text
    assert blocked_confirm.status_code == 400
    assert confirm.status_code == 200, confirm.text
    assert rerun["status"] in {"success", "completed_with_errors", "skipped", "failed"}
    assert after["drawing_no"] == "人工-092"
    assert after["review_status"] == "confirmed"
    assert field_value(dxf_sheet_id, "drawing_no").is_reviewed is True
    assert batch_confirm.status_code == 200, batch_confirm.text
    assert export["ledger_row_count"] == len(sheets)
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(DrawingSheet).where(DrawingSheet.file_id == pdf_file_id)) == 1


def test_v10_export_does_not_change_review_or_issue_status():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 导出一致性")
        file_id, sheet_id, _batch_id = prepare_dxf_sheet(client, project_id, "export-state.dxf")
        assert client.post(f"/api/files/{file_id}/parse-dxf").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            assert sheet is not None
            db.add(
                DrawingIssue(
                    project_id=sheet.project_id,
                    batch_id=sheet.batch_id,
                    file_id=sheet.file_id,
                    sheet_id=sheet.id,
                    issue_code="LOW_CONFIDENCE_NEED_REVIEW",
                    severity="warning",
                    message="最终回归问题",
                    suggestion="人工复核",
                    status="open",
                )
            )
            db.commit()
        before = sheet_snapshot(sheet_id)
        issue_statuses_before = open_issue_statuses(sheet_id)
        export = export_excel(client, project_id)
        after = sheet_snapshot(sheet_id)
        issue_statuses_after = open_issue_statuses(sheet_id)

    assert export["ledger_row_count"] == 1
    assert before == after
    assert issue_statuses_before == issue_statuses_after


def test_v10_backup_restore_verify_delete_and_original_project_are_safe():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 备份恢复最终回归")
        upload = upload_files(client, project_id, [("backup.dxf", title_block_dxf("建施-092B", "备份恢复"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/parse-dxf").status_code == 200
        sheet_id = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"][0]["id"]
        assert client.post(f"/api/sheets/{sheet_id}/cad-preview").status_code == 200
        backup = backup_project(client, project_id)
        with zipfile.ZipFile(settings.root_dir / backup["file_path"]) as archive:
            names = archive.namelist()
        verify = client.get(f"/api/backups/{backup['backup_id']}/verify")
        download = client.get(backup["download_url"])
        restore = client.post(f"/api/backups/{backup['backup_id']}/restore", json={"restore_mode": "new_project"})
        restored_project_id = restore.json()["new_project_id"]
        restored_sheets = client.get(f"/api/projects/{restored_project_id}/sheets?page_size=100").json()["items"]
        restored_image = client.get(f"/api/sheets/{restored_sheets[0]['id']}/cad-preview-image")
        restored_export = export_excel(client, restored_project_id)
        delete_backup = client.delete(f"/api/backups/{backup['backup_id']}")
        original_project = client.get(f"/api/projects/{project_id}")
        restored_project = client.get(f"/api/projects/{restored_project_id}")

    assert verify.status_code == 200, verify.text
    assert verify.json()["valid"] is True
    assert download.status_code == 200
    assert download.content.startswith(b"PK")
    assert any(name.startswith("files/cad/previews/") for name in names)
    assert restore.status_code == 200, restore.text
    assert restored_project_id != project_id
    assert f"project_{restored_project_id}" in restored_sheets[0]["cad_preview_path"]
    assert f"project_{project_id}" not in restored_sheets[0]["cad_preview_path"]
    assert restored_image.status_code == 200
    assert restored_image.content.startswith(b"\x89PNG")
    assert restored_export["ledger_row_count"] == 1
    assert delete_backup.status_code == 204
    assert original_project.status_code == 200
    assert restored_project.status_code == 200


def test_v10_cad_preview_skip_force_and_failed_preview_do_not_block_export():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 CAD 预览最终回归")
        file_id, sheet_id, batch_id = prepare_dxf_sheet(client, project_id, "preview-good.dxf")
        first = client.post(
            f"/api/imports/{batch_id}/cad-preview",
            json={"skip_completed": True, "force": False, "continue_on_error": True},
        )
        skipped = client.post(
            f"/api/imports/{batch_id}/cad-preview",
            json={"skip_completed": True, "force": False, "continue_on_error": True},
        )
        forced = client.post(
            f"/api/imports/{batch_id}/cad-preview",
            json={"skip_completed": True, "force": True, "continue_on_error": True},
        )
        image = client.get(f"/api/sheets/{sheet_id}/cad-preview-image")

        bad_upload = upload_files(client, project_id, [("preview-bad.dxf", b"broken dxf", "application/dxf")])
        bad_file_id = bad_upload["files"][0]["id"]
        bad_prepare = client.post(f"/api/files/{bad_file_id}/prepare-dxf-sheet")
        bad_sheet_id = bad_prepare.json()["sheet_id"]
        failed_preview = client.post(f"/api/sheets/{bad_sheet_id}/cad-preview")
        good_parse = client.post(f"/api/files/{file_id}/parse-dxf")
        good_candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        good_fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)

    assert first.status_code == 200, first.text
    assert first.json()["summary"]["success_count"] == 1
    assert skipped.status_code == 200, skipped.text
    assert skipped.json()["summary"]["skipped_count"] == 1
    assert forced.status_code == 200, forced.text
    assert forced.json()["summary"]["success_count"] == 1
    assert image.status_code == 200
    assert image.headers["content-type"].startswith("image/png")
    assert image.content.startswith(b"\x89PNG")
    assert failed_preview.status_code == 200
    assert failed_preview.json()["status"] == "failed"
    assert good_parse.status_code == 200
    assert good_candidates.status_code == 200
    assert good_fuse.status_code == 200
    assert export["ledger_row_count"] >= 1


def test_v10_data_health_and_temp_cleanup_are_diagnostic_and_safe():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0 健康检查最终回归")
        upload = upload_files(client, project_id, [("health.pdf", make_pdf_bytes("健康检查"), "application/pdf")])
        file_id = upload["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/split").status_code == 200
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)

        temp_file = settings.temp_dir / "v092-cleanup.tmp"
        temp_file.parent.mkdir(parents=True, exist_ok=True)
        temp_file.write_text("temp", encoding="utf-8")

        project_dir = settings.projects_dir / f"project_{project_id}"
        backup_path = settings.root_dir / backup["file_path"]
        export_path = settings.root_dir / export["file_path"]

        health = client.get("/api/system/health-check")
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        backup_health = client.get("/api/backups/health-check")
        export_health = client.get("/api/exports/health-check")
        orphan_scan = client.get(f"/api/projects/{project_id}/orphan-files")
        cleanup = client.post("/api/system/cleanup-temp")
        project_after_cleanup = client.get(f"/api/projects/{project_id}")

    assert health.status_code == 200, health.text
    assert "grouped_summary" in health.json()
    assert project_health.status_code == 200, project_health.text
    assert backup_health.status_code == 200, backup_health.text
    assert export_health.status_code == 200, export_health.text
    assert orphan_scan.status_code == 200, orphan_scan.text
    assert cleanup.status_code == 200, cleanup.text
    assert cleanup.json()["deleted_file_count"] >= 1
    assert not temp_file.exists()
    assert project_dir.exists()
    assert backup_path.exists()
    assert export_path.exists()
    assert project_after_cleanup.status_code == 200


def test_v10_portable_package_contains_required_release_materials_and_clean_app_data():
    summary = build_portable_package(ROOT, version=VERSION, clean=True)
    package_dir = ROOT / "release" / package_name(VERSION)
    package_info = (package_dir / "package_info.txt").read_text(encoding="utf-8")

    assert summary.package_dir == package_dir
    assert summary.integrity_ok is True
    assert (package_dir / "backend").is_dir()
    assert (package_dir / "recognizer").is_dir()
    assert (package_dir / "frontend" / "dist" / "index.html").is_file()
    assert (package_dir / "scripts").is_dir()
    assert (package_dir / "docs" / "RELEASE_CHECKLIST_v1.0-local-stable.md").is_file()
    assert (package_dir / "docs" / "FINAL_ACCEPTANCE_v1.0-local-stable.md").is_file()
    assert (package_dir / "README.md").is_file()
    assert (package_dir / "README_本地使用说明.md").is_file()
    assert (package_dir / "RELEASE_NOTES.md").is_file()
    assert (package_dir / "requirements.txt").is_file()
    assert VERSION in package_info
    assert "包类型：Windows 本地便携正式稳定版" in package_info
    assert "备份目录：app_data/backups/" in package_info
    assert "适用场景：个人本地工程图纸台账识别" in package_info
    assert not (package_dir / ".git").exists()
    assert not (package_dir / "frontend" / "node_modules").exists()
    assert not (package_dir / "app_data" / "database" / "app.db").exists()
    for dirname in ["projects", "backups", "database", "logs", "temp"]:
        entries = sorted(item.name for item in (package_dir / "app_data" / dirname).iterdir())
        assert entries == [".gitkeep"]
