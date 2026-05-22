from datetime import date
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue


def create_project(client: TestClient):
    project = client.post("/api/projects", json={"name": f"交付复核项目-{uuid4().hex[:8]}"})
    assert project.status_code == 201
    project_id = project.json()["id"]
    upload = client.post(
        f"/api/projects/{project_id}/imports",
        data={"batch_name": "v0.5.4 交付复核批次"},
        files=[("files", ("delivery.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"))],
    )
    assert upload.status_code == 201
    return project_id, upload.json()["id"], upload.json()["files"][0]["id"]


def add_file(project_id: int, batch_id: int, name: str, source_format: str) -> int:
    with SessionLocal() as db:
        drawing_file = DrawingFile(
            project_id=project_id,
            batch_id=batch_id,
            original_name=name,
            file_ext=f".{source_format}",
            source_format=source_format,
            file_size=10,
            file_hash=uuid4().hex,
            storage_path=f"app_data/temp/{uuid4().hex}_{name}",
            status="preprocessed",
            convert_status="success" if source_format == "dwg" else None,
            converted_format="dxf" if source_format == "dwg" else None,
            converted_file_path=f"app_data/temp/{uuid4().hex}.dxf" if source_format == "dwg" else None,
        )
        db.add(drawing_file)
        db.flush()
        file_id = drawing_file.id
        db.commit()
    return file_id


def add_sheet(
    project_id: int,
    batch_id: int,
    file_id: int,
    drawing_no: str | None,
    drawing_name: str | None,
    discipline: str | None,
    source: str,
    trust_level: str = "A",
    review_status: str = "confirmed",
) -> int:
    with SessionLocal() as db:
        page_no = db.query(DrawingSheet).filter(DrawingSheet.file_id == file_id).count() + 1
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=page_no,
            sheet_type="unknown",
            drawing_no=drawing_no,
            drawing_name=drawing_name,
            discipline=discipline,
            version="A",
            issue_date=date(2026, 5, 21),
            status="confirmed" if review_status == "confirmed" else "need_review",
            review_status=review_status,
            trust_level=trust_level,
            confidence_score=96 if trust_level == "A" else 42,
        )
        db.add(sheet)
        db.flush()
        sheet_id = sheet.id
        for field_name, value in {
            "drawing_no": drawing_no,
            "drawing_name": drawing_name,
            "discipline": discipline,
            "version": "A",
            "issue_date": "2026-05-21",
        }.items():
            if value:
                db.add(
                    FieldValue(
                        project_id=project_id,
                        batch_id=batch_id,
                        file_id=file_id,
                        sheet_id=sheet_id,
                        field_name=field_name,
                        raw_value=value,
                        normalized_value=value,
                        display_value=value,
                        final_source=source,
                        confidence=96,
                        is_reviewed=review_status == "confirmed",
                    )
                )
        db.commit()
    return sheet_id


def add_issue(project_id: int, batch_id: int, file_id: int, sheet_id: int, code: str, severity: str = "warning") -> int:
    with SessionLocal() as db:
        issue = DrawingIssue(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            sheet_id=sheet_id,
            issue_code=code,
            severity=severity,
            message="图纸信息需要复核",
            suggestion="请结合图纸标题栏和候选值确认。",
            status="open",
        )
        db.add(issue)
        db.flush()
        issue_id = issue.id
        db.commit()
    return issue_id


def export_workbook(client: TestClient, project_id: int):
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200
    return response.json(), load_workbook(settings.root_dir / response.json()["file_path"])


def test_excel_delivery_template_review_and_sources():
    with TestClient(app) as client:
        project_id, batch_id, pdf_file = create_project(client)
        dxf_file = add_file(project_id, batch_id, "delivery.dxf", "dxf")
        dwg_file = add_file(project_id, batch_id, "delivery.dwg", "dwg")
        pdf_sheet = add_sheet(project_id, batch_id, pdf_file, "人工-01", "人工确认图名", "建筑", "manual")
        dxf_sheet = add_sheet(project_id, batch_id, dxf_file, "电施-01", "电气平面图", "电气", "cad_block_attr")
        dwg_sheet = add_sheet(project_id, batch_id, dwg_file, "水施-01", "给排水平面图", "给排水", "cad_text", "D", "unreviewed")
        add_issue(project_id, batch_id, dwg_file, dwg_sheet, "DRAWING_NO_EMPTY", "warning")
        before = state_snapshot([pdf_sheet, dxf_sheet, dwg_sheet])
        result, workbook = export_workbook(client, project_id)
        after = state_snapshot([pdf_sheet, dxf_sheet, dwg_sheet])

    ledger = workbook["图纸总台账"]
    issues = workbook["问题清单"]
    expected_headers = [
        "序号", "项目名称", "专业", "图纸编号", "图纸名称", "版本", "出图日期",
        "文件格式", "原始文件名", "来源文件", "识别来源", "可信等级", "识别评分",
        "校核状态", "确认时间", "问题数量", "错误数量", "警告数量", "备注",
    ]
    assert [cell.value for cell in ledger[1]] == expected_headers
    assert "人工-01" in [ledger.cell(row=row, column=4).value for row in range(2, ledger.max_row + 1)]
    assert "人工确认图名" in [ledger.cell(row=row, column=5).value for row in range(2, ledger.max_row + 1)]
    assert {"PDF", "DXF", "DWG转换"} == {ledger.cell(row=row, column=8).value for row in range(2, ledger.max_row + 1)}
    assert issues.cell(row=2, column=5).value == "警告"
    assert "图纸编号缺失" in issues.cell(row=2, column=6).value
    assert issues.cell(row=2, column=8).value == "请结合图纸标题栏和候选值确认。"
    assert result["summary_sheet_count"] == 2
    assert result["precheck"]["total_sheets"] == 3
    assert before == after


def test_excel_delivery_summary_info_styles_and_precheck():
    with TestClient(app) as client:
        project_id, batch_id, pdf_file = create_project(client)
        s1 = add_sheet(project_id, batch_id, pdf_file, "建施-01", "建筑平面图", "建筑", "pdf_text")
        s2 = add_sheet(project_id, batch_id, pdf_file, None, "结构平面图", "结构", "title_ocr", "D", "unreviewed")
        add_issue(project_id, batch_id, pdf_file, s2, "OCR_TEXT_EMPTY", "error")
        check = client.post(f"/api/projects/{project_id}/exports/check").json()
        result, workbook = export_workbook(client, project_id)

    assert check["can_export"] is True
    assert check["sheet_count"] == 2
    assert check["unconfirmed_count"] == 1
    assert check["empty_drawing_no_count"] == 1
    assert check["open_error_count"] == 1
    assert check["trust_level_d_count"] == 1
    assert "当前项目共有 2 张图纸" in check["summary_message"]
    assert result["warning_count"] > 0

    discipline_rows = {
        workbook["专业汇总"].cell(row=row, column=1).value: workbook["专业汇总"].cell(row=row, column=2).value
        for row in range(2, workbook["专业汇总"].max_row + 1)
    }
    assert discipline_rows["建筑"] == 1
    assert discipline_rows["结构"] == 1
    review_summary = {
        workbook["校核状态汇总"].cell(row=row, column=1).value: workbook["校核状态汇总"].cell(row=row, column=2).value
        for row in range(2, workbook["校核状态汇总"].max_row + 1)
    }
    assert review_summary["图纸总数"] == 2
    assert review_summary["未校核图纸"] == 1
    assert review_summary["D 级图纸"] == 1
    info_values = [workbook["导出说明"].cell(row=row, column=2).value for row in range(2, workbook["导出说明"].max_row + 1)]
    assert any("人工确认值优先于机器识别值" in str(value) for value in info_values)
    assert any("系统不直接解析 DWG" in str(value) for value in info_values)
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None
    assert workbook["图纸总台账"].cell(row=2, column=7).number_format == "yyyy-mm-dd"
    assert workbook["图纸总台账"].column_dimensions["D"].width >= 18
    assert workbook["图纸总台账"].column_dimensions["E"].width >= 32
    assert s1


def test_excel_delivery_empty_project_is_rejected_clearly():
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": f"空交付复核-{uuid4().hex[:8]}"}).json()
        check = client.post(f"/api/projects/{project['id']}/exports/check")
        export = client.post(
            f"/api/projects/{project['id']}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert check.status_code == 200
    assert check.json()["can_export"] is False
    assert check.json()["summary_message"] == "当前项目没有图纸，无法导出 Excel 台账。"
    assert export.status_code == 409
    assert "当前项目没有图纸" in export.json()["detail"]["summary_message"]


def state_snapshot(sheet_ids: list[int]) -> dict[int, dict]:
    with SessionLocal() as db:
        snapshot = {}
        for sheet_id in sheet_ids:
            sheet = db.get(DrawingSheet, sheet_id)
            issue_statuses = [
                issue.status
                for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
            ]
            fields = [
                (value.field_name, value.display_value, value.is_reviewed)
                for value in db.scalars(select(FieldValue).where(FieldValue.sheet_id == sheet_id)).all()
            ]
            snapshot[sheet_id] = {
                "review_status": sheet.review_status,
                "status": sheet.status,
                "issues": issue_statuses,
                "fields": fields,
            }
        return snapshot
