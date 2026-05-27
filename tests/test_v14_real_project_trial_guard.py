"""v1.4 真实项目整体验收试用守护测试。"""
from __future__ import annotations

from pathlib import Path

import ezdxf
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_block_stat import DrawingBlockStat
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.drawing_table import DrawingTable
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    equipment_schedule_dxf,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from test_full_flow_stability_v055 import make_pdf_bytes
from test_project_backup_restore import backup_project, create_project, upload_files
from test_v11_fast_ux import export_excel


VERSION = "v1.4-fast-real-project-trial"


def _count(model: type, project_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(select(func.count()).select_from(model).where(model.project_id == project_id)) or 0


def _sheet_field_state(project_id: int) -> list[tuple[int, str | None, str | None, str | None]]:
    with SessionLocal() as db:
        sheets = db.scalars(select(DrawingSheet).where(DrawingSheet.project_id == project_id).order_by(DrawingSheet.id)).all()
        return [(sheet.id, sheet.drawing_no, sheet.drawing_name, sheet.review_status) for sheet in sheets]


def _trial_block_dxf() -> bytes:
    doc = ezdxf.new("R2010")
    msp = doc.modelspace()
    msp.add_text("图号 E-201 图名 灯具布置图 专业 电气", dxfattribs={"layer": "TITLE", "insert": (0, 30, 0)})
    msp.add_line((0, 0), (100, 0), dxfattribs={"layer": "EE-LIGHT"})
    for block_name in ["LAMP", "A1_FRAME"]:
        block = doc.blocks.new(name=block_name)
        block.add_line((0, 0), (1, 0))
    for x in [0, 10, 20]:
        msp.add_blockref("LAMP", (x, 0), dxfattribs={"layer": "EE-LIGHT"})
    msp.add_blockref("A1_FRAME", (0, 50), dxfattribs={"layer": "0"})
    import io

    stream = io.StringIO()
    doc.write(stream)
    return stream.getvalue().encode("utf-8")


def _run_realistic_trial_project(client: TestClient, tmp_path: Path) -> dict:
    clear_converter_tables()
    create_converter_setting(client, write_mock_converter(tmp_path))
    project_id = create_project(client, "v1.4 真实项目试用")
    upload = upload_files(
        client,
        project_id,
        [
            ("A-101_总平面.pdf", make_pdf_bytes("图号 A-101\n图名 总平面图\n专业 建筑"), "application/pdf"),
            (
                "A-101_设备表.dxf",
                equipment_schedule_dxf(
                    rows=[["送风机", "FJ-1", "2", "备用"], ["排风机", "FJ-2", "3", ""]],
                    header=["设备名", "型号", "数量", "备注"],
                ),
                "application/dxf",
            ),
            (
                "E-201_灯具布置.dxf",
                _trial_block_dxf(),
                "application/dxf",
            ),
            ("S-301_基础.dwg", DWG_BYTES, "application/acad"),
        ],
    )
    pdf_file_id = next(item["id"] for item in upload["files"] if item["source_format"] == "pdf")
    pdf_split = client.post(f"/api/files/{pdf_file_id}/split")
    pipeline = run_cad_pipeline_blocking(
        client,
        upload["id"],
        {
            "steps": [
                "convert_dwg",
                "prepare_dxf_sheet",
                "parse_dxf",
                "generate_candidates",
                "fuse_fields",
                "generate_cad_preview",
            ],
            "skip_completed": True,
            "continue_on_error": True,
        },
    )
    sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
    pdf_sheet = next(item for item in sheets if item["source_format"] == "pdf")
    table_sheet = next(item for item in sheets if item["source_format"] == "dxf")
    for sheet in sheets:
        fields = {
            "drawing_no": "A-101" if sheet["id"] in {pdf_sheet["id"], table_sheet["id"]} else f"试用-{sheet['id']}",
            "drawing_name": "总平面图" if sheet["id"] == pdf_sheet["id"] else ("设备表" if sheet["id"] == table_sheet["id"] else "真实试用图纸"),
            "discipline": "建筑",
        }
        assert client.patch(f"/api/sheets/{sheet['id']}/fields", json={"fields": fields, "note": "v1.4 试用"}).status_code == 200
        assert client.post(f"/api/sheets/{sheet['id']}/confirm", json={"force": True, "note": "v1.4 试用确认"}).status_code == 200
    consistency = client.post(f"/api/projects/{project_id}/consistency-check")
    before_state = _sheet_field_state(project_id)
    export = export_excel(client, project_id)
    after_state = _sheet_field_state(project_id)
    backup = backup_project(client, project_id)
    restore = client.post(f"/api/backups/{backup['backup_id']}/restore", json={"restore_mode": "new_project"})
    restored_project_id = restore.json()["new_project_id"]
    restored_export = export_excel(client, restored_project_id)
    project_health = client.get(f"/api/projects/{project_id}/health-check")
    restored_health = client.get(f"/api/projects/{restored_project_id}/health-check")
    system_health = client.get("/api/system/health-check")

    return {
        "project_id": project_id,
        "restored_project_id": restored_project_id,
        "upload": upload,
        "pdf_split": pdf_split,
        "pipeline": pipeline,
        "sheets": sheets,
        "consistency": consistency,
        "before_state": before_state,
        "after_state": after_state,
        "export": export,
        "restored_export": restored_export,
        "backup": backup,
        "restore": restore,
        "project_health": project_health,
        "restored_health": restored_health,
        "system_health": system_health,
    }


def test_v14_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v14_realistic_project_trial_flow_excel_backup_and_health(tmp_path: Path):
    with TestClient(app) as client:
        result = _run_realistic_trial_project(client, tmp_path)

    assert result["upload"]["file_type_counts"] == {"pdf": 1, "dxf": 2, "dwg": 1, "unsupported": 0}
    assert result["pdf_split"].status_code == 200
    assert len(result["pdf_split"].json()["sheets"]) == 1
    assert result["pipeline"]["status"] == "success"
    assert result["pipeline"]["summary"]["dwg_files"] == 1
    assert result["pipeline"]["summary"]["converted_success"] == 1
    assert result["pipeline"]["summary"]["parse_success"] >= 3
    assert result["consistency"].status_code == 200
    assert result["consistency"].json()["by_code"].get("CROSS_DRAWING_NAME_CONFLICT", 0) >= 1
    assert result["before_state"] == result["after_state"]
    assert result["backup"]["file_size"] > 0
    assert result["restore"].status_code == 200
    assert result["project_health"].status_code == 200
    assert result["restored_health"].status_code == 200
    assert result["system_health"].status_code == 200
    assert _count(DrawingSheet, result["project_id"]) >= 4
    assert _count(DrawingTable, result["project_id"]) >= 1
    assert _count(DrawingBlockStat, result["project_id"]) >= 1
    assert _count(DrawingIssue, result["project_id"]) >= 1
    assert _count(DrawingSheet, result["restored_project_id"]) == _count(DrawingSheet, result["project_id"])

    workbook = load_workbook(settings.root_dir / result["export"]["file_path"], read_only=True)
    restored_workbook = load_workbook(settings.root_dir / result["restored_export"]["file_path"], read_only=True)
    try:
        expected_sheets = {
            "图纸总台账",
            "问题清单",
            "专业汇总",
            "校核状态汇总",
            "导出说明",
            "图纸表格明细",
            "图纸块统计",
        }
        assert expected_sheets.issubset(set(workbook.sheetnames))
        assert workbook["图纸总台账"].max_row - 1 == _count(DrawingSheet, result["project_id"])
        assert restored_workbook["图纸总台账"].max_row == workbook["图纸总台账"].max_row
        ledger_headers = [cell.value for cell in workbook["图纸总台账"][1]]
        assert "表头(JSON)" not in ledger_headers
        assert "数据(JSON)" not in ledger_headers
        assert "块名" not in ledger_headers
        assert "数量" not in ledger_headers
        table_text = "\n".join(str(cell) for row in workbook["图纸表格明细"].iter_rows(values_only=True) for cell in row if cell)
        block_text = "\n".join(str(cell) for row in workbook["图纸块统计"].iter_rows(values_only=True) for cell in row if cell)
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        assert "送风机" in table_text
        assert "LAMP" in block_text
        assert "A1_FRAME" not in block_text
        assert "同图号图名不一致" in issue_text
    finally:
        workbook.close()
        restored_workbook.close()


def test_v14_trial_report_exists_and_records_guard_scope():
    report = settings.root_dir / "docs" / "REAL_PROJECT_TRIAL_REPORT_v1.4.md"
    content = report.read_text(encoding="utf-8")

    assert report.is_file()
    assert "v1.4-fast-real-project-trial" in content
    assert "导入" in content
    assert "Excel 导出" in content
    assert "备份恢复" in content
    assert "健康检查" in content
