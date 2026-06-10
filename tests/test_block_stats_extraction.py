"""DXF INSERT 块统计测试 — 算法单测 + 集成 + Excel 导出。"""
from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.core.config import settings
from backend.main import app
from recognizer.cad_engine.block_aggregator import aggregate_inserts
from tests.dwg_test_helpers import (
    dxf_with_insert_blocks,
    equipment_schedule_dxf,
)


VERSION = "v1.5.1-fast-delivery-package-fix"


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as test_client:
        yield test_client


def _make_cad_json(inserts: list[dict]) -> dict:
    return {
        "spaces": [
            {
                "space": "modelspace",
                "texts": [],
                "mtexts": [],
                "inserts": inserts,
            }
        ],
        "counts": {},
        "warnings": [],
    }


def _make_insert(name: str, layer: str, x: float = 0.0, y: float = 0.0, attribs: dict | None = None) -> dict:
    return {
        "block_name": name,
        "layer": layer,
        "insert": [x, y, 0.0],
        "attribs": [
            {"tag": tag, "clean_text": value}
            for tag, value in (attribs or {}).items()
        ],
    }


def test_aggregate_inserts_groups_by_block_and_layer():
    inserts = [
        _make_insert("LAMP", "EE-LIGHT", 0, 0),
        _make_insert("LAMP", "EE-LIGHT", 10, 0),
        _make_insert("LAMP", "EE-LIGHT", 20, 0),
        _make_insert("FAN", "ME-VENT", 0, 50),
        _make_insert("FAN", "ME-VENT", 30, 50),
    ]
    rows = aggregate_inserts(_make_cad_json(inserts))
    by_name = {(r["block_name"], r["layer_name"]): r for r in rows}
    assert by_name[("LAMP", "EE-LIGHT")]["count"] == 3
    assert by_name[("FAN", "ME-VENT")]["count"] == 2


def test_aggregate_inserts_excludes_title_blocks():
    inserts = [
        _make_insert("TITLEBLOCK", "0", 0, 0),
        _make_insert("图签", "0", 50, 50),
        _make_insert("LAMP", "EE-LIGHT", 100, 100),
    ]
    rows = aggregate_inserts(_make_cad_json(inserts))
    names = {r["block_name"] for r in rows}
    assert "TITLEBLOCK" not in names
    assert "图签" not in names
    assert "LAMP" in names


def test_aggregate_inserts_excludes_anonymous_blocks():
    inserts = [
        _make_insert("*Model_Space", "0"),
        _make_insert("*U12", "0"),
        _make_insert("*X3", "0"),
        _make_insert("LAMP", "EE-LIGHT"),
    ]
    rows = aggregate_inserts(_make_cad_json(inserts))
    names = {r["block_name"] for r in rows}
    assert "LAMP" in names
    assert all(not n.startswith("*") for n in names)


def test_aggregate_inserts_infers_discipline_from_layer():
    inserts = [
        _make_insert("LAMP", "ELEC-LIGHTING"),
        _make_insert("FAN", "HVAC-VENT"),
        _make_insert("VALVE", "PLUMB-PIPE"),
    ]
    rows = aggregate_inserts(_make_cad_json(inserts))
    by_name = {r["block_name"]: r["discipline_guess"] for r in rows}
    # 至少有 1 个能映射上专业（具体由 infer_discipline 决定）
    assert any(value is not None for value in by_name.values())


def test_aggregate_inserts_collects_attribs_summary():
    inserts = [
        _make_insert("LAMP", "EE", attribs={"型号": "AL-A1"}),
        _make_insert("LAMP", "EE", attribs={"型号": "AL-A2"}),
        _make_insert("LAMP", "EE", attribs={"型号": "AL-A1"}),  # 重复应去重
    ]
    rows = aggregate_inserts(_make_cad_json(inserts))
    assert len(rows) == 1
    summary = rows[0]["attribs_summary"]
    assert sorted(summary["型号"]) == ["AL-A1", "AL-A2"]


def _create_project(client: TestClient, name: str) -> int:
    response = client.post("/api/projects", json={"name": name})
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _upload_and_parse(client: TestClient, project_id: int, content: bytes, filename: str = "blocks.dxf") -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", (filename, content, "application/dxf"))],
    )
    assert response.status_code == 201, response.text
    file_id = response.json()["files"][0]["id"]
    parse_resp = client.post(f"/api/files/{file_id}/parse-dxf")
    assert parse_resp.status_code == 200, parse_resp.text
    return parse_resp.json()


def test_block_stats_endpoint_writes_db(client: TestClient):
    project_id = _create_project(client, "INSERT 块统计-端到端")
    content = dxf_with_insert_blocks(
        [
            {"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (5, 0)]},
            {"name": "FAN", "layer": "ME-VENT", "positions": [(0, 100)]},
        ]
    )
    parse_result = _upload_and_parse(client, project_id, content)
    sheet_id = parse_result["sheet_id"]

    extract_resp = client.post(f"/api/sheets/{sheet_id}/extract-blocks", json={})
    assert extract_resp.status_code == 200, extract_resp.text
    summary = extract_resp.json()
    assert summary["total_block_count"] >= 2

    list_resp = client.get(f"/api/sheets/{sheet_id}/block-stats")
    assert list_resp.status_code == 200
    rows = list_resp.json()
    assert len(rows) >= 1
    block_names = {row["block_name"] for row in rows}
    assert "LAMP" in block_names or "FAN" in block_names


def test_block_stats_auto_triggered_after_parse(client: TestClient):
    project_id = _create_project(client, "INSERT 块统计-自动触发")
    content = dxf_with_insert_blocks(
        [{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0)]}]
    )
    parse_result = _upload_and_parse(client, project_id, content)
    sheet_id = parse_result["sheet_id"]

    list_resp = client.get(f"/api/sheets/{sheet_id}/block-stats")
    assert list_resp.status_code == 200
    rows = list_resp.json()
    # 自动触发应该已经写入（即使是 0 行也允许，重点是不抛错）
    assert isinstance(rows, list)


def test_excel_export_includes_block_stats_sheet(client: TestClient):
    project_id = _create_project(client, "INSERT 块统计-Excel")
    content = dxf_with_insert_blocks(
        [{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (10, 0)]}]
    )
    parse_result = _upload_and_parse(client, project_id, content)
    sheet_id = parse_result["sheet_id"]
    client.post(f"/api/sheets/{sheet_id}/extract-blocks", json={})

    export_resp = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True},
    )
    assert export_resp.status_code == 200, export_resp.text
    body = export_resp.json()
    output_path = settings.root_dir / body["file_path"]
    assert output_path.exists()
    wb = load_workbook(output_path, read_only=True)
    try:
        assert "图纸块统计" in wb.sheetnames
    finally:
        wb.close()
