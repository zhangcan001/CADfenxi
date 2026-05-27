"""v1.3 深度抽取能力真实项目回归测试。"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from types import ModuleType
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.services import cad_table_service, block_stats_service
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    dxf_with_insert_blocks,
    equipment_schedule_dxf,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from recognizer.cad_engine.block_aggregator import aggregate_inserts
from recognizer.cad_engine.table_extractor import extract_acad_tables, extract_text_cluster_tables
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf
from test_project_backup_restore import backup_project, create_project, upload_files
from test_v11_fast_ux import export_excel


VERSION = "v1.4-fast-real-project-trial"


def post_import(client: TestClient, project_id: int, files: list[tuple[str, bytes, str]]) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/imports",
        files=[("files", (name, content, mime_type)) for name, content, mime_type in files],
    )
    assert response.status_code == 201, response.text
    return response.json()


def upload_and_parse_dxf(
    client: TestClient,
    project_id: int,
    content: bytes,
    *,
    filename: str = "deep.dxf",
) -> tuple[dict, dict]:
    upload = post_import(client, project_id, [(filename, content, "application/dxf")])
    file_id = upload["files"][0]["id"]
    parse = client.post(f"/api/files/{file_id}/parse-dxf")
    assert parse.status_code == 200, parse.text
    assert parse.json()["status"] == "success"
    return upload, parse.json()


def text_item(text: str, x: float, y: float, *, height: float = 2.5, layer: str = "TABLE") -> dict:
    return {"clean_text": text, "insert": [x, y, 0], "height": height, "layer": layer}


def mtext_item(text: str, x: float, y: float, *, char_height: float = 2.5, layer: str = "TABLE") -> dict:
    return {"clean_text": text, "insert": [x, y, 0], "char_height": char_height, "layer": layer}


def make_cad_json(*, texts: list[dict] | None = None, mtexts: list[dict] | None = None, inserts: list[dict] | None = None) -> dict:
    return {
        "spaces": [
            {
                "space": "modelspace",
                "texts": texts or [],
                "mtexts": mtexts or [],
                "inserts": inserts or [],
            }
        ],
        "counts": {},
        "warnings": [],
    }


def add_confirmed_sheet(
    project_id: int,
    batch_id: int,
    file_id: int,
    *,
    drawing_no: str | None,
    drawing_name: str,
    discipline: str | None = "建筑",
    version: str | None = "A",
    issue_date: date | None = None,
    trust_level: str = "A",
) -> int:
    with SessionLocal() as db:
        page_no = db.query(DrawingSheet).filter(DrawingSheet.file_id == file_id).count() + 1
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=page_no,
            sheet_type="unknown",
            drawing_no=drawing_no,
            drawing_name=drawing_name,
            discipline=discipline,
            version=version,
            issue_date=issue_date,
            status="confirmed",
            review_status="confirmed",
            trust_level=trust_level,
            confidence_score=95.0,
        )
        db.add(sheet)
        db.flush()
        sheet_id = sheet.id
        db.commit()
    return sheet_id


def sheet_state(sheet_id: int) -> tuple[str, str]:
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        assert sheet is not None
        return sheet.status, sheet.review_status


def issue_statuses(project_id: int) -> dict[int, str]:
    with SessionLocal() as db:
        rows = db.scalars(select(DrawingIssue).where(DrawingIssue.project_id == project_id)).all()
        return {row.id: row.status for row in rows}


def test_v13_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v13_acad_table_extractor_path_does_not_regress(monkeypatch: pytest.MonkeyPatch):
    class FakeModelspace:
        def query(self, query: str) -> list[object]:
            assert query == "ACAD_TABLE"
            return [SimpleNamespace(dxf=SimpleNamespace(layer="TABLE", insert=(10, 20, 0)))]

    class FakeDocument:
        def modelspace(self) -> FakeModelspace:
            return FakeModelspace()

    def fake_reader(_entity: object) -> list[list[str]]:
        return [["材料名", "规格", "数量"], ["钢管", "DN25", "10"], ["阀门", "DN32", "2"]]

    fake_module = ModuleType("ezdxf.entities.acad_table")
    fake_module.read_acad_table_content = fake_reader
    monkeypatch.setitem(sys.modules, "ezdxf.entities.acad_table", fake_module)
    tables = extract_acad_tables(FakeDocument())

    assert len(tables) == 1
    assert tables[0]["extraction_method"] == "acad_table"
    assert tables[0]["header"] == ["材料名", "规格", "数量"]
    assert tables[0]["row_count"] == 2
    assert tables[0]["col_count"] == 3


def test_v13_text_and_mtext_cluster_table_extracts_without_polluting_title_fields():
    rows = [
        ["设备名", "型号", "数量"],
        ["送风机", "FJ-1", "2"],
        ["排风机", "FJ-2", "3"],
    ]
    texts: list[dict] = []
    for r_idx, row in enumerate(rows):
        y = 200.0 - r_idx * 8.0
        for c_idx, cell in enumerate(row):
            texts.append(text_item(cell, c_idx * 60.0, y))
    texts.extend(
        [
            text_item("建施-101", 500.0, 20.0, layer="TITLE"),
            text_item("首层平面图", 560.0, 20.0, layer="TITLE"),
        ]
    )
    cad_json = make_cad_json(texts=texts, mtexts=[mtext_item("备注\n备用", 180.0, 200.0)])
    tables = extract_text_cluster_tables(cad_json, exclude_bbox=[490.0, 0.0, 700.0, 80.0])

    assert len(tables) == 1
    table = tables[0]
    assert table["extraction_method"] == "text_cluster"
    assert table["row_count"] >= 2
    assert table["col_count"] >= 3
    flat = json.dumps(table, ensure_ascii=False)
    assert "设备名" in flat and "送风机" in flat
    assert "建施-101" not in flat


def test_v13_empty_table_and_plain_notes_are_not_misdetected():
    empty_like = make_cad_json(
        texts=[
            text_item("A", 0, 100),
            text_item("B", 60, 100),
            text_item("C", 0, 92),
            text_item("D", 60, 92),
        ]
    )
    notes = make_cad_json(
        texts=[
            text_item("本图尺寸以毫米计", 0, 100),
            text_item("施工前应复核现场条件", 0, 92),
            text_item("管线综合需与各专业协调", 0, 84),
            text_item("材料表另见说明", 0, 76),
            text_item("如有疑问请联系设计", 0, 68),
            text_item("不得擅自修改图纸", 0, 60),
        ]
    )

    assert extract_text_cluster_tables(empty_like) == []
    assert extract_text_cluster_tables(notes) == []


def test_v13_table_and_block_endpoints_and_excel_sheets_do_not_pollute_ledger():
    content = equipment_schedule_dxf(
        rows=[["送风机", "FJ-1", "2", "备用"], ["排风机", "FJ-2", "3", ""], ["水泵", "P-1", "1", ""]],
        header=["设备名", "型号", "数量", "备注"],
    )
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3 深度抽取端到端")
        _upload, parse = upload_and_parse_dxf(client, project_id, content, filename="equipment-table.dxf")
        sheet_id = parse["sheet_id"]
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-v13-01", "drawing_name": "设备表回归", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3"})
        tables = client.get(f"/api/sheets/{sheet_id}/tables")
        project_tables = client.get(f"/api/projects/{project_id}/tables")
        blocks = client.get(f"/api/sheets/{sheet_id}/block-stats")
        before_state = sheet_state(sheet_id)
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        after_state = sheet_state(sheet_id)

    assert update.status_code == 200
    assert confirm.status_code == 200
    assert tables.status_code == 200
    assert len(tables.json()) >= 1
    assert project_tables.status_code == 200
    assert len(project_tables.json()) >= 1
    assert blocks.status_code == 200
    assert isinstance(blocks.json(), list)
    assert before_state == ("confirmed", "confirmed")
    assert after_state == before_state
    assert export.status_code == 200, export.text
    assert export.json()["ledger_row_count"] == 1

    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        assert workbook.sheetnames == [
            "图纸总台账",
            "问题清单",
            "专业汇总",
            "图纸表格明细",
            "图纸块统计",
            "校核状态汇总",
            "导出说明",
        ]
        assert workbook["图纸总台账"].max_row == 2
        assert workbook["图纸表格明细"].max_row >= 2
        assert workbook["图纸总台账"]["D2"].value == "建施-v13-01"
        assert workbook["图纸总台账"]["E2"].value == "设备表回归"
    finally:
        workbook.close()


def test_v13_insert_block_stats_filters_low_value_blocks_and_exports_sheet():
    cad_json = make_cad_json(
        inserts=[
            {"block_name": "LAMP", "layer": "EE-LIGHT", "insert": [0, 0, 0], "attribs": []},
            {"block_name": "LAMP", "layer": "EE-LIGHT", "insert": [10, 0, 0], "attribs": []},
            {"block_name": "FAN", "layer": "ME-VENT", "insert": [20, 0, 0], "attribs": []},
            {"block_name": "TITLEBLOCK", "layer": "0", "insert": [30, 0, 0], "attribs": []},
            {"block_name": "图签", "layer": "0", "insert": [40, 0, 0], "attribs": []},
            {"block_name": "*U12", "layer": "0", "insert": [50, 0, 0], "attribs": []},
        ]
    )
    rows = aggregate_inserts(cad_json)
    names = {row["block_name"] for row in rows}

    assert {("LAMP", 2), ("FAN", 1)} <= {(row["block_name"], row["count"]) for row in rows}
    assert "TITLEBLOCK" not in names
    assert "图签" not in names
    assert all(not name.startswith("*") for name in names)

    content = dxf_with_insert_blocks(
        [
            {"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (10, 0)]},
            {"name": "FAN", "layer": "ME-VENT", "positions": [(30, 0)]},
            {"name": "TITLEBLOCK", "layer": "0", "positions": [(0, 50)]},
            {"name": "图签", "layer": "0", "positions": [(10, 50)]},
        ]
    )
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3 块统计")
        _upload, parse = upload_and_parse_dxf(client, project_id, content, filename="blocks.dxf")
        sheet_id = parse["sheet_id"]
        extract = client.post(f"/api/sheets/{sheet_id}/extract-blocks", json={"force": True})
        rows_resp = client.get(f"/api/sheets/{sheet_id}/block-stats")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert extract.status_code == 200
    assert extract.json()["total_block_count"] >= 3
    assert rows_resp.status_code == 200
    stats_names = {row["block_name"] for row in rows_resp.json()}
    assert "LAMP" in stats_names and "FAN" in stats_names
    assert "TITLEBLOCK" not in stats_names and "图签" not in stats_names
    assert export.status_code == 200
    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        assert "图纸块统计" in workbook.sheetnames
        assert workbook["图纸块统计"].max_row >= 2
        assert workbook["图纸总台账"].max_row == 2
    finally:
        workbook.close()


def test_v13_consistency_check_warnings_enter_issue_list_and_excel_without_blocking_export():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3 跨图一致性")
        upload = post_import(client, project_id, [("placeholder.pdf", make_pdf_bytes("v13"), "application/pdf")])
        batch_id = upload["id"]
        file_id = upload["files"][0]["id"]
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="建施-900", drawing_name="首层平面图", version="A")
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="建施-900", drawing_name="二层平面图", version="C")
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no=None, drawing_name="无图号图纸", version="A")
        consistency = client.post(f"/api/projects/{project_id}/consistency-check")
        issues = client.get(f"/api/projects/{project_id}/issues?status=open&page_size=100")
        before_statuses = issue_statuses(project_id)
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        after_statuses = issue_statuses(project_id)

    assert consistency.status_code == 200
    by_code = consistency.json()["by_code"]
    assert by_code.get("CROSS_DRAWING_NO_DUPLICATE", 0) >= 2
    assert by_code.get("CROSS_DRAWING_NAME_CONFLICT", 0) >= 2
    assert by_code.get("CROSS_VERSION_SKIP", 0) >= 2
    assert issues.status_code == 200
    issue_items = issues.json()["items"] if isinstance(issues.json(), dict) and "items" in issues.json() else issues.json()
    cross_codes = {item["issue_code"] for item in issue_items if item["issue_code"].startswith("CROSS_")}
    assert "CROSS_DRAWING_NO_DUPLICATE" in cross_codes
    assert "CROSS_DRAWING_NAME_CONFLICT" in cross_codes
    assert "CROSS_VERSION_SKIP" in cross_codes
    assert before_statuses == after_statuses
    assert export.status_code == 200, export.text
    assert export.json()["ledger_row_count"] == 3
    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        assert "问题清单" in workbook.sheetnames
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        assert "CROSS_DRAWING_NO_DUPLICATE" in issue_text or "重复" in issue_text
        assert workbook["图纸总台账"].max_row == 4
    finally:
        workbook.close()


def test_v13_deep_extract_failures_do_not_block_cad_parse_fusion_or_excel(monkeypatch: pytest.MonkeyPatch):
    def fail_tables(*_args, **_kwargs):
        raise RuntimeError("table extraction failed")

    def fail_blocks(*_args, **_kwargs):
        raise RuntimeError("block stats failed")

    monkeypatch.setattr(cad_table_service, "extract_tables_from_sheet", fail_tables)
    monkeypatch.setattr(block_stats_service, "extract_block_stats_from_sheet", fail_blocks)

    with TestClient(app) as client:
        project_id = create_project(client, "v1.3 失败不阻断")
        _upload, parse = upload_and_parse_dxf(
            client,
            project_id,
            title_block_dxf("建施-v13-fail", "失败不阻断"),
            filename="failure-isolated.dxf",
        )
        sheet_id = parse["sheet_id"]
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-v13-fail", "drawing_name": "失败不阻断", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3 failure"})
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert parse["status"] == "success"
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert export.status_code == 200
    assert export.json()["ledger_row_count"] == 1


def test_v13_pdf_dxf_dwg_cad_pipeline_preview_review_backup_health_regression(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        pdf_project_id = create_project(client, "v1.3 PDF 回归")
        pdf_upload = upload_files(client, pdf_project_id, [("v13.pdf", make_pdf_bytes("图号 建施-v13"), "application/pdf")])
        pdf_split = client.post(f"/api/files/{pdf_upload['files'][0]['id']}/split")

        project_id = create_project(client, "v1.3 CAD 闭环回归")
        cad_upload = upload_files(
            client,
            project_id,
            [
                ("v13-pipeline.dxf", dxf_with_insert_blocks([{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0)]}]), "application/dxf"),
                ("v13-pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        pipeline = run_cad_pipeline_blocking(
            client,
            cad_upload["id"],
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        first_sheet_id = sheets[0]["id"]
        update = client.patch(
            f"/api/sheets/{first_sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-v13-reg", "drawing_name": "深度回归", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{first_sheet_id}/confirm", json={"force": True, "note": "v1.3"})
        preview = client.post(f"/api/sheets/{first_sheet_id}/cad-preview")
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        system_health = client.get("/api/system/health-check")

    assert pdf_split.status_code == 200
    assert len(pdf_split.json()["sheets"]) >= 1
    assert pipeline["summary"]["dwg_files"] == 1
    assert pipeline["summary"]["converted_success"] == 1
    assert pipeline["summary"]["parse_success"] >= 2
    assert len(sheets) >= 2
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert preview.status_code == 200
    assert export["ledger_row_count"] >= 1
    assert backup["backup_id"] > 0
    assert project_health.status_code == 200
    assert system_health.status_code == 200
