"""v1.4.2 真实项目 Excel 交付质量优化测试。"""
from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from dwg_test_helpers import DWG_BYTES, clear_converter_tables, create_converter_setting, write_mock_converter
from test_project_backup_restore import backup_project
from test_v13_deep_extract_review import make_cad_json, text_item
from test_v132_excel_safety import test_v132_pdf_dxf_dwg_cad_preview_backup_health_do_not_regress
from test_v141_real_project_fix import test_v141_real_project_import_stats_next_actions_and_unsupported_do_not_regress
from test_v14_real_project_trial_guard import _run_realistic_trial_project


VERSION = "v1.5.1-fast-delivery-package-fix"
EXPECTED_SHEETS = [
    "图纸总台账",
    "问题清单",
    "专业汇总",
    "校核状态汇总",
    "图纸表格明细",
    "图纸块统计",
    "导出说明",
]


def _status_snapshot(project_id: int) -> tuple[dict[int, str], dict[int, str]]:
    with SessionLocal() as db:
        sheets = db.scalars(select(DrawingSheet).where(DrawingSheet.project_id == project_id)).all()
        issues = db.scalars(select(DrawingIssue).where(DrawingIssue.project_id == project_id)).all()
        return (
            {sheet.id: sheet.review_status for sheet in sheets},
            {issue.id: issue.status for issue in issues},
        )


def _headers(workbook, sheet_name: str) -> list[str]:
    return [cell.value for cell in workbook[sheet_name][1]]


def test_v142_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v142_excel_delivery_order_headers_styles_and_safe_details(tmp_path: Path):
    with TestClient(app) as client:
        result = _run_realistic_trial_project(client, tmp_path)
        project_id = result["project_id"]
        before = _status_snapshot(project_id)
        export_again = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        after = _status_snapshot(project_id)

    assert export_again.status_code == 200
    assert before == after

    workbook = load_workbook(settings.root_dir / export_again.json()["file_path"], read_only=False)
    try:
        assert workbook.sheetnames == EXPECTED_SHEETS
        for worksheet in workbook.worksheets:
            assert worksheet.freeze_panes == "A2"
            assert worksheet.auto_filter.ref is not None

        ledger = workbook["图纸总台账"]
        ledger_headers = _headers(workbook, "图纸总台账")
        assert ledger.max_row - 1 == len(result["sheets"])
        assert "表头(JSON)" not in ledger_headers
        assert "数据(JSON)" not in ledger_headers
        assert "块名" not in ledger_headers
        assert "关键属性" not in ledger_headers
        ledger_text = "\n".join(str(cell) for row in ledger.iter_rows(values_only=True) for cell in row if cell)
        assert "A-101" in ledger_text
        assert "总平面图" in ledger_text

        issue_headers = _headers(workbook, "问题清单")
        assert {"问题级别", "问题类型", "图纸编号", "图纸名称", "问题描述", "建议处理"}.issubset(set(issue_headers))
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        assert "同图号图名不一致" in issue_text
        assert "warning" not in issue_text.lower()

        table_headers = _headers(workbook, "图纸表格明细")
        assert {
            "Sheet ID",
            "图号",
            "图名",
            "表格序号",
            "行号",
            "列号",
            "单元格内容",
            "可信度",
            "低可信标记",
        }.issubset(set(table_headers))
        table_text = "\n".join(str(cell) for row in workbook["图纸表格明细"].iter_rows(values_only=True) for cell in row if cell)
        assert "送风机" in table_text

        block_headers = _headers(workbook, "图纸块统计")
        assert {
            "Sheet ID",
            "图号",
            "图名",
            "块名",
            "归一化块名",
            "数量",
            "是否过滤",
            "过滤原因",
        }.issubset(set(block_headers))
        block_text = "\n".join(str(cell) for row in workbook["图纸块统计"].iter_rows(values_only=True) for cell in row if cell)
        assert "LAMP" in block_text
        assert "已在抽取阶段过滤" in block_text
        assert "A1_FRAME" not in block_text

        info_text = "\n".join(str(cell) for row in workbook["导出说明"].iter_rows(values_only=True) for cell in row if cell)
        assert "本 Excel 由系统自动生成" in info_text
        assert "图纸总台账以人工确认字段优先" in info_text
        assert "图纸表格明细仅作为辅助参考" in info_text
        assert "图纸块统计仅作为辅助参考" in info_text
        assert "CAD 预览仅作为辅助查看" in info_text
        assert "系统不直接解析 DWG" in info_text
        assert "最终成果应由人工复核确认" in info_text
    finally:
        workbook.close()


def test_v142_low_confidence_table_marker_and_block_filter_reason(tmp_path: Path):
    with TestClient(app) as client:
        result = _run_realistic_trial_project(client, tmp_path)
        project_id = result["project_id"]
        with SessionLocal() as db:
            sheet = db.scalar(select(DrawingSheet).where(DrawingSheet.project_id == project_id).order_by(DrawingSheet.id.asc()))
            assert sheet is not None
            from backend.models.drawing_table import DrawingTable

            table = db.scalar(select(DrawingTable).where(DrawingTable.project_id == project_id))
            assert table is not None
            table.warnings_json = '["LOW_CONFIDENCE_TABLE"]'
            db.commit()
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert export.status_code == 200
    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        table_text = "\n".join(str(cell) for row in workbook["图纸表格明细"].iter_rows(values_only=True) for cell in row if cell)
        block_text = "\n".join(str(cell) for row in workbook["图纸块统计"].iter_rows(values_only=True) for cell in row if cell)
        assert "低可信" in table_text
        assert "是" in table_text
        assert "过滤原因" in block_text
        assert "图框块、标题栏块、匿名块" in block_text
    finally:
        workbook.close()


def test_v142_import_and_core_flows_do_not_regress(tmp_path: Path):
    test_v141_real_project_import_stats_next_actions_and_unsupported_do_not_regress(tmp_path)
    test_v132_pdf_dxf_dwg_cad_preview_backup_health_do_not_regress(tmp_path)
