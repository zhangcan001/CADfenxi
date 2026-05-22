from datetime import date
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue


def create_project_file(client: TestClient):
    project = client.post("/api/projects", json={"name": f"Excel一致性项目-{uuid4().hex[:8]}"})
    assert project.status_code == 201
    project_id = project.json()["id"]
    upload = client.post(
        f"/api/projects/{project_id}/imports",
        data={"batch_name": "Excel 一致性批次"},
        files=[("files", ("consistency.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"))],
    )
    assert upload.status_code == 201
    return project_id, upload.json()["id"], upload.json()["files"][0]["id"]


def add_sheet(project_id: int, batch_id: int, file_id: int) -> int:
    with SessionLocal() as db:
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=1,
            sheet_type="unknown",
            drawing_no="机器-01",
            drawing_name="机器图名",
            discipline="建筑",
            version="A",
            issue_date=date(2026, 5, 21),
            status="need_review",
            review_status="unreviewed",
            trust_level="D",
            confidence_score=42,
        )
        db.add(sheet)
        db.flush()
        sheet_id = sheet.id
        db.add(
            FieldValue(
                project_id=project_id,
                batch_id=batch_id,
                file_id=file_id,
                sheet_id=sheet_id,
                field_name="drawing_no",
                raw_value="机器-01",
                normalized_value="机器-01",
                display_value="机器-01",
                final_source="pdf_text",
                confidence=42,
                is_reviewed=False,
            )
        )
        db.add(
            DrawingIssue(
                project_id=project_id,
                batch_id=batch_id,
                file_id=file_id,
                sheet_id=sheet_id,
                issue_code="LOW_CONFIDENCE_NEED_REVIEW",
                severity="warning",
                message="低可信",
                suggestion="请复核",
                status="open",
            )
        )
        db.commit()
    return sheet_id


def export_excel(client: TestClient, project_id: int):
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200
    return response.json(), load_workbook(settings.root_dir / response.json()["file_path"])


def snapshot(sheet_id: int):
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        issue = db.scalar(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id))
        fields = [
            (value.field_name, value.display_value, value.is_reviewed)
            for value in db.scalars(select(FieldValue).where(FieldValue.sheet_id == sheet_id)).all()
        ]
        return {
            "status": sheet.status,
            "review_status": sheet.review_status,
            "issue_status": issue.status,
            "fields": fields,
        }


def test_excel_export_uses_manual_values_and_does_not_mutate_state():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project_file(client)
        sheet_id = add_sheet(project_id, batch_id, file_id)
        update = client.post(
            f"/api/review/sheets/{sheet_id}/update-fields",
            json={"fields": {"drawing_no": "人工-01", "drawing_name": "人工确认图名"}, "note": "导出一致性"},
        )
        assert update.status_code == 200
        before = snapshot(sheet_id)
        result, workbook = export_excel(client, project_id)
        after = snapshot(sheet_id)

    ledger = workbook["图纸总台账"]
    assert result["ledger_row_count"] == 1
    assert ledger.cell(row=2, column=4).value == "人工-01"
    assert ledger.cell(row=2, column=5).value == "人工确认图名"
    assert ledger.cell(row=2, column=11).value == "人工确认"
    assert before == after


def test_export_check_reports_warnings_without_blocking():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project_file(client)
        sheet_id = add_sheet(project_id, batch_id, file_id)
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            sheet.drawing_no = None
            sheet.drawing_name = None
            db.commit()
        check = client.post(f"/api/projects/{project_id}/exports/check")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": False, "include_issues": True, "filter": None},
        )

    assert check.status_code == 200
    data = check.json()
    assert data["can_export"] is True
    assert data["empty_drawing_no_count"] == 1
    assert data["empty_drawing_name_count"] == 1
    assert data["trust_level_d_count"] == 1
    assert data["warning_count"] >= 3
    assert export.status_code == 200


def test_export_empty_project_is_blocked_clearly():
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": f"空导出-{uuid4().hex[:8]}"}).json()
        check = client.post(f"/api/projects/{project['id']}/exports/check")
        export = client.post(
            f"/api/projects/{project['id']}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert check.status_code == 200
    assert check.json()["can_export"] is False
    assert export.status_code == 409
