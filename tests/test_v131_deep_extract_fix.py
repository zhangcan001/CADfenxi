"""v1.3.1 深度抽取真实问题快速修复测试。"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.drawing_table import DrawingTable
from backend.services import block_stats_service, cad_table_service
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    dxf_with_insert_blocks,
    equipment_schedule_dxf,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from recognizer.cad_engine.block_aggregator import aggregate_inserts
from recognizer.cad_engine.table_extractor import extract_text_cluster_tables
from recognizer.cad_engine.cad_json_writer import cad_parse_output_path, write_cad_json
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf
from test_project_backup_restore import backup_project, create_project, upload_files
from test_v11_fast_ux import export_excel
from test_v13_deep_extract_review import (
    add_confirmed_sheet,
    make_cad_json,
    post_import,
    text_item,
    upload_and_parse_dxf,
)


VERSION = "v1.4-fast-real-project-trial"


def test_v131_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v131_plain_notes_and_single_column_candidates_are_filtered_or_downgraded():
    notes = make_cad_json(
        texts=[
            text_item("本图尺寸以毫米计，施工前应复核现场条件", 0, 120),
            text_item("管线综合需与各专业协调，详见设计说明", 0, 112),
            text_item("材料表另见说明，不得擅自修改图纸", 0, 104),
            text_item("如有疑问请联系设计，按规范执行", 0, 96),
            text_item("现场复核后方可施工", 0, 88),
            text_item("本图未尽事宜详见总说明", 0, 80),
        ]
    )
    one_col = make_cad_json(
        texts=[
            text_item("设备名", 0, 120),
            text_item("送风机", 0, 112),
            text_item("排风机", 0, 104),
            text_item("水泵", 0, 96),
            text_item("阀门", 0, 88),
            text_item("风口", 0, 80),
        ]
    )

    assert extract_text_cluster_tables(notes) == []
    assert extract_text_cluster_tables(one_col) == []


def test_v131_title_fields_are_not_overwritten_by_table_detail():
    content = equipment_schedule_dxf(
        rows=[["建施-999", "目录表内图号", "1"], ["建施-998", "目录表内图号2", "2"]],
        header=["图号", "图名", "数量"],
    )
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.1 表格不污染标题栏字段")
        _upload, parse = upload_and_parse_dxf(client, project_id, content, filename="table-title-isolated.dxf")
        sheet_id = parse["sheet_id"]
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-131", "drawing_name": "人工确认图名", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3.1"})
        tables = client.get(f"/api/sheets/{sheet_id}/tables")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert update.status_code == 200
    assert confirm.status_code == 200
    assert tables.status_code == 200
    assert len(tables.json()) >= 1
    assert export.status_code == 200
    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        ledger = workbook["图纸总台账"]
        assert ledger["D2"].value == "人工-131"
        assert ledger["E2"].value == "人工确认图名"
        table_text = "\n".join(str(cell) for row in workbook["图纸表格明细"].iter_rows(values_only=True) for cell in row if cell)
        assert "建施-999" in table_text
    finally:
        workbook.close()


def test_v131_table_extract_failure_warns_and_does_not_block_excel(monkeypatch: pytest.MonkeyPatch):
    def fail_tables(*_args, **_kwargs):
        raise RuntimeError("forced table failure")

    monkeypatch.setattr(cad_table_service, "extract_tables_from_sheet", fail_tables)
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.1 表格失败不阻断")
        _upload, parse = upload_and_parse_dxf(client, project_id, title_block_dxf("建施-131A", "表格失败"))
        sheet_id = parse["sheet_id"]
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-131A", "drawing_name": "表格失败", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3.1"})
        issues = client.get(f"/api/projects/{project_id}/issues?issue_code=CAD_TABLE_EXTRACT_WARNING")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert parse["status"] == "success"
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert issues.status_code == 200
    issue_items = issues.json()["items"] if isinstance(issues.json(), dict) else issues.json()
    assert len(issue_items) >= 1
    assert export.status_code == 200
    assert export.json()["ledger_row_count"] == 1


def test_v131_frame_title_and_anonymous_blocks_are_filtered():
    cad_json = make_cad_json(
        inserts=[
            {"block_name": "LAMP", "layer": "EE-LIGHT", "insert": [0, 0, 0], "attribs": []},
            {"block_name": "A1_FRAME", "layer": "0", "insert": [10, 0, 0], "attribs": []},
            {"block_name": "图框_BLOCK", "layer": "0", "insert": [20, 0, 0], "attribs": []},
            {"block_name": "TITLE_BLOCK", "layer": "0", "insert": [30, 0, 0], "attribs": []},
            {"block_name": "标题栏", "layer": "0", "insert": [40, 0, 0], "attribs": []},
            {"block_name": "*U123", "layer": "0", "insert": [50, 0, 0], "attribs": []},
        ]
    )
    rows = aggregate_inserts(cad_json)
    names = {row["block_name"] for row in rows}

    assert names == {"LAMP"}
    assert rows[0]["count"] == 1


def test_v131_block_stats_failure_warns_and_does_not_block_excel(monkeypatch: pytest.MonkeyPatch):
    def fail_blocks(*_args, **_kwargs):
        raise RuntimeError("forced block failure")

    monkeypatch.setattr(block_stats_service, "extract_block_stats_from_sheet", fail_blocks)
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.1 块统计失败不阻断")
        _upload, parse = upload_and_parse_dxf(client, project_id, title_block_dxf("建施-131B", "块失败"))
        sheet_id = parse["sheet_id"]
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-131B", "drawing_name": "块失败", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3.1"})
        issues = client.get(f"/api/projects/{project_id}/issues?issue_code=CAD_BLOCK_STATS_WARNING")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert update.status_code == 200
    assert confirm.status_code == 200
    assert issues.status_code == 200
    issue_items = issues.json()["items"] if isinstance(issues.json(), dict) else issues.json()
    assert len(issue_items) >= 1
    assert export.status_code == 200


def test_v131_empty_and_low_quality_drawing_no_do_not_enter_duplicate_check():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.1 跨图误报控制")
        upload = post_import(client, project_id, [("placeholder.pdf", make_pdf_bytes("v131"), "application/pdf")])
        batch_id = upload["id"]
        file_id = upload["files"][0]["id"]
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no=None, drawing_name="无图号 A")
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="", drawing_name="无图号 B")
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="A0", drawing_name="纸张幅面")
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="TEMP", drawing_name="临时图号")
        response = client.post(f"/api/projects/{project_id}/consistency-check")

    assert response.status_code == 200
    assert response.json()["by_code"].get("CROSS_DRAWING_NO_DUPLICATE", 0) == 0


def test_v131_same_drawing_no_different_name_warns_once_per_run_without_issue_pileup():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.1 图名冲突不堆积")
        upload = post_import(client, project_id, [("placeholder.pdf", make_pdf_bytes("v131"), "application/pdf")])
        batch_id = upload["id"]
        file_id = upload["files"][0]["id"]
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="建施-131", drawing_name="一层平面图", version="A")
        add_confirmed_sheet(project_id, batch_id, file_id, drawing_no="建施-131", drawing_name="二层平面图", version="A")
        first = client.post(f"/api/projects/{project_id}/consistency-check")
        second = client.post(f"/api/projects/{project_id}/consistency-check")
        issues = client.get(f"/api/projects/{project_id}/issues?issue_code=CROSS_DRAWING_NAME_CONFLICT&status=open&page_size=100")

    assert first.status_code == 200
    assert second.status_code == 200
    assert first.json()["by_code"].get("CROSS_DRAWING_NAME_CONFLICT") == 2
    assert second.json()["by_code"].get("CROSS_DRAWING_NAME_CONFLICT") == 2
    issue_items = issues.json()["items"] if isinstance(issues.json(), dict) else issues.json()
    assert len(issue_items) == 2
    assert {item["severity"] for item in issue_items} == {"warning"}


def test_v131_excel_detail_sheets_are_clear_and_do_not_pollute_ledger():
    content = dxf_with_insert_blocks(
        [
            {"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (10, 0)]},
            {"name": "A1_FRAME", "layer": "0", "positions": [(0, 50)]},
            {"name": "TITLE_BLOCK", "layer": "0", "positions": [(10, 50)]},
        ]
    )
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.1 Excel 明细结构")
        _upload, parse = upload_and_parse_dxf(client, project_id, content, filename="excel-detail.dxf")
        sheet_id = parse["sheet_id"]
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            assert sheet is not None
            output_path = cad_parse_output_path(settings.root_dir, sheet.project_id, sheet.id)
            cad_json = make_cad_json(
                texts=[
                    text_item("设备名", 0, 100),
                    text_item("型号", 60, 100),
                    text_item("数量", 120, 100),
                    text_item("灯具", 0, 92),
                    text_item("L-1", 60, 92),
                    text_item("2", 120, 92),
                    text_item("风机", 0, 84),
                    text_item("F-1", 60, 84),
                    text_item("1", 120, 84),
                ],
                inserts=[
                    {"block_name": "LAMP", "layer": "EE-LIGHT", "insert": [0, 0, 0], "attribs": []},
                    {"block_name": "A1_FRAME", "layer": "0", "insert": [0, 50, 0], "attribs": []},
                    {"block_name": "TITLE_BLOCK", "layer": "0", "insert": [10, 50, 0], "attribs": []},
                ],
            )
            write_cad_json(cad_json, output_path)
            db.commit()
        client.post(f"/api/sheets/{sheet_id}/extract-tables", json={"force": True})
        client.post(f"/api/sheets/{sheet_id}/extract-blocks", json={"force": True})
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-131C", "drawing_name": "Excel 明细", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3.1"})
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert update.status_code == 200
    assert confirm.status_code == 200
    assert export.status_code == 200
    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        ledger_headers = [cell.value for cell in workbook["图纸总台账"][1]]
        assert "块名" not in ledger_headers
        assert "表头(JSON)" not in ledger_headers
        assert workbook["图纸总台账"]["D2"].value == "建施-131C"
        table_headers = [cell.value for cell in workbook["图纸表格明细"][1]]
        block_headers = [cell.value for cell in workbook["图纸块统计"][1]]
        assert table_headers == [
            "序号",
            "Sheet ID",
            "原始文件名",
            "专业",
            "图号",
            "图名",
            "表格类型",
            "表格序号",
            "抽取方式",
            "可信度",
            "行号",
            "行数",
            "列数",
            "表头(JSON)",
            "数据(JSON)",
            "抽取提示",
        ]
        assert block_headers == [
            "序号",
            "Sheet ID",
            "原始文件名",
            "专业",
            "图号",
            "图名",
            "块名",
            "图层",
            "推断专业",
            "数量",
            "关键属性",
        ]
        block_text = "\n".join(str(cell) for row in workbook["图纸块统计"].iter_rows(values_only=True) for cell in row if cell)
        assert "LAMP" in block_text
        assert "A1_FRAME" not in block_text
        assert "TITLE_BLOCK" not in block_text
    finally:
        workbook.close()


def test_v131_pdf_dxf_dwg_cad_preview_excel_backup_health_do_not_regress(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        pdf_project_id = create_project(client, "v1.3.1 PDF 回归")
        pdf_upload = upload_files(client, pdf_project_id, [("v131.pdf", make_pdf_bytes("图号 建施-131"), "application/pdf")])
        pdf_split = client.post(f"/api/files/{pdf_upload['files'][0]['id']}/split")

        project_id = create_project(client, "v1.3.1 核心闭环")
        cad_upload = upload_files(
            client,
            project_id,
            [
                ("v131-pipeline.dxf", dxf_with_insert_blocks([{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0)]}]), "application/dxf"),
                ("v131-pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        pipeline = run_cad_pipeline_blocking(
            client,
            cad_upload["id"],
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        first_sheet_id = sheets[0]["id"]
        update = client.patch(
            f"/api/sheets/{first_sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-131R", "drawing_name": "深度修复回归", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{first_sheet_id}/confirm", json={"force": True, "note": "v1.3.1"})
        preview = client.post(f"/api/sheets/{first_sheet_id}/cad-preview")
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        system_health = client.get("/api/system/health-check")

    assert pdf_split.status_code == 200
    assert len(pdf_split.json()["sheets"]) >= 1
    assert pipeline["summary"]["dwg_files"] == 1
    assert pipeline["summary"]["converted_success"] == 1
    assert pipeline["summary"]["parse_success"] >= 2
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert preview.status_code == 200
    assert export["ledger_row_count"] >= 1
    assert backup["backup_id"] > 0
    assert project_health.status_code == 200
    assert system_health.status_code == 200
