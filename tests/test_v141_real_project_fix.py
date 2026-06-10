"""v1.4.1 真实项目问题快速修复守护测试。"""
from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_block_stat import DrawingBlockStat
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.drawing_table import DrawingTable
from dwg_test_helpers import DWG_BYTES, clear_converter_tables, create_converter_setting, write_mock_converter
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf
from test_project_backup_restore import backup_project, create_project
from test_v123_fast_import_stable import post_import
from test_v13_deep_extract_review import add_confirmed_sheet
from test_v131_deep_extract_fix import (
    test_v131_empty_and_low_quality_drawing_no_do_not_enter_duplicate_check,
    test_v131_excel_detail_sheets_are_clear_and_do_not_pollute_ledger,
    test_v131_frame_title_and_anonymous_blocks_are_filtered,
    test_v131_plain_notes_and_single_column_candidates_are_filtered_or_downgraded,
    test_v131_same_drawing_no_different_name_warns_once_per_run_without_issue_pileup,
    test_v131_title_fields_are_not_overwritten_by_table_detail,
)
from test_v14_real_project_trial_guard import _run_realistic_trial_project


VERSION = "v1.5.1-fast-delivery-package-fix"


def _count(model: type, project_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(select(func.count()).select_from(model).where(model.project_id == project_id)) or 0


def _paths_for_project(project_id: int) -> dict[str, list[str]]:
    with SessionLocal() as db:
        files = db.scalars(select(DrawingFile).where(DrawingFile.project_id == project_id)).all()
        sheets = db.scalars(select(DrawingSheet).where(DrawingSheet.project_id == project_id)).all()
        return {
            "storage": [item.storage_path for item in files if item.storage_path],
            "converted": [item.converted_file_path for item in files if item.converted_file_path],
            "cad_preview": [item.cad_preview_path for item in sheets if item.cad_preview_path],
            "preview": [item.preview_path for item in sheets if item.preview_path],
        }


def test_v141_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v141_real_project_import_stats_next_actions_and_unsupported_do_not_regress(tmp_path: Path):
    clear_converter_tables()
    with TestClient(app) as client:
        no_converter_project = create_project(client, "v1.4.1 DWG 工具提示")
        no_converter = post_import(client, no_converter_project, [("missing-tool.dwg", DWG_BYTES, "application/acad")])

    assert no_converter["file_type_counts"] == {"pdf": 0, "dxf": 0, "dwg": 1, "unsupported": 0}
    assert no_converter["next_actions"] == ["configure_dwg_converter"]
    assert "dwg_converter_not_configured" in no_converter["warnings"]
    assert "run_cad_pipeline" not in no_converter["next_actions"]

    clear_converter_tables()
    with TestClient(app) as client:
        create_converter_setting(client, write_mock_converter(tmp_path))
        project_id = create_project(client, "v1.4.1 混合导入")
        duplicate_pdf = make_pdf_bytes("A-141")
        first = post_import(
            client,
            project_id,
            [
                ("mix.pdf", duplicate_pdf, "application/pdf"),
                ("mix.dxf", title_block_dxf("A-141", "混合导入"), "application/dxf"),
                ("mix.dwg", DWG_BYTES, "application/acad"),
                ("note.jpg", b"jpg", "image/jpeg"),
                ("archive.zip", b"zip", "application/zip"),
            ],
        )
        second = post_import(
            client,
            project_id,
            [
                ("mix-duplicate.pdf", duplicate_pdf, "application/pdf"),
                ("mix-new.dxf", title_block_dxf("A-142", "新 DXF"), "application/dxf"),
            ],
        )
        summary = client.get(f"/api/projects/{project_id}/workbench-summary")
        files = client.get(f"/api/projects/{project_id}/files")

    assert first["total_selected"] == 5
    assert first["imported_count"] == 3
    assert first["unsupported_count"] == 2
    assert first["file_type_counts"] == {"pdf": 1, "dxf": 1, "dwg": 1, "unsupported": 2}
    assert first["next_actions"] == ["split_pdf", "convert_dwg", "run_cad_pipeline"]
    assert "unsupported_files_rejected" in first["warnings"]
    assert {item["file_name"] for item in first["items"] if item["status"] == "unsupported"} == {"note.jpg", "archive.zip"}
    assert second["duplicate_count"] == 1
    assert [item["status"] for item in second["items"]].count("duplicate") == 1
    assert summary.status_code == 200
    assert summary.json()["drawing_file_count"] == 5
    assert files.status_code == 200
    assert {item["source_format"] for item in files.json()} <= {"pdf", "dxf", "dwg"}


def test_v141_deep_extract_false_positive_block_filter_and_consistency_issue_guards():
    test_v131_plain_notes_and_single_column_candidates_are_filtered_or_downgraded()
    test_v131_title_fields_are_not_overwritten_by_table_detail()
    test_v131_frame_title_and_anonymous_blocks_are_filtered()
    test_v131_empty_and_low_quality_drawing_no_do_not_enter_duplicate_check()
    test_v131_same_drawing_no_different_name_warns_once_per_run_without_issue_pileup()


def test_v141_excel_delivery_uses_manual_values_and_keeps_deep_extract_isolated():
    test_v131_excel_detail_sheets_are_clear_and_do_not_pollute_ledger()


def test_v141_real_project_full_flow_excel_backup_restore_paths_and_health(tmp_path: Path):
    with TestClient(app) as client:
        result = _run_realistic_trial_project(client, tmp_path)
        project_id = result["project_id"]
        restored_project_id = result["restored_project_id"]
        with SessionLocal() as db:
            before_issue_status = {
                row.id: row.status
                for row in db.scalars(select(DrawingIssue).where(DrawingIssue.project_id == project_id)).all()
            }
        export_again = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        system_health = client.get("/api/system/health-check")
        temp_file = settings.temp_dir / "v141_cleanup.tmp"
        temp_file.parent.mkdir(parents=True, exist_ok=True)
        temp_file.write_text("temp", encoding="utf-8")
        project_marker_file = settings.projects_dir / f"project_{project_id}" / "keep_after_cleanup.tmp"
        project_marker_file.write_text("project", encoding="utf-8")
        cleanup = client.post("/api/system/cleanup-temp")
        backup = backup_project(client, project_id)
        delete_backup = client.delete(f"/api/backups/{backup['backup_id']}")

    assert result["upload"]["file_type_counts"] == {"pdf": 1, "dxf": 2, "dwg": 1, "unsupported": 0}
    assert result["pipeline"]["status"] == "success"
    assert result["pipeline"]["summary"]["converted_success"] == 1
    assert result["pipeline"]["summary"]["parse_success"] >= 3
    assert _count(DrawingSheet, project_id) == _count(DrawingSheet, restored_project_id)
    assert _count(DrawingTable, project_id) >= 1
    assert _count(DrawingBlockStat, project_id) >= 1
    assert result["restore"].status_code == 200
    assert export_again.status_code == 200
    assert project_health.status_code == 200
    assert system_health.status_code == 200
    assert cleanup.status_code == 200
    assert not temp_file.exists()
    assert project_marker_file.exists()
    project_marker_file.unlink()
    assert delete_backup.status_code == 204

    original_paths = _paths_for_project(project_id)
    restored_paths = _paths_for_project(restored_project_id)
    original_marker = f"project_{project_id}"
    restored_marker = f"project_{restored_project_id}"
    assert all(original_marker in path for group in original_paths.values() for path in group)
    assert all(restored_marker in path and original_marker not in path for group in restored_paths.values() for path in group)

    workbook = load_workbook(settings.root_dir / result["export"]["file_path"], read_only=True)
    try:
        ledger = workbook["图纸总台账"]
        ledger_headers = [cell.value for cell in ledger[1]]
        assert ledger.max_row - 1 == _count(DrawingSheet, project_id)
        assert "表头(JSON)" not in ledger_headers
        assert "数据(JSON)" not in ledger_headers
        assert "块名" not in ledger_headers
        assert "数量" not in ledger_headers
        ledger_text = "\n".join(str(cell) for row in ledger.iter_rows(values_only=True) for cell in row if cell)
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        assert "A-101" in ledger_text
        assert "同图号图名不一致" in issue_text
    finally:
        workbook.close()

    with SessionLocal() as db:
        after_issue_status = {
            row.id: row.status
            for row in db.scalars(select(DrawingIssue).where(DrawingIssue.project_id == project_id)).all()
        }
        sheets = db.scalars(select(DrawingSheet).where(DrawingSheet.project_id == project_id)).all()
        assert all(sheet.review_status == "confirmed" for sheet in sheets)
        assert before_issue_status == after_issue_status
