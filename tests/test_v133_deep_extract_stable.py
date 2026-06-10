"""v1.3.3 深度抽取稳定包整理测试。"""
from __future__ import annotations

from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.core.config import settings
from backend.main import app
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    dxf_with_insert_blocks,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from scripts.build_portable_package import build_portable_package, package_name
from test_full_flow_stability_v055 import make_pdf_bytes
from test_project_backup_restore import backup_project, create_project, upload_files
from test_v11_fast_ux import export_excel
from test_v132_excel_safety import test_v132_excel_deep_extract_sheets_are_isolated_and_safe
from test_v13_deep_extract_review import (
    add_confirmed_sheet,
    make_cad_json,
    post_import,
    text_item,
    upload_and_parse_dxf,
)


VERSION = "v1.5.1-fast-delivery-package-fix"


def test_v133_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION


def test_v133_deep_extract_excel_safety_regression():
    test_v132_excel_deep_extract_sheets_are_isolated_and_safe()


def test_v133_table_block_consistency_and_import_flow_do_not_regress():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.3.3 深度抽取稳定")
        import_result = post_import(
            client,
            project_id,
            [
                (
                    "stable-table.dxf",
                    dxf_with_insert_blocks(
                        [
                            {"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0), (10, 0)]},
                            {"name": "A1_FRAME", "layer": "0", "positions": [(0, 50)]},
                        ]
                    ),
                    "application/dxf",
                )
            ],
        )
        file_id = import_result["files"][0]["id"]
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        tables = client.post(f"/api/sheets/{sheet_id}/extract-tables", json={"force": True})
        blocks = client.post(f"/api/sheets/{sheet_id}/extract-blocks", json={"force": True})
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-133", "drawing_name": "稳定包台账", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": True, "note": "v1.3.3"})
        upload = post_import(client, project_id, [("placeholder.pdf", make_pdf_bytes("v133"), "application/pdf")])
        add_confirmed_sheet(project_id, upload["id"], upload["files"][0]["id"], drawing_no="建施-133", drawing_name="稳定包冲突", version="A")
        consistency = client.post(f"/api/projects/{project_id}/consistency-check")
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert import_result["file_type_counts"]["dxf"] == 1
    assert parse.status_code == 200
    assert parse.json()["status"] == "success"
    assert tables.status_code == 200
    assert blocks.status_code == 200
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert consistency.status_code == 200
    assert consistency.json()["by_code"].get("CROSS_DRAWING_NAME_CONFLICT", 0) >= 2
    assert export.status_code == 200

    workbook = load_workbook(settings.root_dir / export.json()["file_path"], read_only=True)
    try:
        assert "图纸表格明细" in workbook.sheetnames
        assert "图纸块统计" in workbook.sheetnames
        ledger_headers = [cell.value for cell in workbook["图纸总台账"][1]]
        assert "表头(JSON)" not in ledger_headers
        assert "块名" not in ledger_headers
        issue_text = "\n".join(str(cell) for row in workbook["问题清单"].iter_rows(values_only=True) for cell in row if cell)
        assert "同图号图名不一致" in issue_text
    finally:
        workbook.close()


def test_v133_pdf_dxf_dwg_pipeline_preview_review_excel_backup_health_regression(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        pdf_project_id = create_project(client, "v1.3.3 PDF 回归")
        pdf_upload = upload_files(client, pdf_project_id, [("v133.pdf", make_pdf_bytes("图号 建施-133"), "application/pdf")])
        pdf_split = client.post(f"/api/files/{pdf_upload['files'][0]['id']}/split")

        project_id = create_project(client, "v1.3.3 核心闭环")
        cad_upload = upload_files(
            client,
            project_id,
            [
                ("v133-pipeline.dxf", dxf_with_insert_blocks([{"name": "LAMP", "layer": "EE-LIGHT", "positions": [(0, 0)]}]), "application/dxf"),
                ("v133-pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        pipeline = run_cad_pipeline_blocking(
            client,
            cad_upload["id"],
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        first_sheet_id = sheets[0]["id"]
        update = client.patch(
            f"/api/sheets/{first_sheet_id}/fields",
            json={"fields": {"drawing_no": "建施-133R", "drawing_name": "稳定包回归", "discipline": "建筑"}},
        )
        confirm = client.post(f"/api/sheets/{first_sheet_id}/confirm", json={"force": True, "note": "v1.3.3"})
        preview = client.post(f"/api/sheets/{first_sheet_id}/cad-preview")
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        system_health = client.get("/api/system/health-check")

    assert pdf_split.status_code == 200
    assert len(pdf_split.json()["sheets"]) >= 1
    assert pipeline["summary"]["dwg_files"] == 1
    assert pipeline["summary"]["converted_success"] == 1
    assert pipeline["summary"]["parse_success"] >= 2
    assert update.status_code == 200
    assert confirm.status_code == 200
    assert preview.status_code == 200
    assert export["ledger_row_count"] >= 1
    assert backup["backup_id"] > 0
    assert project_health.status_code == 200
    assert system_health.status_code == 200


def test_v133_portable_package_can_be_generated_and_contains_expected_files():
    summary = build_portable_package(version=VERSION, clean=True, skip_tests=True)
    package_dir = settings.root_dir / "release" / package_name(VERSION)
    package_info = summary.package_info_path.read_text(encoding="utf-8")

    required_paths = [
        "backend/main.py",
        "recognizer",
        "frontend/dist/index.html",
        "scripts/local_launcher.py",
        "app_data/projects",
        "docs/FAST_RELEASE_REPORT_v1.3.3.md",
        "requirements.txt",
        "README.md",
        "README_本地使用说明.md",
        "RELEASE_NOTES.md",
        "package_info.txt",
        "start.bat",
        "stop.bat",
        "check_env.bat",
    ]
    forbidden_parts = {".git", "node_modules", "__pycache__", ".pytest_cache", "coverage", "htmlcov"}

    assert summary.integrity_ok is True
    assert summary.package_dir == package_dir
    assert f"版本：{VERSION}" in package_info
    for relative in required_paths:
        assert (package_dir / relative).exists(), relative
    for path in package_dir.rglob("*"):
        relative_parts = set(path.relative_to(package_dir).parts)
        assert not (relative_parts & forbidden_parts)
        assert path.name not in {".env", ".env.local"}
        assert path.suffix not in {".log", ".tmp", ".pyc"}
