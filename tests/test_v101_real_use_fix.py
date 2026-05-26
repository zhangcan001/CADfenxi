from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from dwg_test_helpers import (
    DWG_BYTES,
    clear_converter_tables,
    create_converter_setting,
    run_cad_pipeline_blocking,
    write_mock_converter,
)
from scripts.build_portable_package import DEFAULT_VERSION, build_portable_package, package_name
from test_full_flow_stability_v055 import make_pdf_bytes, title_block_dxf
from test_project_backup_restore import backup_project, create_project, upload_files
from test_recognition_raw import _wait_for_ocr_job


VERSION = "v1.1.5-deep-extract"
ROOT = Path(__file__).resolve().parents[1]


def export_excel(client: TestClient, project_id: int) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/exports/excel",
        json={"confirm_incomplete": True, "include_issues": True, "filter": None},
    )
    assert response.status_code == 200, response.text
    return response.json()


def sheet_snapshot(sheet_id: int) -> dict:
    with SessionLocal() as db:
        sheet = db.get(DrawingSheet, sheet_id)
        assert sheet is not None
        values = sorted(
            (value.field_name, value.display_value, value.is_reviewed, value.final_source)
            for value in db.scalars(select(FieldValue).where(FieldValue.sheet_id == sheet_id)).all()
        )
        issues = sorted(
            (issue.issue_code, issue.severity, issue.status)
            for issue in db.scalars(select(DrawingIssue).where(DrawingIssue.sheet_id == sheet_id)).all()
        )
        return {
            "status": sheet.status,
            "review_status": sheet.review_status,
            "drawing_no": sheet.drawing_no,
            "drawing_name": sheet.drawing_name,
            "discipline": sheet.discipline,
            "values": values,
            "issues": issues,
        }


def field_value(sheet_id: int, field_name: str) -> FieldValue:
    with SessionLocal() as db:
        value = db.scalar(
            select(FieldValue).where(FieldValue.sheet_id == sheet_id, FieldValue.field_name == field_name)
        )
        assert value is not None
        db.expunge(value)
        return value


def candidate_count(sheet_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(
            select(func.count()).select_from(RecognitionCandidate).where(RecognitionCandidate.sheet_id == sheet_id)
        ) or 0


def open_issue_count(sheet_id: int) -> int:
    with SessionLocal() as db:
        return db.scalar(
            select(func.count())
            .select_from(DrawingIssue)
            .where(DrawingIssue.sheet_id == sheet_id, DrawingIssue.status == "open")
        ) or 0


def test_v101_health_version_and_portable_package_output():
    with TestClient(app) as client:
        health = client.get("/api/health")

    summary = build_portable_package(ROOT, version=VERSION, clean=True)
    package_dir = ROOT / "release" / package_name(VERSION)
    package_info = (package_dir / "package_info.txt").read_text(encoding="utf-8")

    assert health.status_code == 200
    assert health.json() == {"status": "ok", "version": VERSION, "database": "ok", "storage": "ok"}
    assert settings.version == VERSION
    assert DEFAULT_VERSION == VERSION
    assert summary.package_dir == package_dir
    assert summary.integrity_ok is True
    assert package_dir.name == f"工程图纸智能台账识别系统-{VERSION}"
    assert VERSION in package_info
    assert "包类型：Windows 本地便携正式稳定版" in package_info
    assert (package_dir / "app_data" / "backups" / ".gitkeep").is_file()
    assert (ROOT / "docs" / "V1_0_1_REAL_USE_FIX_REPORT.md").is_file()


def test_v101_pdf_minimal_flow_does_not_regress():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.1 PDF 最小流程")
        upload = upload_files(client, project_id, [("v101.pdf", make_pdf_bytes("图号 建施-101"), "application/pdf")])
        batch_id = upload["id"]
        file_id = upload["files"][0]["id"]
        split = client.post(f"/api/files/{file_id}/split")
        sheet_id = split.json()["sheets"][0]["id"]
        title_crop = client.post(f"/api/imports/{batch_id}/title-crops")
        extract_text = client.post(f"/api/imports/{batch_id}/extract-text")
        ocr = client.post(f"/api/imports/{batch_id}/ocr-titles")
        _wait_for_ocr_job(client, batch_id)
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "PDF-人工-101", "drawing_name": "PDF 人工图名", "discipline": "建筑"}},
        )
        export = export_excel(client, project_id)
        preview = client.get(f"/api/sheets/{sheet_id}/preview")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert split.status_code == 200
    assert title_crop.status_code == 200
    assert extract_text.status_code == 200
    assert ocr.status_code == 200
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert update.status_code == 200
    assert preview.status_code == 200
    assert workbook["图纸总台账"].cell(2, 4).value == "PDF-人工-101"
    assert export["ledger_row_count"] == 1


def test_v101_dxf_minimal_flow_does_not_regress():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.1 DXF 最小流程")
        upload = upload_files(client, project_id, [("v101.dxf", title_block_dxf("建施-101", "DXF 回归"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        prepare = client.post(f"/api/files/{file_id}/prepare-dxf-sheet")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        summary = client.get(f"/api/sheets/{sheet_id}/cad-parse")
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)

    assert prepare.status_code == 200
    assert parse.status_code == 200
    assert parse.json()["status"] == "success"
    assert (settings.root_dir / parse.json()["output_path"]).is_file()
    assert summary.status_code == 200
    assert summary.json()["counts"]["text_count"] + summary.json()["counts"]["mtext_count"] + summary.json()["counts"]["attrib_count"] > 0
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert candidate_count(sheet_id) > 0
    assert export["ledger_row_count"] == 1


def test_v101_dwg_mock_conversion_flow_does_not_regress(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.1 DWG mock 转换")
        upload = upload_files(client, project_id, [("v101.dwg", DWG_BYTES, "application/acad")])
        file_id = upload["files"][0]["id"]
        parse_before_convert = client.post(f"/api/files/{file_id}/parse-dxf")
        create_converter_setting(client, converter)
        convert = client.post(f"/api/files/{file_id}/convert-dwg-to-dxf")
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        candidates = client.post(f"/api/sheets/{sheet_id}/generate-candidates")
        fuse = client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        export = export_excel(client, project_id)
        runs = client.get(f"/api/files/{file_id}/cad-conversion-runs")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert parse_before_convert.status_code == 400
    assert parse_before_convert.json()["detail"]["error_code"] == "DWG_NOT_CONVERTED"
    assert convert.status_code == 200
    assert convert.json()["status"] == "success"
    assert (settings.root_dir / convert.json()["converted_file_path"]).is_file()
    assert parse.status_code == 200
    assert candidates.status_code == 200
    assert fuse.status_code == 200
    assert runs.status_code == 200
    assert runs.json()[0]["status"] == "success"
    assert workbook["图纸总台账"].cell(2, 8).value == "DWG转换"


def test_v101_cad_pipeline_idempotent_and_protects_manual_fields(tmp_path: Path):
    clear_converter_tables()
    converter = write_mock_converter(tmp_path)
    with TestClient(app) as client:
        create_converter_setting(client, converter)
        project_id = create_project(client, "v1.0.1 CAD pipeline")
        upload = upload_files(
            client,
            project_id,
            [
                ("pipeline.pdf", make_pdf_bytes("PDF 不应被 CAD pipeline 处理"), "application/pdf"),
                ("pipeline.dxf", title_block_dxf("建施-101P", "Pipeline DXF"), "application/dxf"),
                ("pipeline.dwg", DWG_BYTES, "application/acad"),
            ],
        )
        batch_id = upload["id"]
        pdf_file_id = next(item["id"] for item in upload["files"] if item["source_format"] == "pdf")
        assert client.post(f"/api/files/{pdf_file_id}/split").status_code == 200
        pipeline = run_cad_pipeline_blocking(
            client,
            batch_id,
            {
                "steps": [
                    "convert_dwg",
                    "prepare_dxf_sheet",
                    "parse_dxf",
                    "generate_candidates",
                    "fuse_fields",
                    "generate_cad_preview",
                ],
                "skip_completed": True,
                "continue_on_error": True,
            },
        )
        sheets = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"]
        dxf_sheet_id = next(item["id"] for item in sheets if item["source_format"] == "dxf")
        candidates_before = candidate_count(dxf_sheet_id)
        issues_before = open_issue_count(dxf_sheet_id)
        manual = client.patch(
            f"/api/sheets/{dxf_sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-PIPE-101", "drawing_name": "人工 Pipeline", "discipline": "建筑"}},
        )
        rerun = run_cad_pipeline_blocking(
            client,
            batch_id,
            {"steps": ["generate_candidates", "fuse_fields"], "skip_completed": False, "continue_on_error": True},
        )
        after = client.get(f"/api/sheets/{dxf_sheet_id}").json()
        with SessionLocal() as db:
            pdf_sheet_count = db.scalar(select(func.count()).select_from(DrawingSheet).where(DrawingSheet.file_id == pdf_file_id))

    assert pipeline["summary"]["pdf_files"] == 1
    assert pipeline["summary"]["dwg_files"] == 1
    assert pipeline["summary"]["dxf_files"] == 1
    assert pipeline["summary"]["converted_success"] == 1
    assert pipeline["summary"]["parse_success"] == 2
    assert pipeline["summary"]["cad_preview_success"] >= 1
    assert manual.status_code == 200
    assert rerun["status"] in {"success", "completed_with_errors", "skipped", "failed"}
    assert after["drawing_no"] == "人工-PIPE-101"
    assert field_value(dxf_sheet_id, "drawing_no").is_reviewed is True
    assert candidate_count(dxf_sheet_id) == candidates_before
    assert open_issue_count(dxf_sheet_id) <= issues_before
    assert pdf_sheet_count == 1


def test_v101_review_workbench_batch_confirm_and_export_consistency():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.1 校核导出一致性")
        upload = upload_files(client, project_id, [("review.dxf", title_block_dxf("建施-101R", "校核图"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        parse = client.post(f"/api/files/{file_id}/parse-dxf")
        sheet_id = parse.json()["sheet_id"]
        assert client.post(f"/api/sheets/{sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{sheet_id}/fuse-fields").status_code == 200
        update = client.patch(
            f"/api/sheets/{sheet_id}/fields",
            json={"fields": {"drawing_no": "人工-校核-101", "drawing_name": "人工校核图名", "discipline": "建筑"}},
        )
        with SessionLocal() as db:
            sheet = db.get(DrawingSheet, sheet_id)
            assert sheet is not None
            db.add(
                DrawingIssue(
                    project_id=sheet.project_id,
                    batch_id=sheet.batch_id,
                    file_id=sheet.file_id,
                    sheet_id=sheet.id,
                    issue_code="OPEN_ERROR",
                    severity="error",
                    message="v1.0.1 open error",
                    suggestion="处理后再确认",
                    status="open",
                )
            )
            db.commit()
        blocked = client.post(
            f"/api/projects/{project_id}/batch-confirm",
            json={"sheet_ids": [sheet_id], "note": "不应确认"},
        )
        confirm_without_force = client.post(f"/api/sheets/{sheet_id}/confirm", json={"force": False})
        before = sheet_snapshot(sheet_id)
        export = export_excel(client, project_id)
        after = sheet_snapshot(sheet_id)
        logs = client.get(f"/api/sheets/{sheet_id}/audit-logs")

    workbook = load_workbook(settings.root_dir / export["file_path"])
    assert update.status_code == 200
    assert blocked.status_code == 200
    assert blocked.json()["confirmed_count"] == 0
    assert blocked.json()["skipped"][0]["reason"] in {"存在 open error", "存在未处理错误"}
    assert confirm_without_force.status_code == 400
    assert before == after
    assert workbook["图纸总台账"].cell(2, 4).value == "人工-校核-101"
    assert logs.status_code == 200
    assert any(log["action_type"] == "field_edit" for log in logs.json())


def test_v101_backup_restore_health_and_temp_cleanup_are_safe():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.1 备份恢复健康检查")
        upload = upload_files(client, project_id, [("backup.dxf", title_block_dxf("建施-101B", "备份图"), "application/dxf")])
        file_id = upload["files"][0]["id"]
        assert client.post(f"/api/files/{file_id}/parse-dxf").status_code == 200
        sheet_id = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"][0]["id"]
        assert client.post(f"/api/sheets/{sheet_id}/cad-preview").status_code == 200
        export = export_excel(client, project_id)
        backup = backup_project(client, project_id)

        temp_file = settings.temp_dir / "v101-cleanup.tmp"
        temp_file.parent.mkdir(parents=True, exist_ok=True)
        temp_file.write_text("temp", encoding="utf-8")
        project_dir = settings.projects_dir / f"project_{project_id}"
        backup_path = settings.root_dir / backup["file_path"]
        export_path = settings.root_dir / export["file_path"]

        system_health = client.get("/api/system/health-check")
        project_health = client.get(f"/api/projects/{project_id}/health-check")
        backup_health = client.get("/api/backups/health-check")
        export_health = client.get("/api/exports/health-check")
        cleanup = client.post("/api/system/cleanup-temp")
        backup_exists_before_delete = backup_path.exists()
        restore = client.post(f"/api/backups/{backup['backup_id']}/restore", json={"restore_mode": "new_project"})
        restored_id = restore.json()["new_project_id"]
        delete_backup = client.delete(f"/api/backups/{backup['backup_id']}")
        original_project = client.get(f"/api/projects/{project_id}")
        restored_project = client.get(f"/api/projects/{restored_id}")
        restored_export = export_excel(client, restored_id)

    assert system_health.status_code == 200
    assert project_health.status_code == 200
    assert backup_health.status_code == 200
    assert export_health.status_code == 200
    assert cleanup.status_code == 200
    assert not temp_file.exists()
    assert project_dir.exists()
    assert backup_exists_before_delete is True
    assert not backup_path.exists()
    assert export_path.exists()
    assert restore.status_code == 200
    assert restored_id != project_id
    assert delete_backup.status_code == 204
    assert original_project.status_code == 200
    assert restored_project.status_code == 200
    assert restored_export["ledger_row_count"] == 1
    assert settings.projects_dir.exists()
    assert settings.backups_dir.exists()
    assert (settings.projects_dir / f"project_{restored_id}").exists()


def test_v101_cad_preview_failure_does_not_block_excel_export():
    with TestClient(app) as client:
        project_id = create_project(client, "v1.0.1 CAD 预览失败隔离")
        good = upload_files(client, project_id, [("good.dxf", title_block_dxf("建施-101G", "好图"), "application/dxf")])
        good_file_id = good["files"][0]["id"]
        assert client.post(f"/api/files/{good_file_id}/parse-dxf").status_code == 200
        good_sheet_id = client.get(f"/api/projects/{project_id}/sheets?page_size=100").json()["items"][0]["id"]
        assert client.post(f"/api/sheets/{good_sheet_id}/generate-candidates").status_code == 200
        assert client.post(f"/api/sheets/{good_sheet_id}/fuse-fields").status_code == 200

        bad = upload_files(client, project_id, [("bad.dxf", b"broken dxf", "application/dxf")])
        bad_file_id = bad["files"][0]["id"]
        bad_prepare = client.post(f"/api/files/{bad_file_id}/prepare-dxf-sheet")
        bad_sheet_id = bad_prepare.json()["sheet_id"]
        failed_preview = client.post(f"/api/sheets/{bad_sheet_id}/cad-preview")
        export = export_excel(client, project_id)

    assert bad_prepare.status_code == 200
    assert failed_preview.status_code == 200
    assert failed_preview.json()["status"] == "failed"
    assert export["ledger_row_count"] >= 1
