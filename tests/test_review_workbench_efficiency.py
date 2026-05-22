from datetime import date
from pathlib import Path
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from backend.core.config import settings
from backend.core.database import SessionLocal
from backend.main import app
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.field_value import FieldValue
from backend.models.recognition_candidate import RecognitionCandidate
from backend.models.review_audit_log import ReviewAuditLog


def create_project(client: TestClient, name: str = "校核效率测试项目"):
    project = client.post("/api/projects", json={"name": f"{name}-{uuid4().hex[:8]}"})
    assert project.status_code == 201
    project_id = project.json()["id"]
    upload = client.post(
        f"/api/projects/{project_id}/imports",
        data={"batch_name": "v0.5.2 校核效率批次"},
        files=[("files", ("review.pdf", b"%PDF-1.4\n%%EOF", "application/pdf"))],
    )
    assert upload.status_code == 201
    return project_id, upload.json()["id"], upload.json()["files"][0]["id"]


def add_file(project_id: int, batch_id: int, name: str, source_format: str) -> int:
    with SessionLocal() as db:
        drawing_file = DrawingFile(
            project_id=project_id,
            batch_id=batch_id,
            original_name=name,
            file_ext=f".{source_format}",
            source_format=source_format,
            file_size=10,
            file_hash=uuid4().hex,
            storage_path=f"app_data/temp/{uuid4().hex}_{name}",
            status="preprocessed",
            convert_status="success" if source_format == "dwg" else None,
            converted_format="dxf" if source_format == "dwg" else None,
        )
        db.add(drawing_file)
        db.flush()
        file_id = drawing_file.id
        db.commit()
    return file_id


def add_sheet(
    project_id: int,
    batch_id: int,
    file_id: int,
    *,
    drawing_no: str | None = "建施-01",
    drawing_name: str | None = "首层平面图",
    discipline: str | None = "建筑",
    issue_date: date | None = date(2026, 5, 21),
    status: str = "need_review",
    review_status: str = "unreviewed",
    trust_level: str = "A",
    confidence_score: float = 92,
    error_code: str | None = None,
) -> int:
    with SessionLocal() as db:
        page_no = db.query(DrawingSheet).filter(DrawingSheet.file_id == file_id).count() + 1
        sheet = DrawingSheet(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            page_no=page_no,
            sheet_type="unknown",
            drawing_no=drawing_no,
            drawing_name=drawing_name,
            discipline=discipline,
            version="A",
            issue_date=issue_date,
            status=status,
            review_status=review_status,
            trust_level=trust_level,
            confidence_score=confidence_score,
            error_code=error_code,
        )
        db.add(sheet)
        db.flush()
        sheet_id = sheet.id
        for field_name, value in {
            "drawing_no": drawing_no,
            "drawing_name": drawing_name,
            "discipline": discipline,
            "issue_date": issue_date.isoformat() if issue_date else None,
        }.items():
            if value:
                db.add(
                    FieldValue(
                        project_id=project_id,
                        batch_id=batch_id,
                        file_id=file_id,
                        sheet_id=sheet_id,
                        field_name=field_name,
                        raw_value=value,
                        normalized_value=value,
                        display_value=value,
                        final_source="pdf_text",
                        confidence=confidence_score,
                        is_reviewed=False,
                    )
                )
        db.commit()
    return sheet_id


def add_issue(project_id: int, batch_id: int, file_id: int, sheet_id: int, code: str, severity: str) -> None:
    with SessionLocal() as db:
        db.add(
            DrawingIssue(
                project_id=project_id,
                batch_id=batch_id,
                file_id=file_id,
                sheet_id=sheet_id,
                issue_code=code,
                severity=severity,
                message="测试问题",
                suggestion="请处理",
                status="open",
            )
        )
        db.commit()


def add_candidate(project_id: int, batch_id: int, file_id: int, sheet_id: int, value: str, confidence: int = 95) -> int:
    with SessionLocal() as db:
        candidate = RecognitionCandidate(
            project_id=project_id,
            batch_id=batch_id,
            file_id=file_id,
            sheet_id=sheet_id,
            field_name="drawing_no",
            candidate_value=value,
            normalized_value=value,
            source_type="title_ocr",
            confidence=confidence,
            raw_text=f"图号：{value}",
            parser_name="test",
            parser_version="1.0",
        )
        db.add(candidate)
        db.flush()
        candidate_id = candidate.id
        db.commit()
    return candidate_id


def test_review_filters_status_trust_missing_and_source_format():
    with TestClient(app) as client:
        project_id, batch_id, pdf_file = create_project(client)
        dxf_file = add_file(project_id, batch_id, "cad.dxf", "dxf")
        dwg_file = add_file(project_id, batch_id, "converted.dwg", "dwg")
        unreviewed = add_sheet(project_id, batch_id, pdf_file, trust_level="C", confidence_score=55)
        confirmed = add_sheet(project_id, batch_id, pdf_file, review_status="confirmed", status="confirmed")
        missing_no = add_sheet(project_id, batch_id, dxf_file, drawing_no=None, trust_level="D", confidence_score=35)
        dwg_sheet = add_sheet(project_id, batch_id, dwg_file, drawing_no="结施-02", trust_level="B")

        unreviewed_items = client.get(f"/api/projects/{project_id}/sheets?review_status=unreviewed&page_size=100").json()["items"]
        assert {item["id"] for item in unreviewed_items} == {unreviewed, missing_no, dwg_sheet}
        low = client.get(f"/api/projects/{project_id}/sheets?low_confidence=true&page_size=100").json()["items"]
        assert {item["id"] for item in low} == {unreviewed, missing_no}
        missing = client.get(f"/api/projects/{project_id}/sheets?missing_field=drawing_no&page_size=100").json()["items"]
        assert [item["id"] for item in missing] == [missing_no]
        assert client.get(f"/api/projects/{project_id}/sheets?source_format=pdf&page_size=100").json()["total"] == 2
        assert client.get(f"/api/projects/{project_id}/sheets?source_format=dxf&page_size=100").json()["items"][0]["id"] == missing_no
        assert client.get(f"/api/projects/{project_id}/sheets?source_format=dwg&page_size=100").json()["items"][0]["id"] == dwg_sheet
        assert confirmed


def test_review_sorting_by_issue_count_and_confidence():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project(client)
        quiet = add_sheet(project_id, batch_id, file_id, drawing_no="建施-01", confidence_score=95)
        noisy = add_sheet(project_id, batch_id, file_id, drawing_no="建施-02", confidence_score=45, trust_level="D")
        add_issue(project_id, batch_id, file_id, noisy, "DRAWING_NO_EMPTY", "error")
        add_issue(project_id, batch_id, file_id, noisy, "VERSION_EMPTY", "warning")

        by_issues = client.get(f"/api/projects/{project_id}/sheets?sort_by=issue_count&sort_order=desc&page_size=100").json()["items"]
        by_confidence = client.get(f"/api/projects/{project_id}/sheets?sort_by=confidence_score&sort_order=asc&page_size=100").json()["items"]

    assert by_issues[0]["id"] == noisy
    assert by_confidence[0]["id"] == noisy
    assert by_confidence[-1]["id"] == quiet


def test_batch_confirm_a_level_skips_open_error_and_missing_drawing_no_and_writes_audit():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project(client)
        ok = add_sheet(project_id, batch_id, file_id, drawing_no="建施-01", trust_level="A")
        open_error = add_sheet(project_id, batch_id, file_id, drawing_no="建施-02", trust_level="A")
        missing_no = add_sheet(project_id, batch_id, file_id, drawing_no=None, trust_level="A")
        add_issue(project_id, batch_id, file_id, open_error, "OPEN_ERROR", "error")
        response = client.post(
            "/api/review/batch-confirm",
            json={"project_id": project_id, "sheet_ids": [ok, open_error, missing_no], "confirm_mode": "trust_a"},
        )

    data = response.json()
    assert response.status_code == 200
    assert data["confirmed_count"] == 1
    assert data["skipped_count"] == 2
    reasons = {item["sheet_id"]: item.get("reason") for item in data["items"] if item["status"] == "skipped"}
    assert reasons[open_error] == "存在 open error"
    assert reasons[missing_no] == "drawing_no 缺失"
    with SessionLocal() as db:
        assert db.get(DrawingSheet, ok).review_status == "confirmed"
        assert db.scalar(select(ReviewAuditLog.id).where(ReviewAuditLog.sheet_id == ok, ReviewAuditLog.action_type == "sheet_batch_confirmed")) is not None


def test_candidate_adopt_restore_and_manual_fusion_protection():
    with TestClient(app) as client:
        project_id, batch_id, file_id = create_project(client)
        sheet_id = add_sheet(project_id, batch_id, file_id, drawing_no="建施-01")
        low = add_candidate(project_id, batch_id, file_id, sheet_id, "建施-08", 75)
        add_candidate(project_id, batch_id, file_id, sheet_id, "建施-09", 98)

        adopted = client.post(f"/api/sheets/{sheet_id}/adopt-candidate", json={"candidate_id": low, "note": "采用候选值"})
        restored = client.post(f"/api/sheets/{sheet_id}/restore-recommended", json={"field_name": "drawing_no", "note": "恢复机器推荐值"})
        client.post(f"/api/sheets/{sheet_id}/fuse-fields")
        values = client.get(f"/api/sheets/{sheet_id}/field-values").json()
        logs = client.get(f"/api/sheets/{sheet_id}/audit-logs").json()

    assert adopted.json()["field_value"]["display_value"] == "建施-08"
    assert restored.json()["field_value"]["display_value"] == "建施-09"
    reviewed = [value for value in values if value["field_name"] == "drawing_no" and value["is_reviewed"]]
    assert reviewed[0]["display_value"] == "建施-09"
    assert any(log["action_type"] == "candidate_adopted" for log in logs)
    assert any(log["action_type"] == "recommended_restored" for log in logs)


def test_excel_export_uses_reviewed_values_and_formats_remain_reviewable():
    with TestClient(app) as client:
        project_id, batch_id, pdf_file = create_project(client, "导出一致性")
        dxf_file = add_file(project_id, batch_id, "review.dxf", "dxf")
        dwg_file = add_file(project_id, batch_id, "review.dwg", "dwg")
        pdf_sheet = add_sheet(project_id, batch_id, pdf_file, drawing_no="建施-01", review_status="confirmed", status="confirmed")
        dxf_sheet = add_sheet(project_id, batch_id, dxf_file, drawing_no="电施-01", review_status="confirmed", status="confirmed")
        dwg_sheet = add_sheet(project_id, batch_id, dwg_file, drawing_no="水施-01", review_status="confirmed", status="confirmed")
        update = client.post(
            f"/api/review/sheets/{pdf_sheet}/update-fields",
            json={"fields": {"drawing_no": "人工-01", "drawing_name": "人工确认图名"}, "note": "导出前人工校核"},
        )
        export = client.post(
            f"/api/projects/{project_id}/exports/excel",
            json={"confirm_incomplete": True, "include_issues": True, "filter": None},
        )

    assert update.status_code == 200
    assert export.status_code == 200
    workbook = load_workbook(settings.root_dir / export.json()["file_path"])
    ledger = workbook["图纸总台账"]
    exported_numbers = [ledger.cell(row=row, column=4).value for row in range(2, ledger.max_row + 1)]
    exported_names = [ledger.cell(row=row, column=5).value for row in range(2, ledger.max_row + 1)]
    assert "人工-01" in exported_numbers
    assert "人工确认图名" in exported_names
    assert {pdf_sheet, dxf_sheet, dwg_sheet}


def test_review_efficiency_static_portable_assets_still_exist():
    assert Path("scripts/start_local.bat").exists()
    assert Path("scripts/local_launcher.py").exists()
    assert Path("scripts/build_portable_package.py").exists()
