from collections import Counter, defaultdict
from datetime import datetime
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.core.config import settings
from backend.models.drawing_block_stat import DrawingBlockStat
from backend.models.drawing_file import DrawingFile
from backend.models.drawing_issue import DrawingIssue
from backend.models.drawing_sheet import DrawingSheet
from backend.models.drawing_table import DrawingTable
from backend.models.field_value import FieldValue
from backend.models.project import Project
from backend.models.recognition_candidate import RecognitionCandidate
from backend.models.review_audit_log import ReviewAuditLog
from backend.schemas.export import ExportCheckResult
from backend.services.sheet_query_service import issue_counts_subquery

SYSTEM_NAME = "工程图纸智能台账识别系统"
SYSTEM_VERSION = settings.version
CORE_FIELDS = ["drawing_no", "drawing_name", "discipline", "version", "issue_date"]


def build_excel(
    db: Session,
    project: Project,
    output_path: Path,
    check_result: ExportCheckResult,
) -> tuple[int, int]:
    workbook = Workbook()
    context = export_context(db, project.id)

    ledger_ws = workbook.active
    ledger_ws.title = "图纸总台账"
    write_ledger_sheet(ledger_ws, project, context)

    issue_ws = workbook.create_sheet("问题清单")
    write_issue_sheet(issue_ws, context)

    discipline_ws = workbook.create_sheet("专业汇总")
    write_discipline_summary_sheet(discipline_ws, context)

    review_ws = workbook.create_sheet("校核状态汇总")
    write_review_summary_sheet(review_ws, context)

    table_detail_ws = workbook.create_sheet("图纸表格明细")
    write_table_detail_sheet(table_detail_ws, db, project.id, context)

    block_stats_ws = workbook.create_sheet("图纸块统计")
    write_block_stats_sheet(block_stats_ws, db, project.id, context)

    info_ws = workbook.create_sheet("导出说明")
    write_info_sheet(info_ws, project, check_result, len(context["sheets"]), len(context["issues"]))

    workbook.save(output_path)
    return len(context["sheets"]), len(context["issues"])


def export_context(db: Session, project_id: int) -> dict:
    rows = ledger_rows(db, project_id)
    issues = issue_rows(db, project_id)
    values = db.scalars(
        select(FieldValue)
        .join(DrawingSheet, DrawingSheet.id == FieldValue.sheet_id)
        .where(DrawingSheet.project_id == project_id)
        .order_by(FieldValue.field_name.asc(), FieldValue.is_reviewed.desc(), FieldValue.id.asc())
    ).all()
    candidates = db.scalars(
        select(RecognitionCandidate)
        .join(DrawingSheet, DrawingSheet.id == RecognitionCandidate.sheet_id)
        .where(DrawingSheet.project_id == project_id)
        .order_by(RecognitionCandidate.confidence.desc(), RecognitionCandidate.id.asc())
    ).all()
    logs = db.scalars(
        select(ReviewAuditLog)
        .where(ReviewAuditLog.project_id == project_id)
        .order_by(ReviewAuditLog.created_at.desc(), ReviewAuditLog.id.desc())
    ).all()
    return {
        "sheets": rows,
        "issues": issues,
        "field_values": values_by_sheet(values),
        "candidates": candidates_by_sheet(candidates),
        "confirmed_times": confirmed_times_by_sheet(values, logs),
    }


def ledger_rows(db: Session, project_id: int):
    counts = issue_counts_subquery()
    return db.execute(
        select(
            DrawingSheet,
            DrawingFile,
            counts.c.issue_count,
            counts.c.error_count,
            counts.c.warning_count,
        )
        .join(DrawingFile, DrawingFile.id == DrawingSheet.file_id)
        .outerjoin(counts, counts.c.sheet_id == DrawingSheet.id)
        .where(DrawingSheet.project_id == project_id)
        .order_by(DrawingSheet.discipline.asc(), DrawingSheet.drawing_no.asc(), DrawingSheet.id.asc())
    ).all()


def issue_rows(db: Session, project_id: int):
    return db.execute(
        select(DrawingIssue, DrawingSheet, DrawingFile)
        .join(DrawingSheet, DrawingSheet.id == DrawingIssue.sheet_id)
        .join(DrawingFile, DrawingFile.id == DrawingIssue.file_id)
        .where(DrawingIssue.project_id == project_id)
        .order_by(DrawingIssue.status.asc(), DrawingIssue.severity.asc(), DrawingIssue.id.asc())
    ).all()


def values_by_sheet(values: list[FieldValue]) -> dict[int, dict[str, FieldValue]]:
    grouped: dict[int, dict[str, FieldValue]] = defaultdict(dict)
    for value in values:
        current = grouped[value.sheet_id].get(value.field_name)
        if current is None or (value.is_reviewed and not current.is_reviewed):
            grouped[value.sheet_id][value.field_name] = value
    return grouped


def candidates_by_sheet(candidates: list[RecognitionCandidate]) -> dict[int, dict[str, list[RecognitionCandidate]]]:
    grouped: dict[int, dict[str, list[RecognitionCandidate]]] = defaultdict(lambda: defaultdict(list))
    for candidate in candidates:
        grouped[candidate.sheet_id][candidate.field_name].append(candidate)
    return grouped


def confirmed_times_by_sheet(values: list[FieldValue], logs: list[ReviewAuditLog]) -> dict[int, datetime]:
    confirmed: dict[int, datetime] = {}
    for log in logs:
        if log.action_type in {"sheet_confirmed", "sheet_batch_confirmed"} and log.sheet_id not in confirmed:
            confirmed[log.sheet_id] = log.created_at
    for value in values:
        if value.is_reviewed and value.reviewed_at and value.sheet_id not in confirmed:
            confirmed[value.sheet_id] = value.reviewed_at
    return confirmed


def write_ledger_sheet(ws, project: Project, context: dict) -> None:
    headers = [
        "序号", "项目名称", "专业", "图纸编号", "图纸名称", "版本", "出图日期",
        "文件格式", "原始文件名", "来源文件", "识别来源", "可信等级", "识别评分",
        "校核状态", "确认时间", "问题数量", "错误数量", "警告数量", "备注",
    ]
    ws.append(headers)
    values = context["field_values"]
    confirmed_times = context["confirmed_times"]
    for index, (sheet, drawing_file, issue_count, error_count, warning_count) in enumerate(context["sheets"], start=1):
        row = [
            index,
            project.name,
            sheet.discipline or "",
            sheet.drawing_no or "",
            sheet.drawing_name or "",
            sheet.version or "",
            sheet.issue_date,
            source_format_label(drawing_file.source_format),
            drawing_file.original_name or "",
            source_file_label(drawing_file),
            source_summary(values.get(sheet.id, {})),
            sheet.trust_level or "",
            int(sheet.confidence_score) if sheet.confidence_score is not None else "",
            review_status_label(sheet.review_status),
            confirmed_times.get(sheet.id),
            issue_count or 0,
            error_count or 0,
            warning_count or 0,
            "",
        ]
        ws.append(row)
    style_sheet(ws, date_columns={7}, datetime_columns={15}, number_columns={1, 13, 16, 17, 18})


def write_issue_sheet(ws, context: dict) -> None:
    headers = [
        "序号", "专业", "图纸编号", "图纸名称", "问题级别", "问题类型", "问题描述",
        "建议处理", "字段名", "当前值", "候选值", "来源", "状态", "创建时间",
    ]
    ws.append(headers)
    candidates = context["candidates"]
    for index, (issue, sheet, _drawing_file) in enumerate(context["issues"], start=1):
        field_name = issue_field_name(issue.issue_code)
        candidate_items = candidates.get(sheet.id, {}).get(field_name or "", [])
        ws.append([
            index,
            sheet.discipline or "",
            sheet.drawing_no or "",
            sheet.drawing_name or "",
            severity_label(issue.severity),
            issue_type_label(issue.issue_code),
            issue.message,
            issue.suggestion,
            field_name_label(field_name),
            current_field_value(sheet, field_name),
            " / ".join(candidate_display(candidate) for candidate in candidate_items[:3]),
            " / ".join(source_type_label(candidate.source_type) for candidate in candidate_items[:3]),
            issue_status_label(issue.status),
            issue.created_at,
        ])
    style_sheet(ws, datetime_columns={14}, number_columns={1})


def write_discipline_summary_sheet(ws, context: dict) -> None:
    headers = ["专业", "图纸数量", "已确认数量", "未校核数量", "A 级数量", "B 级数量", "C 级数量", "D 级数量", "错误数量", "警告数量"]
    ws.append(headers)
    summary: dict[str, Counter] = defaultdict(Counter)
    for sheet, _drawing_file, _issue_count, error_count, warning_count in context["sheets"]:
        discipline = sheet.discipline or "未识别"
        summary[discipline]["图纸数量"] += 1
        summary[discipline]["已确认数量"] += 1 if sheet.review_status == "confirmed" else 0
        summary[discipline]["未校核数量"] += 1 if sheet.review_status != "confirmed" else 0
        summary[discipline][f"{sheet.trust_level or 'D'} 级数量"] += 1
        summary[discipline]["错误数量"] += error_count or 0
        summary[discipline]["警告数量"] += warning_count or 0
    for discipline in sorted(summary):
        item = summary[discipline]
        ws.append([discipline] + [item[header] for header in headers[1:]])
    style_sheet(ws, number_columns=set(range(2, len(headers) + 1)))


def write_table_detail_sheet(ws, db: Session, project_id: int, context: dict) -> None:
    """每行 = 一张表的一个非空单元格，保留行级 JSON 供旧版本核对。"""
    headers = [
        "序号",
        "Sheet ID",
        "原始文件名",
        "专业",
        "图号",
        "图名",
        "表格类型",
        "表格序号",
        "抽取方式",
        "可信度",
        "低可信标记",
        "行号",
        "列号",
        "单元格内容",
        "行数",
        "列数",
        "表头(JSON)",
        "数据(JSON)",
        "抽取提示",
    ]
    ws.append(headers)

    sheet_meta: dict[int, tuple[str, str, str, str]] = {}
    for sheet, drawing_file, _issue_count, _err, _warn in context["sheets"]:
        sheet_meta[sheet.id] = (
            drawing_file.original_name or "",
            sheet.discipline or "未识别",
            sheet.drawing_no or "",
            sheet.drawing_name or "",
        )

    tables = db.scalars(
        select(DrawingTable)
        .where(DrawingTable.project_id == project_id)
        .order_by(
            DrawingTable.sheet_id.asc(),
            DrawingTable.table_index.asc(),
            DrawingTable.id.asc(),
        )
    ).all()

    counter = 1
    for table in tables:
        original_name, discipline, drawing_no, drawing_name = sheet_meta.get(
            table.sheet_id, ("", "未识别", "", "")
        )
        header_json = table.header_json or "[]"
        rows = _parse_rows_json(table.rows_json)
        warnings = _json_list_text(table.warnings_json)
        confidence = table_confidence_label(table.warnings_json)
        low_confidence = "是" if confidence == "低可信" else "否"
        if not rows:
            continue
        for row_idx, row in enumerate(rows, start=1):
            row_text = _row_to_text(row)
            for col_idx, cell_text in enumerate(row, start=1):
                if not str(cell_text).strip():
                    continue
                ws.append([
                    counter,
                    table.sheet_id,
                    original_name,
                    discipline,
                    drawing_no,
                    drawing_name,
                    table_kind_label(table.table_kind),
                    table.table_index,
                    extraction_method_label(table.extraction_method),
                    confidence,
                    low_confidence,
                    row_idx,
                    col_idx,
                    str(cell_text),
                    table.row_count,
                    table.col_count,
                    header_json,
                    row_text,
                    warnings,
                ])
                counter += 1

    style_sheet(ws, number_columns={1, 2, 8, 12, 13, 15, 16})


def _parse_rows_json(value: str | None) -> list[list[str]]:
    if not value:
        return []
    try:
        data = json.loads(value)
    except (ValueError, TypeError):
        return []
    if not isinstance(data, list):
        return []
    out: list[list[str]] = []
    for row in data:
        if isinstance(row, list):
            out.append([str(cell) for cell in row])
    return out


def _row_to_text(row: list[str]) -> str:
    return " | ".join(cell for cell in row)


def _json_list_text(value: str | None) -> str:
    if not value:
        return ""
    try:
        data = json.loads(value)
    except (ValueError, TypeError):
        return ""
    if not isinstance(data, list):
        return ""
    return " / ".join(str(item) for item in data if item)


def table_confidence_label(warnings_json: str | None) -> str:
    warnings = set(_json_list(warnings_json))
    return "低可信" if "LOW_CONFIDENCE_TABLE" in warnings else "正常"


def _json_list(value: str | None) -> list[str]:
    if not value:
        return []
    try:
        data = json.loads(value)
    except (ValueError, TypeError):
        return []
    if not isinstance(data, list):
        return []
    return [str(item) for item in data]


def table_kind_label(kind: str) -> str:
    return {
        "equipment": "设备表",
        "material": "材料表",
        "drawing_index": "图纸目录",
        "legend": "图例表",
        "other": "其他",
    }.get(kind, kind)


def extraction_method_label(method: str) -> str:
    return {
        "acad_table": "ACAD 表格实体",
        "text_cluster": "文字坐标聚类",
    }.get(method, method)


def write_block_stats_sheet(ws, db: Session, project_id: int, context: dict) -> None:
    """每行 = 一个 (block_name, layer) 在某 sheet 的辅助统计行。"""
    headers = [
        "序号",
        "Sheet ID",
        "原始文件名",
        "专业",
        "图号",
        "图名",
        "块名",
        "归一化块名",
        "图层",
        "推断专业",
        "数量",
        "是否过滤",
        "过滤原因",
        "关键属性",
    ]
    ws.append(headers)

    sheet_meta: dict[int, tuple[str, str, str, str]] = {}
    for sheet, drawing_file, _issue_count, _err, _warn in context["sheets"]:
        sheet_meta[sheet.id] = (
            drawing_file.original_name or "",
            sheet.discipline or "未识别",
            sheet.drawing_no or "",
            sheet.drawing_name or "",
        )

    stats = db.scalars(
        select(DrawingBlockStat)
        .where(DrawingBlockStat.project_id == project_id)
        .order_by(
            DrawingBlockStat.sheet_id.asc(),
            DrawingBlockStat.count.desc(),
            DrawingBlockStat.id.asc(),
        )
    ).all()

    counter = 1
    for stat in stats:
        original_name, discipline, drawing_no, drawing_name = sheet_meta.get(
            stat.sheet_id, ("", "未识别", "", "")
        )
        ws.append(
            [
                counter,
                stat.sheet_id,
                original_name,
                discipline,
                drawing_no,
                drawing_name,
                stat.block_name,
                normalize_block_name_for_export(stat.block_name),
                stat.layer_name or "",
                stat.discipline_guess or "未识别",
                stat.count,
                "否",
                "已保留为辅助块统计；图框块、标题栏块、匿名块已在抽取阶段过滤",
                _attribs_text(stat.attribs_summary_json),
            ]
        )
        counter += 1

    style_sheet(ws, number_columns={1, 2, 11})


def _attribs_text(value: str | None) -> str:
    if not value:
        return ""
    try:
        data = json.loads(value)
    except (ValueError, TypeError):
        return ""
    if not isinstance(data, dict):
        return ""
    parts: list[str] = []
    for tag, values in data.items():
        if isinstance(values, list):
            parts.append(f"{tag}: {', '.join(str(v) for v in values[:5])}")
    return " | ".join(parts)


def normalize_block_name_for_export(value: str | None) -> str:
    return " ".join((value or "").strip().split())


def write_review_summary_sheet(ws, context: dict) -> None:
    sheets = context["sheets"]
    rows = [
        ("图纸总数", len(sheets), "当前导出范围内图纸数量"),
        ("已确认图纸", count_sheets(sheets, lambda sheet, _file, *_: sheet.review_status == "confirmed"), "校核状态为已确认"),
        ("未校核图纸", count_sheets(sheets, lambda sheet, _file, *_: sheet.review_status != "confirmed"), "校核状态未确认"),
        ("A 级图纸", count_sheets(sheets, lambda sheet, _file, *_: sheet.trust_level == "A"), "可信等级 A"),
        ("B 级图纸", count_sheets(sheets, lambda sheet, _file, *_: sheet.trust_level == "B"), "可信等级 B"),
        ("C 级图纸", count_sheets(sheets, lambda sheet, _file, *_: sheet.trust_level == "C"), "可信等级 C"),
        ("D 级图纸", count_sheets(sheets, lambda sheet, _file, *_: sheet.trust_level == "D"), "可信等级 D"),
        ("存在错误图纸", count_sheets(sheets, lambda _sheet, _file, _issues, errors, _warnings: (errors or 0) > 0), "存在 open error"),
        ("存在警告图纸", count_sheets(sheets, lambda _sheet, _file, _issues, _errors, warnings: (warnings or 0) > 0), "存在 open warning"),
        ("缺图号图纸", count_sheets(sheets, lambda sheet, _file, *_: not sheet.drawing_no), "图纸编号为空"),
        ("缺图名图纸", count_sheets(sheets, lambda sheet, _file, *_: not sheet.drawing_name), "图纸名称为空"),
        ("缺专业图纸", count_sheets(sheets, lambda sheet, _file, *_: not sheet.discipline), "专业为空"),
        ("PDF 图纸", count_sheets(sheets, lambda _sheet, drawing_file, *_: drawing_file.source_format == "pdf"), "来源格式 PDF"),
        ("DXF 图纸", count_sheets(sheets, lambda _sheet, drawing_file, *_: drawing_file.source_format == "dxf"), "来源格式 DXF"),
        ("DWG 转换图纸", count_sheets(sheets, lambda _sheet, drawing_file, *_: drawing_file.source_format == "dwg"), "DWG 转 DXF 后识别"),
    ]
    ws.append(["统计项", "数量", "说明"])
    for row in rows:
        ws.append(list(row))
    style_sheet(ws, number_columns={2})


def write_info_sheet(ws, project: Project, check_result: ExportCheckResult, sheet_count: int, issue_count: int) -> None:
    rows = [
        ("系统名称", SYSTEM_NAME),
        ("系统版本", SYSTEM_VERSION),
        ("项目名称", project.name),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("生成说明", "本 Excel 由系统自动生成。"),
        ("导出范围", "当前项目全部图纸"),
        ("图纸总数", sheet_count),
        ("已确认图纸数量", sheet_count - check_result.unconfirmed_count),
        ("未校核图纸数量", check_result.unconfirmed_count),
        ("问题总数", issue_count),
        ("错误数量", check_result.open_error_count),
        ("警告数量", check_result.open_warning_count),
        ("图纸表格明细数量", check_result.drawing_table_count),
        ("低可信表格数量", check_result.low_confidence_table_count),
        ("块统计图纸数量", check_result.block_stats_sheet_count),
        ("跨图一致性问题数量", check_result.consistency_issue_count),
        ("台账字段说明", "图纸总台账以人工确认字段优先；人工确认值优先于机器识别值。"),
        ("图纸表格明细说明", "图纸表格明细仅作为辅助参考，不会写入或覆盖图纸总台账。"),
        ("图纸块统计说明", "图纸块统计仅作为辅助参考，不作为工程量或造价依据；图框块、标题栏块、匿名块已在抽取阶段过滤。"),
        ("跨图一致性说明", "跨图一致性问题用于辅助发现重复图号、图名冲突、版本异常等问题，已进入问题清单供人工复核。"),
        ("CAD 预览说明", "CAD 预览仅作为辅助查看，不保证与专业 CAD 软件完全一致。"),
        ("DWG 处理说明", "系统不直接解析 DWG，DWG 通过外部工具转换为 DXF 后识别。"),
        ("数据来源说明", "PDF 来源：PDF 文本、标题栏 OCR、文件名、规则推断。DXF 来源：CAD 块属性、CAD 文字、CAD 多行文字、CAD 图层、文件名。"),
        ("重要限制说明", "本系统用于辅助生成图纸台账。低可信图纸、存在错误或警告的图纸，应人工复核。最终成果应由人工复核确认。"),
        ("深度抽取导出说明", "图纸表格明细、图纸块统计和跨图一致性问题均为辅助信息；正式图纸总台账只使用主字段，不会被表格明细或块统计覆盖。"),
        ("台账状态", "完整台账" if check_result.is_complete_ledger else "存在需复核项"),
    ]
    ws.append(["项目", "内容"])
    for row in rows:
        ws.append(list(row))
    style_sheet(ws, number_columns={2})


def count_sheets(rows, predicate) -> int:
    return sum(1 for row in rows if predicate(*row))


def style_sheet(ws, date_columns: set[int] | None = None, datetime_columns: set[int] | None = None, number_columns: set[int] | None = None) -> None:
    date_columns = date_columns or set()
    datetime_columns = datetime_columns or set()
    number_columns = number_columns or set()
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="17212F")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in date_columns:
                cell.number_format = "yyyy-mm-dd"
            if cell.column in datetime_columns:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            if cell.column in number_columns:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    auto_width(ws)
    apply_preferred_widths(ws)


def auto_width(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 42)


def apply_preferred_widths(ws) -> None:
    preferred = {
        "专业": 12,
        "Sheet ID": 10,
        "图纸编号": 18,
        "图纸名称": 32,
        "原始文件名": 36,
        "来源文件": 36,
        "识别来源": 18,
        "问题描述": 40,
        "建议处理": 40,
        "问题类型": 32,
        "候选值": 28,
        "来源": 18,
        "备注": 30,
        "内容": 80,
        "说明": 42,
        "数据(JSON)": 48,
        "表头(JSON)": 48,
        "单元格内容": 32,
        "抽取提示": 32,
        "关键属性": 42,
        "过滤原因": 50,
        "归一化块名": 24,
    }
    for cell in ws[1]:
        width = preferred.get(str(cell.value))
        if width:
            ws.column_dimensions[cell.column_letter].width = width


def source_summary(values: dict[str, FieldValue]) -> str:
    sources = []
    for field in CORE_FIELDS:
        value = values.get(field)
        if value and value.final_source not in sources:
            sources.append(value.final_source)
    if any(values.get(field) and values[field].is_reviewed for field in CORE_FIELDS):
        return "人工确认"
    return " / ".join(source_type_label(source) for source in sources) if sources else ""


def source_file_label(drawing_file: DrawingFile) -> str:
    if drawing_file.source_format == "dwg":
        return drawing_file.converted_file_path or drawing_file.storage_path or "DWG 转换图纸"
    return drawing_file.storage_path or drawing_file.original_name


def candidate_display(candidate: RecognitionCandidate) -> str:
    return candidate.normalized_value or candidate.candidate_value


def issue_field_name(issue_code: str) -> str | None:
    text = issue_code.lower()
    if "drawing_no" in text:
        return "drawing_no"
    if "drawing_name" in text:
        return "drawing_name"
    if "discipline" in text:
        return "discipline"
    if "version" in text:
        return "version"
    if "date" in text:
        return "issue_date"
    return None


def current_field_value(sheet: DrawingSheet, field_name: str | None) -> str:
    if not field_name:
        return ""
    value = getattr(sheet, field_name, None)
    if value is None:
        return ""
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def format_datetime(value) -> str:
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def source_format_label(value: str) -> str:
    return {"pdf": "PDF", "dxf": "DXF", "dwg": "DWG转换"}.get(value, value.upper())


def source_type_label(source_type: str) -> str:
    labels = {
        "cad_block_attr": "CAD 块属性",
        "cad_text": "CAD 文字",
        "cad_mtext": "CAD 多行文字",
        "cad_layer": "CAD 图层",
        "cad_filename": "CAD 文件名",
        "pdf_text": "PDF 文本",
        "title_ocr": "标题栏 OCR",
        "filename": "文件名",
        "rule": "规则推断",
        "mixed": "多来源一致",
        "manual": "人工确认",
    }
    return labels.get(source_type, source_type)


def issue_type_label(issue_code: str) -> str:
    labels = {
        "DRAWING_NO_EMPTY": "图纸编号缺失",
        "DRAWING_NAME_EMPTY": "图纸名称缺失",
        "DISCIPLINE_EMPTY": "专业缺失",
        "VERSION_EMPTY": "版本缺失",
        "ISSUE_DATE_EMPTY": "出图日期缺失",
        "DRAWING_NO_DUPLICATE": "图纸编号重复",
        "DRAWING_NO_SUSPECT": "图纸编号疑似异常",
        "DRAWING_NAME_SUSPECT": "图纸名称疑似异常",
        "CAD_DRAWING_NO_SUSPECT": "CAD 图纸编号疑似异常",
        "CAD_DRAWING_NAME_SUSPECT": "CAD 图纸名称疑似异常",
        "CAD_FIELD_CONFLICT": "CAD 字段候选冲突",
        "FIELD_CONFLICT_HIGH": "字段候选高风险冲突",
        "OCR_TEXT_EMPTY": "OCR 文本为空",
        "PDF_TEXT_EMPTY": "PDF 文本为空",
        "CAD_BLOCK_ATTR_MISSING": "CAD 块属性缺失",
        "CAD_TABLE_EXTRACT_WARNING": "CAD 表格抽取提示",
        "CAD_BLOCK_STATS_WARNING": "CAD 块统计提示",
        "CAD_PARSE_EMPTY_CONTENT": "CAD 解析内容为空",
        "LOW_CONFIDENCE": "低可信",
        "LOW_CONFIDENCE_NEED_REVIEW": "低可信需复核",
        "ONLY_FROM_FILENAME": "仅来自文件名",
        "FIELD_ONLY_FROM_FILENAME": "字段仅来自文件名",
        "OPEN_ERROR": "文件打开错误",
        "DXF_OPEN_FAILED": "DXF 打开失败",
        "CROSS_DRAWING_NO_DUPLICATE": "跨图重复图号",
        "CROSS_DRAWING_NAME_CONFLICT": "同图号图名不一致",
        "CROSS_VERSION_SKIP": "版本号跳号",
        "CROSS_DISCIPLINE_PREFIX_MISMATCH": "专业与图号前缀不一致",
        "CROSS_ISSUE_DATE_INCONSISTENT": "同图号同版本出图日期不一致",
        "CROSS_VERSION_DATE_REGRESS": "新版本出图日期倒退",
    }
    label = labels.get(issue_code, "其他问题")
    return f"{label}（{issue_code}）"


def severity_label(value: str) -> str:
    return {"error": "错误", "warning": "警告", "info": "提示"}.get(value, value)


def issue_status_label(value: str) -> str:
    return {"open": "待处理", "resolved": "已解决", "ignored": "已忽略"}.get(value, value)


def review_status_label(value: str) -> str:
    return "已确认" if value == "confirmed" else "未校核"


def field_name_label(field_name: str | None) -> str:
    labels = {
        "drawing_no": "图纸编号",
        "drawing_name": "图纸名称",
        "discipline": "专业",
        "version": "版本",
        "issue_date": "出图日期",
    }
    return labels.get(field_name or "", "")
